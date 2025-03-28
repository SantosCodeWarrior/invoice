from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

from openpyxl.styles import PatternFill
#####################################################


@csrf_exempt
def monthly_summary(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		login_user 		= request.user
		
		context={	
			'login_user' : login_user,
		}

		return render_to_response("invoice_display/monthly_summary.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')


def load_data(request):
	client_list 	= models.Client.objects.filter(proj_name='BOSS',status=1,currency_type='USD')
	login_user 		= request.user
	start_date  	= '2023-06-30'		
	bill_type   	= 'monthly_flat'
	boss_details    = []	
	s_no = 0
	for c in client_list:
		if c.id==20002:
			c_id = 136
		elif c.id==20011:
			c_id = 135
		else:
			c_id = c.id

		get_date 	 = start_date				
		api_url 	 = "https://aboss.bwesglobal.com/api/get_ship_to_be_billed/"			
		api_method   = "GET"			
		parameters   = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'client_id': c_id,'billing_type':bill_type,'billing_date':(get_date)}
		response     = requests.get(api_url, params=parameters,verify=False)			
		
		try:
			boss_array  = json.loads(response.content)
			for i in boss_array:
				ccv  	= models.api_client_data.objects.filter(client_id=c_id,ship_name=i['Ship Name']).first()
				try:
					acc = ccv.account_tab
				except:
					acc = ''
				boss_details.append({
					'client_name'  	: c.client_name,
					'ship_name'    	: i['Ship Name'],
					'billing'	 	: bill_type,
					'account_tab'	: acc,					
				})
		except:
			pass			


	context={	
		'boss_details' : boss_details,
	}

	return HttpResponse(json.dumps(context))

@csrf_exempt
def export_load_data(request):
	show_table 	  = json.loads(request.POST['show_table'])
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/MASTER LIST.xlsx'
	dstroot       = select_dir+'/MASTER LIST.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/MASTER LIST.xlsx')
	wb 	   	      = load_workbook(my_dir+'/MASTER LIST.xlsx')
	ws_usd     	  = wb.get_sheet_by_name("Sheet1")	
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	ii   		  = 2
	jj 	 		  = 1

	for t in show_table:		
		thin_border    				   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style       				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
		_row_style     				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)		
		ws_usd.cell('A'+str(ii)).value = jj
		ws_usd.cell('B'+str(ii)).value = t['Client']
		ws_usd.cell('C'+str(ii)).value = t['Vessel']
		ws_usd.cell('D'+str(ii)).value = t['Account Tab']		
		ws_usd.cell('A'+str(ii)).style = _row_style	 
		ws_usd.cell('B'+str(ii)).style = _row_style	 
		ws_usd.cell('C'+str(ii)).style = _row_style	 
		ws_usd.cell('D'+str(ii)).style = _row_style	 
		
		ii+=1
		jj+=1	
		
		wb.save(my_dir+'/MASTER LIST.xlsx')
	return HttpResponse(json.dumps('done'))


def download_load_data(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)
	location 	   					= (my_dir+'/MASTER LIST.xlsx')
	filename 	   					= my_dir+'/MASTER LIST.xlsx'
	download_name  					= "MASTER LIST.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response