from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################

def ledger(request):
	if request.user.is_authenticated():		
		login_user 	= request.user		
		QRYS 		= "select id,client_name,proj_name,currency_type,rclient_name from invoice2.it_client where status='1' and client_name<>'' group by client_name;"
		cursor 		= connection.cursor()
		cursor.execute(QRYS)
		client_list = cursor.fetchall()
		clnt_array  = []
		for i in client_list:
			if i[4]!=None:
				client_name = i[1]
			else:
				client_name = i[1]

			clnt_array.append({
				'id' 	  : i[0],
				'clients' : client_name,
				})

		context={
			'login_user' 	: login_user,
			'client_list' 	: clnt_array,
		}
		return render_to_response("invoice_display/ledger.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')


def ledger_datas(request):
	client_ids 		= request.GET['client_ids']
	fin_date 		= request.GET['fin_date']
	split 			= fin_date.split('-')
	years 			= int(split[0])
	months 			= split[1]
	strt_date 		= str(years)+'-01-01'
	end_date 		= str(years)+'-12-31'

	

	
	QRY 			= "SELECT v.invoice_no,v.invoice_date,v.proj_name,v.inr,v.usd,v.rate,v.price,v.qty,sum(v.qty*v.rate*v.price),v.month FROM invoice.it_invoice v,invoice.it_client cl where cl.client_name='"+str(client_ids)+"' and (v.client_id=cl.id) and YEAR(v.invoice_date)='"+str(years)+"' group by v.invoice_no;"
	cursor 			= connection.cursor()
	cursor.execute(QRY)
	get_invoices 	= cursor.fetchall()
	invoice_arr 	= []
	ledger_list     = []

	
	all_inward_amt  = 0
	all_usd_amt 	= 0
	get_remittance_date = models.remittance_data.objects.filter(remittance_date__gte=strt_date,remittance_date__lte=end_date,client_name=client_ids)
	sum_remittance_date = models.remittance_data.objects.filter(remittance_date__gte=strt_date,remittance_date__lte=end_date,client_name=client_ids).aggregate(Sum('inward_amount'))	
	remitt_amt = 0
	for e in get_remittance_date:
		if e.services1!="":
			e_services = e.services1
		else:
			e_services =  e.services2

		if e.inward_amount!=e.net_amount_ref:
			statuss = 'Not Reconciled'
		else:
			statuss = 'Reconciled'
		try:
			invoice_arr.append({				
				'payment_firc' 	: e.certificate_no,				
				'credit_date'   : e.remittance_date.strftime('%d-%b-%Y'),
				'currency1' 	: e.currency,
				'service1' 		: e_services,
				'inward_amt'    : e.inward_amount,
				'status' 		: statuss,	
				'remitt_amt'    : sum_remittance_date["inward_amount__sum"],
			})

			

		except:
			pass

	
	for c in get_invoices:
		# get_certificate = models.remittance_data.objects.filter(certificate_no=c).first()
		# print '------',get_certificate.invoice_no1
		split_month = c[9].split('-')
		last_month  = split_month[2]
		if c[7]!=1:
			get_amount = (c[5]*c[6]*c[7]/int(last_month))
			all_usd_amt+=get_amount
		else:
			get_amount = c[8]
			all_usd_amt+=c[8]
		
		ledger_list.append({
			'invoice_no'	: c[0],
			'invoice_date'	: c[1].strftime('%d-%b-%Y'),
			'currency'		: c[4],
			'usd_amt'		: round(get_amount,0),#c[8],
			'services'		: c[2],	
			
		})


	context={
		'invoice_arr' : invoice_arr,
		'ledger_list' : ledger_list,
		'all_usd_amt' : all_usd_amt
	}
	

	return HttpResponse(json.dumps(context))


@csrf_exempt
def export_ledger_data(request):
	client_name 	= json.loads(request.POST['client_name'])
	fin_date 		= json.loads(request.POST['fin_date'])
	invoice_data   	= json.loads(request.POST['invoice_data'])
	remittance_data = json.loads(request.POST['remittance_data'])	
	select_dir    	= os.path.dirname(__file__)
	srcfile       	= select_dir+'/static/Ledger.xlsx'
	dstroot       	= select_dir+'/Ledger.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      	= os.path.dirname(__file__)		
	location      	= (my_dir+'/Ledger.xlsx')
	wb 	   	      	= load_workbook(my_dir+'/Ledger.xlsx')
	ws_usd        	= wb.get_sheet_by_name("Sheet1")
	wsusd        	= wb.get_sheet_by_name("Sheet1")
	thin_border   	= Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      	= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	_styles6 	 	= Style(font=Font(name='Tahoma', size=9, italic=False,bold=False),alignment=Alignment(horizontal='center'),border=Border(right=Side(style='thin',color='9c9a95'),top=Side(style='thin',color='9c9a95'),bottom=Side(style='thin',color='9c9a95')))
	ii   		  	= 5
	jj 	 		  	= 0
	rr 				= 0
	w 				= 5
	

	ws_usd.cell('C2').value = client_name
	ws_usd.cell('M2').value = fin_date
	for c in remittance_data:
		ws_usd.cell('G'+str(ii)).value = jj+1
		ws_usd.cell('H'+str(ii)).value = c['credit_date']
		ws_usd.cell('I'+str(ii)).value = c['payment_firc']
		ws_usd.cell('J'+str(ii)).value = c['currency1']
		ws_usd.cell('K'+str(ii)).value = c['service1']
		ws_usd.cell('L'+str(ii)).value = c['inward_amt']
		ws_usd.cell('M'+str(ii)).value = c['status']
		ws_usd.cell('G'+str(ii)).style = _styles6	 
		ws_usd.cell('H'+str(ii)).style = _styles6
		ws_usd.cell('I'+str(ii)).style = _styles6
		ws_usd.cell('J'+str(ii)).style = _styles6
		ws_usd.cell('K'+str(ii)).style = _styles6
		ws_usd.cell('L'+str(ii)).style = _styles6
		ws_usd.cell('M'+str(ii)).style = _styles6
		ii+=1
		jj+=1

	for t in invoice_data:
		wsusd.cell('A'+str(w)).value = rr+1
		wsusd.cell('B'+str(w)).value = t['invoice_no']
		wsusd.cell('C'+str(w)).value = t['invoice_date']
		wsusd.cell('D'+str(w)).value = t['services']
		wsusd.cell('E'+str(w)).value = t['currency']
		wsusd.cell('F'+str(w)).value = t['usd_amt']
		wsusd.cell('A'+str(w)).style = _styles6	 
		wsusd.cell('B'+str(w)).style = _styles6
		wsusd.cell('C'+str(w)).style = _styles6
		wsusd.cell('D'+str(w)).style = _styles6
		wsusd.cell('E'+str(w)).style = _styles6
		wsusd.cell('F'+str(w)).style = _styles6

		w+=1
		rr+=1

	sum_row 				=  wsusd.max_row
	sum_rem_amt				=  ws_usd.max_row
	_row_ledger   			= 'F'+str((sum_row))
	_sum_remitt   			= 'L'+str((sum_rem_amt))
	wsusd.cell('F3').value 	= "=SUM(F5:"+_row_ledger+")"
	ws_usd.cell('L3').value = "=SUM(L5:"+_sum_remitt+")"	
	
	wb.save(my_dir+'/Ledger.xlsx')

	return HttpResponse(json.dumps('done'))


def export_legder_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	today 							= datetime.now().date()
	format_year 					= (today).strftime('%Y')
	format_month 					= (today).strftime('%b')
	int_mon							= (today).strftime('%m')
	format_date 					= (today).strftime('%d')
	int_mont 						= int(int_mon)
	mrge_date 						= str(format_date)+'-'+str(format_month)+'-'+str(format_year)	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/Ledger.xlsx')
	filename 	   					= my_dir+'/Ledger.xlsx'
	download_name  					= 'Ledger.xlsx'	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response