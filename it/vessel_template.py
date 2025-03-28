from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
# import wget
#####################################################
def get_calling_for_vessel_template(calling_pdf,client_name,get_month,invoice_no,format_year,format_month,split_address1,split_address2,vm_name,format_invoice_date,customer_no,merge,currency,total_amount,split_address3,split_address4,split_address5,check_vessel):
	#dir      = '/home/munish/Documents/Finance/Finance/Current/Finance/HIM/'+str(client_name)+'/Online/'
	print '-=================================--f-dfdfdfdf'
	dir      = '/var/www/html/invoice/it/static/pdf/HIM/'+str(client_name)+'/Online/'
	if not os.path.exists(dir):
		os.makedirs(dir)

	vvds 				= calling_pdf[0].invoice_no
	split_invoice_no  	= vvds.replace('/','_')
	select_dir    		= os.path.dirname(__file__)
	srcfile       		= select_dir+'/static/Empty.xlsx'
	#dstroot       		= '/home/munish/Documents/Finance/Finance/Current/Finance/HIM/'+str(client_name)+'/Online/'+str(split_invoice_no)+'.xlsx'	
	dstroot       		= '/var/www/html/invoice/it/static/pdf/HIM/'+str(client_name)+'/Online/'+str(split_invoice_no)+'.xlsx' 
	copyfile(srcfile, dstroot)
	my_dir 	      		= os.path.dirname(__file__)		
	#location      		= '/home/munish/Documents/Finance/Finance/Current/Finance/HIM/'+str(client_name)+'/Online/'+str(split_invoice_no)+'.xlsx'	 #(my_dir+'/Clearlake/Empty.xlsx')
	location      		= '/var/www/html/invoice/it/static/pdf/HIM/'+str(client_name)+'/Online/'+str(split_invoice_no)+'.xlsx'
	#wb 			  	= load_workbook('/home/munish/Documents/Finance/Finance/Current/Finance/HIM/'+str(client_name)+'/Online/'+str(split_invoice_no)+'.xlsx')
	wb 			  		= load_workbook('/var/www/html/invoice/it/static/pdf/HIM/'+str(client_name)+'/Online/'+str(split_invoice_no)+'.xlsx')		
	ws 			  		= wb.get_sheet_by_name("Invoice")
	# s_no = 1
	# i    = 25
	h 	 = 7
	arr  = [split_address1,split_address2,split_address3,split_address4,split_address5]
	for e in range(0,4):
		ws.cell('C'+str(h+e)).value  = arr[e]

	sums_amt = 0
	for b in range(0,check_vessel):
		ws.cell('I7').value  = datetime.strptime(calling_pdf[b].month, "%Y-%m-%d").strftime('%d-%b-%y')
		ws.cell('I9').value  = calling_pdf[b].invoice_no
		ws.cell('J6').value  = calling_pdf[b].invoice_no
		ws.cell('I11').value = calling_pdf[b].ship_name
		ws.cell('K11').value = calling_pdf[b].voyage_no
		ws.cell('I12').value = calling_pdf[b].disch_port
		disch_date = calling_pdf[b].disch_date.date().strftime('%d-%b-%y')
		ws.cell('K12').value = disch_date
		ws.cell('E16').value = str(client_name)+'/BW/CHS'
		ws.cell('E18').value = 'Cargo Heating Services'
		ws.cell('E20').value = 'Cargo Heating Management'
		ws.cell('F11').value = calling_pdf[b].vm_name
		img = Image(select_dir+'/static/pdf/1.png')
	 	ws.add_image(img, 'A1')
	 	img = Image(select_dir+'/static/pdf/4.png')
	 	ws.add_image(img, 'J69')
		
		thin_border  			  = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style     			  = Style(font=Font(name='Calibri', size=9, bold=False),border=thin_border)
		_row_style   			  = Style(font=Font(name='Calibri', size=9, bold=False),border=thin_border)		
		ws.cell('C25').value = '1'
		ws.cell('D25').value = 'Cargo Heating Service charges'
		ws.cell('I25').value = calling_pdf[0].qty
		ws.cell('J25').value = calling_pdf[0].client.price
		ws.cell('K25').value = calling_pdf[0].client.price*calling_pdf[0].qty
		ws.cell('K43').value = '=SUM(K25:L42)' #'=SUM(K'+str(i)+':L42)'
		sums_amt+=(calling_pdf[0].client.price*calling_pdf[0].qty)		
		word_amount 		 = convert_money_to_text(sums_amt)
		ws.cell('C47').value = word_amount
		remove_splash_invoice_no = calling_pdf[b].invoice_no.replace('/', '_')	
		extension  		 		 = remove_splash_invoice_no+'_'+client_name+".xlsx"
		
		dir = my_dir+'/static/pdf/'+str(client_name)+'/Online/'
		if not os.path.exists(dir):
			os.makedirs(dir)
		#wb.save(my_dir+'/static/pdf/'+str(client_name)+'/Online/'+str(extension))
		#wb.save('/home/munish/Documents/Finance/Finance/Current/Finance/HIM/'+str(client_name)+'/Online/'+str(extension))	
		wb.save('/var/www/html/invoice/it/static/pdf/HIM/'+str(client_name)+'/Online/'+str(extension))
	return HttpResponse(json.dumps('done'))

def other_template_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/Clearlake/clearlake_template.xlsx')
	filename 	   					= my_dir+'/Clearlake/clearlake_template.xlsx'
	download_name  					= "clearlake_template.xlsx"	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response

def convert_money_to_text(money):
    try:
        money_number = int(money)
    except ValueError:
        return "You must enter Number"
    if money_number == 0:
        return None
    positions 		= [None for i in range(4)]
    key_range 		= 0
    one_place 		= ["", "One ", "Two ", "Three ", "Four ","Five ", "Six ", "Seven ", "Eight ", "Nine "]
    one_ten_place   = ["Ten ", "Eleven ", "Twelve ", "Thirteen ", "Fourteen ","Fifteen ", "Sixteen ", "Seventeen ", "Eighteen ","Nineteen "]
    ten_place 		= ["Twenty ", "Thirty ", "Forty ", "Fifty ","Sixty ", "Seventy ", "Eighty ", "Ninety "]
    name_of_number  = ["Thousand ", "Lakh ", "Crore "]
    money_number_money_text = ''
    positions[0] = money_number % 1000  # units
    positions[1] = money_number / 1000
    positions[2] = money_number / 100000
    positions[1] = positions[1] - 100 * positions[2]  # thousands
    positions[3] = money_number / 10000000  # crores
    positions[2] = positions[2] - 100 * positions[3]  # lakhs
    for counter in range(3, 0, -1):
        if positions[counter] != 0:
            key_range = counter
            break
    for i in range(key_range, -1, -1):
        if positions[i] == 0:
            continue
        ones = positions[i] % 10  # ones
        tens = positions[i] / 10
        hundreds = positions[i] / 100  # hundreds
        tens = tens - 10 * hundreds  # tens
        if (hundreds > 0):
            money_number_money_text += one_place[hundreds] + "Hundred "
        if (ones > 0 or tens > 0):
            if (hundreds > 0 and i == 0):
                money_number_money_text += "and "
            if (tens == 0):
                money_number_money_text += one_place[ones] 
            elif (tens == 1):
                money_number_money_text += one_ten_place[ones] 
            else:
                money_number_money_text += ten_place[tens - 2] + \
                    one_place[ones]  
        if (i != 0):
            money_number_money_text += name_of_number[i - 1]  
    return money_number_money_text  + "Dollars and Zero Cents"
