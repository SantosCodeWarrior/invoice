from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler
from openpyxl.styles.borders import Border, Side, BORDER_THIN

import pytz

#####################################################
def client_data(request):
	if request.user.is_authenticated():
		try:			
			models.api_client_data.objects.all().delete()
		except:
			pass

		login_user 			= request.user
		start_date_format 	= '2024-01-01'
		end_date_format 	= '2024-08-30'
		api_urlx 			= "https://aboss.bwesglobal.com/api/get_static_details/"		
		api_methodx     	= "GET"			
		parametersx     	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_date_format),'end_date': str(end_date_format)}
		responsex       	= requests.get(api_urlx, params=parametersx,verify=False)		
		client_listx    	= json.loads(responsex.content)
		cl_dets 			= client_listx['ship_details']
		error 				= 'Done'
		#print responsex
		for key in cl_dets:			
			#if cl_dets[key]["clients_id"][0]==63:
			try:
				account_tab = cl_dets[key]["account_tab"]				
			except:
				account_tab = ''


			
			patn 			= re.sub(r"[\([{})\]]", "", str(cl_dets[key]["clients_id"]))			
			ship_id     	= cl_dets[key]["id"]
			client_id   	= patn	
			ship_name   	= cl_dets[key]["ship_name"]			
			
			db1				= models.api_client_data()				
			db1.account_tab = account_tab			 				 	
			db1.ship_name   = ship_name
			db1.address 	= cl_dets[key]["address"]
			db1.client_id 	= str(patn)
			db1.ship_id     = ship_id			 	
			db1.save()		

		context={
		
		'login_user' :  login_user,
		}
		return render_to_response("invoice_display/client_json.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')

@csrf_exempt
def client_list(request):
	if request.user.is_authenticated():
		start_date_format 	= '2024-01-01'
		end_date_format 	= '2024-08-30'
		current_date 		= datetime.now().date()
		tabs_name 			= json.loads(request.POST['tabs_name'])
		
		api_urlx 			= "https://aboss.bwesglobal.com/api/get_static_details/"		
		api_methodx     	= "GET"			
		parametersx     	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_date_format),'end_date': str(end_date_format)}
		responsex       	= requests.get(api_urlx, params=parametersx,verify=False)		
		client_listx    	= json.loads(responsex.content)		
		cl_dets 			= client_listx['client_details']
		client_array 		= []
		#print api_urlx
		t=0
		for key in cl_dets:
			get_status   	= tabs_name[t][6]
			proj_name    	= tabs_name[t][3]
			currency_type 	= tabs_name[t][4]
			client_name  	= cl_dets[key]["name"]		
			check_client 	= models.boss_client.objects.filter(client_name=client_name,id=int(key)).count()
			try:
				get_prev_date  = models.boss_client.objects.filter(client_name=client_name,id=int(key)).order_by('-curr_date').first()
			except:
				get_prev_date  = current_date			
			
			if check_client!=0:		
				db = models.boss_client.objects.filter(client_name=client_name,id=int(key)).first()
				new_client 		= ''
				db.curr_date    = str(get_prev_date.curr_date)				
			else:
				db = models.boss_client()
				new_client 		= client_name
				db.curr_date 	= (current_date)
			
			db.id 				= key
			db.client_name 		= client_name
			db.proj_name 		= proj_name	
			db.currency_type 	= currency_type	
			if get_status==1:
				checked = True				
			else:
				checked = False
			db.status   		= checked						
			db.save()			
			t+=1

		login_user 	    = request.user
		app_url 		= request.path		
		split_url 		= app_url.split('/')
		url_nme 		= split_url[2].capitalize()
		IST 			= pytz.timezone('Asia/Kolkata') 
		datetime_ist 	= datetime.now(IST)		 
		now_date 		= datetime_ist.strftime('%Y-%m-%d %H:%M:%S')
		InID 			= models.log_sessions()
		InID.date  		= now_date 
		InID.user_name  = login_user
		InID.url_name   = "New Client API"#url_nme
		#InID.invoice_no = new_client
		InID.save()	
		print '-------',now_date

		return HttpResponse(json.dumps('done'))
	else:
		return HttpResponseRedirect('/it/user_login')



@csrf_exempt
def client_log(request):
	start_date_format 	= '2024-01-01'
	end_date_format 	= '2024-08-30'
	current_date 		= datetime.now().date()
	api_urlx 			= "https://aboss.bwesglobal.com/api/get_static_details/"		
	api_methodx     	= "GET"			
	parametersx     	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_date_format),'end_date': str(end_date_format)}
	responsex       	= requests.get(api_urlx, params=parametersx,verify=False)		
	client_listx   	 	= json.loads(responsex.content)
	
	cl_dets 			= client_listx['client_details']
	#print '----------',api_urlx,'---',parametersx,'----',responsex
	client_array 		= []
	currency_array      = []
	s_no=0
	for key in cl_dets:
		client_name  = cl_dets[key]["name"]		
		#print '--------->>>', client_name
		check_client = models.boss_client.objects.filter(client_name=client_name,id=int(key)).count()		
		
		if check_client!=0:
			db = models.boss_client.objects.filter(client_name=client_name,id=int(key)).first()
			new_client = ''
		else:
			db = models.boss_client()
			new_client = client_name
		
		db.id 			= key
		db.client_name 	= client_name		
		if db.status==1:
			checked = True
		else:
			checked = False
		#db.save()

		try:
			proj_name = db.proj_name
		except:
			proj_name = ''

		try:
			currency = db.currency_type
		except:
			currency = ''

		try:
			entry_date = db.curr_date
		except:
			entry_date = ''


		#print '---------',proj_name
		client_array.append({
			'id' 			: key,
			'name' 			: client_name,	
			'new'   		: new_client,
			'status' 		: checked,
			'proj_name' 	: proj_name,
			'currency_type' : currency,
			's_no'			: int(s_no)+1,
			'entry_date'    : str(entry_date),

			})
		s_no+=1

	context={
	'client_array'   : client_array,	
	}

	
	
	return HttpResponse(json.dumps(context))

def get_currency_list(request):	
	currency_list  = models.currency_data.objects.all().order_by('currency')
	currency_array = []
	for e in currency_list:
		currency_array.append({
			'id' 		: e.id,
			'currency' 	: e.currency,
			})
	return HttpResponse(json.dumps(currency_array))


def currency_entry(request):
	n_currency 	= request.GET['n_currency']
	r_currency  = n_currency.strip()
	chk  		= models.currency_data.objects.filter(currency=r_currency).count()
	if chk!=0:
		db  	= models.currency_data.objects.filter(currency=r_currency).first()
	else:
		db  	= models.currency_data()
	db.currency = r_currency.upper()
	db.save()

	return HttpResponse(json.dumps('done'))


@csrf_exempt
def export_client_json(request):
	try:
		client_json_details = json.loads(request.POST['datas'])
		ii = 3
		jj = 1
		select_dir    = os.path.dirname(__file__)
		srcfile       = select_dir+'/static/API Clients.xlsx'
		dstroot       = select_dir+'/API Clients.xlsx'	 
		copyfile(srcfile,dstroot)
		my_dir 	      = os.path.dirname(__file__)		
		location      = (my_dir+'/API Clients.xlsx')
		wb 	   	      = load_workbook(my_dir+'/API Clients.xlsx')
		ws_usd        = wb.get_sheet_by_name("Sheet1")
		thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)

		for f in client_json_details:
		 	if f[0]!=None or f[1]!=None or f[2]!=None or f[3]!=None or f[4]!=None or f[5]!=None or f[6]!=None or f[7]!=None:
		 		client_id 		= f[0]
		 		client_name  	= f[2]
		 		proj_name 		= f[3]
		 		currencys 		= f[4]
		 		new_client 		= f[5]
		 		active			= f[6]
		 		entry_date 		= f[7]				

				thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
				my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
				_row_style    = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
				srow_style    = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='center'))

				ws_usd.cell('A'+str(ii)).value = client_id
				ws_usd.cell('B'+str(ii)).value = jj
				ws_usd.cell('C'+str(ii)).value = client_name
				ws_usd.cell('D'+str(ii)).value = proj_name
				ws_usd.cell('E'+str(ii)).value = currencys
				ws_usd.cell('F'+str(ii)).value = new_client
				if active==1:
					flag = 'Yes'
				else:
					flag = 'No'

				try:
					format_entry_date = datetime.strptime(entry_date,"%Y-%m-%d").strftime('%d-%b-%Y')
				except:
					format_entry_date = ''

				ws_usd.cell('G'+str(ii)).value = flag
				ws_usd.cell('H'+str(ii)).value = format_entry_date
				
				ws_usd.cell('A'+str(ii)).style = srow_style	 
				ws_usd.cell('B'+str(ii)).style = srow_style	
				ws_usd.cell('C'+str(ii)).style = _row_style	 	
				ws_usd.cell('D'+str(ii)).style = srow_style	
				ws_usd.cell('E'+str(ii)).style = srow_style	
				ws_usd.cell('F'+str(ii)).style = srow_style	
				ws_usd.cell('G'+str(ii)).style = srow_style	
				ws_usd.cell('H'+str(ii)).style = srow_style	
				
				ii+=1
				jj+=1

				wb.save(my_dir+'/API Clients.xlsx')

	except:
		pass


	return HttpResponse(json.dumps('done'))


def export_client_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/API Clients.xlsx')
	filename 	   					= my_dir+'/API Clients.xlsx'
	download_name  					= "API Clients.xlsx"	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response