from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
#####################################################


@csrf_exempt
def boss_export_tracker(request):
	track_details = json.loads(request.POST['track_details'])
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/TRACKER-EXPORT.xlsx'
	dstroot       = select_dir+'/Tracker_Export/TRACKER-EXPORT.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/Tracker_Export/TRACKER-EXPORT.xlsx')
	wb 	   	      = load_workbook(my_dir+'/Tracker_Export/TRACKER-EXPORT.xlsx')
	ws     	      = wb.get_sheet_by_name("Invoice")
	ws_inr 	      = wb.get_sheet_by_name("Invoice(INR)")
	ws_usd 	      = wb.get_sheet_by_name("Invoice(USD)")
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	ii   		  = 3
	jj 	 		  = 1

	for t in track_details:		
		thin_border    				   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style       				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
		_row_style     				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)		
		invoice_no 	   				   = t['Invoice No.'][:9]
		invoicesx 	   				   = models.invoice.objects.filter(invoice_no=invoice_no,proj_name='BOSS').first()		
		month_number      	           = datetime.strptime(invoicesx.month, "%Y-%m-%d").strftime('%m')	
		month_name      	           = datetime.strptime(invoicesx.month, "%Y-%m-%d").strftime('%b')	
		year_name      	   	           = datetime.strptime(invoicesx.month, "%Y-%m-%d").strftime('%Y')		
		get_month 	 	  	           = monthrange(int(year_name),int(month_number))[1]	
		start_date_for 	  	           = '01 '+str(month_name)
		end_date_for   	  	           = str(month_name)+'-'+str(get_month)			
		#print '----------',invoicesx.month,'----',invoicesx.invoice_no

		ws_usd.cell('A'+str(ii)).value = jj
		ws_usd.cell('B'+str(ii)).value = t['Invoice No.']
		ws_usd.cell('C'+str(ii)).value = t['Dated']
		ws_usd.cell('D'+str(ii)).value = t['Vessel/Month']
		ws_usd.cell('E'+str(ii)).value = t['Period']			
		ws_usd.cell('F'+str(ii)).value = t['Voy No']
		ws_usd.cell('F'+str(ii)).value = t['PIC']
		ws_usd.cell('I'+str(ii)).value = t['Client']					
		ws_usd.cell('J'+str(ii)).value = t['USD']					
		ws_usd.cell('K'+str(ii)).value = t['Recvd Date']
		ws_usd.cell('A'+str(ii)).style = _row_style	 
		ws_usd.cell('B'+str(ii)).style = _row_style	 
		ws_usd.cell('C'+str(ii)).style = _row_style	 
		ws_usd.cell('D'+str(ii)).style = _row_style	 
		ws_usd.cell('E'+str(ii)).style = _row_style	 
		ws_usd.cell('F'+str(ii)).style = _row_style
		ws_usd.cell('G'+str(ii)).style = _row_style	 	
		ws_usd.cell('H'+str(ii)).style = _row_style	 	
		ws_usd.cell('I'+str(ii)).style = _row_style	 	
		ws_usd.cell('J'+str(ii)).style = _row_style	 	
		ws_usd.cell('K'+str(ii)).style = _row_style	 	
		ws_usd.cell('L'+str(ii)).style = _row_style	 	
		ws_usd.cell('M'+str(ii)).style = _row_style	 	
		ws_usd.cell('N'+str(ii)).style = _row_style	 
		ws_usd.cell('O'+str(ii)).style = _row_style	 	
		ws_usd.cell('P'+str(ii)).style = _row_style	
		ws_usd.cell('Q'+str(ii)).style = _row_style
		ws_usd.cell('R'+str(ii)).style = _row_style		
		ii+=1
		jj+=1	
		
		wb.save(my_dir+'/Tracker_Export/TRACKER-EXPORT.xlsx')
	return HttpResponse(json.dumps('done'))

def invoice_export(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)
	location 	   					= (my_dir+'/Tracker_Export/TRACKER-EXPORT.xlsx')
	filename 	   					= my_dir+'/Tracker_Export/TRACKER-EXPORT.xlsx'
	download_name  					= "TRACKER-EXPORT.xlsx.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response