from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################
def account_data(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		login_user 	= request.user

		context={
		
		'login_user' :  login_user,
		}
		return render_to_response("invoice_display/account_json.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')

@csrf_exempt
def account_list(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		start_date_format 	= '2023-02-01'
		end_date_format 	= '2023-02-28'
		current_date 		= datetime.now().date()
		tabs_name 			= json.loads(request.POST['tabs_name'])
		
		api_urlx 			= "https://bossv2.bwesglobal.com/api/get_static_details/"		
		api_methodx     	= "GET"			
		parametersx     	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_date_format),'end_date': str(end_date_format)}
		responsex       	= requests.get(api_urlx, params=parametersx,verify=False)		
		client_listx    	= json.loads(responsex.content)		
		account_list 		= client_listx['ship_details']
		client_array 		= []

		t=0
		for key in account_list:
			print '-------',key


			get_status   	= tabs_name[t][6]
			proj_name    	= tabs_name[t][3]
			currency_type 	= tabs_name[t][4]
			client_name  	= cl_dets[key]["name"]		
			check_client 	= models.boss_client.objects.filter(client_name=client_name,id=int(key)).count()
			try:
				get_prev_date  = models.boss_client.objects.filter(client_name=client_name,id=int(key)).order_by('-curr_date').first()
			except:
				get_prev_date  = current_date			
			
			if check_client!=0:		
				db = models.boss_client.objects.filter(client_name=client_name,id=int(key)).first()
				new_client 		= ''
				db.curr_date    = str(get_prev_date.curr_date)
				
			else:
				db = models.boss_client()
				new_client 		= client_name
				db.curr_date 	= (current_date)
			
			db.id 				= key
			db.client_name 		= client_name
			db.proj_name 		= proj_name	
			db.currency_type 	= currency_type	
			if get_status==1:
				checked = True				
			else:
				checked = False
			db.status   		= checked						
			db.save()
			
			t+=1		
		

		return HttpResponse(json.dumps('done'))
	else:
		return HttpResponseRedirect('/it/user_login')



@csrf_exempt
def account_log(request):
	todays  			= datetime.now().date()
	format_year 		= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%Y')
	format_month 		= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%b')
	int_month 			= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%m')
	get_month 			= monthrange(int(format_year),int(int_month)-1)[1]
	prev 				= str(int(int_month)-1)
	if prev<="9":
		prev_month = "0"+str(prev)
	else:
		prev_month = prev

	st_date = str(format_year)+'-'+(prev_month)+'-01'
	et_date = str(format_year)+'-'+(prev_month)+str(get_month)

	start_date_format 	= st_date
	end_date_format 	= et_date

	api_urlx 			= "https://bossv2.bwesglobal.com/api/get_static_details/"		
	api_methodx     	= "GET"			
	parametersx     	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_date_format),'end_date': str(end_date_format)}
	responsex       	= requests.get(api_urlx, params=parametersx,verify=False)		
	client_listx   	 	= json.loads(responsex.content)	
	cl_dets 			= client_listx['ship_details']	
	account_array 		= []
	
	s_no=0
	for key in cl_dets:
		cl_id 	= re.sub(r"[\([{})\]]", "", str(cl_dets[key]["clients_id"]))	
		aa 		= cl_id.split(',')
		cl_ids 	= aa[0]
		for v in cl_dets[key]["clients_id"]:
			client_name = models.boss_client.objects.filter(id=v,proj_name='BOSS').first()
			try:
				cl_name = client_name.client_name
			except:
				cl_name = ''

		try:
			get_address = cl_dets[key]["address"]
		except:
			get_address = ""

		account_array.append({
			's_no'			: s_no,
			'account_tab'   : cl_dets[key]["account_tab"],
			'address' 		: cl_dets[key]["address"],
			'ship_name' 	: cl_dets[key]["ship_name"],
			'client_id' 	: cl_id,
			'client_name'   : cl_name,
			'get_address'   : get_address,		
		})
		s_no+=1

	
	context={
	'account_array'   : account_array,	
	}

	
	
	return HttpResponse(json.dumps(context))

def get_currency_list(request):
	
	currency_list  = models.currency_data.objects.all().order_by('currency')
	currency_array = []
	for e in currency_list:
		currency_array.append({
			'id' 		: e.id,
			'currency' 	: e.currency,
			})

	return HttpResponse(json.dumps(currency_array))


def currency_entry(request):
	n_currency 	= request.GET['n_currency']
	r_currency  = n_currency.strip()
	chk  		= models.currency_data.objects.filter(currency=r_currency).count()
	if chk!=0:
		db  	= models.currency_data.objects.filter(currency=r_currency).first()
	else:
		db  	= models.currency_data()

	db.currency = r_currency.upper()
	db.save()

	return HttpResponse(json.dumps('done'))
