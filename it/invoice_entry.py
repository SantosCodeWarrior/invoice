
from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
import pytz
#####################################################


def invoice_entry(request):
	if request.user.is_authenticated():		
		login_user = request.user
		currency_details = models.currency_data.objects.all().order_by('currency')
		context={
			'login_user' 		: login_user,
			'currency_details' 	: currency_details
		}			
		return render_to_response("invoice_display/invoice_entry.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')


def get_other_invoice(request):
	if request.user.is_authenticated():
		typex      = request.GET['type']
		proj_name  = request.GET['proj_name']
		login_user = request.user
		cl_array   = []
		single_arr = []
		if proj_name=='BOSS':
			cl_list	  = models.Client.objects.filter(proj_name=proj_name,currency_type=typex,status=1).exclude(client_name='')
		else:
			cl_list	  = models.Client.objects.filter(proj_name=proj_name,currency_type=typex,status=1).exclude(client_name='')
			
		for t in cl_list:
			if t.client_name not in single_arr:
				single_arr.append(t.client_name)

		for c in single_arr:
			cl	  = models.Client.objects.filter(proj_name=proj_name,currency_type=typex,client_name=c).first()
			
			cl_array.append({
				'id'          : cl.id,
				'client_name' : c
			})
		
		if typex=='USD':
			invoice_details = models.invoice.objects.filter(proj_name=proj_name,usd='USD').order_by('-invoice_no').first()
		elif typex=='INR':
			invoice_details = models.invoice.objects.filter(proj_name=proj_name,inr='INR').order_by('-invoice_no').first()

		get_invoice = []
		slp         = invoice_details.invoice_no.split('/')
		first_no    = int(slp[0])+1
		second_no   = str(first_no)+'/'+str(slp[1])

		get_invoice.append({
			'invoice_no' : second_no,
		})

		context={

			'get_invoice' : get_invoice,
			'cl_array'    : sorted(cl_array)
		}
		
		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def submit_other_invoice(request):
	if request.user.is_authenticated():
		table_details = json.loads(request.POST['table_details'])
		invoice_nox	  = request.POST['invoice_no']
		proj_name	  = request.POST['proj_name']
		curr_type 	  = request.POST['curr_type']
		vm_name       = request.POST['vm_name']
		in_date 	  = request.POST['in_date']
		cl_name       = request.POST['cl_name']
		gtin_no       = request.POST['gtin_no']
		address 	  = request.POST['address']
		pool_name     = request.POST['pool_name']
		period_dat 	  = request.POST['period_dat']
		txt_remarks   = request.POST['tt_remarks']
		login_user 	  = request.user
		#disch_port    = request.GET['disch_port']
		#disch_date    = request.GET['disch_date']
		#voyage_no     = request.GET['voyage_no']
		#nominate_date = request.GET['nominate_date']
		
		#print '======',invoice_nox,'====',proj_name,'=====',curr_type,'=====',vm_name,'=====',in_date,'====',cl_name,'====',address
		invoice_date  = datetime.strptime(in_date, "%m/%d/%Y").strftime('%Y-%m-%d')	
		period_date   = datetime.strptime(period_dat, "%m/%d/%Y").strftime('%Y-%m-%d')	

		#print '----------',	invoice_date,'----',period_date
		s     		  = 0
		check_client  = models.Client.objects.filter(client_name=cl_name,currency_type=curr_type,proj_name=proj_name).count()		
		if check_client!=0:
			db_client =  models.Client.objects.filter(client_name=cl_name,currency_type=curr_type,proj_name=proj_name).first()
		else:
			db_client =  models.Client()

		db_client.client_name   = cl_name
		db_client.proj_name     = proj_name
		db_client.currency_type = curr_type
		db_client.rate 			= 1
		#db_client.price 		= 1
		db_client.price_type    = None
		db_client.duration_type = 'Voyagewise'
		db_client.vm_name       = cl_name
		db_client.tin_number    = gtin_no		
		db_client.status 		= 1
		db_client.save()

		check_dbx = models.pool_master.objects.filter(client_id=db_client.id,pool=cl_name).count()
		if check_dbx!=0:
			save_db = models.pool_master.objects.filter(client_id=db_client.id,pool=cl_name).first()
		else:
			save_db = models.pool_master()
		save_db.address   = address
		save_db.pool      = cl_name
		save_db.client_id = db_client.id
		save_db.vm_name   = vm_name		
		save_db.save()
		
		for f in table_details:
			if f['description']!=None:
				ship_name  = f['description']
				unit 	   = f['unit_cost']
				qty        = f['qty']	
				rate 	   = f['rate_inr']
				passage    = f['passage']
				report_id  = f['report_id']
				heading    = f['heading']

				#print '---------',cl_name
				s+=float(unit)*float(qty)
				if cl_name!='Shell':
					if report_id==0 or report_id=='' or report_id==None:				
						check_db = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,ship_name=ship_name,client__client_name=cl_name).count()				
					else:
						check_db = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,ship_name=ship_name,client__client_name=cl_name,voyage_no=report_id).count()				
					
					if check_db!=0:
						if report_id=='' or report_id==0 or report_id==None:					
							submit_invoice = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,ship_name=ship_name,client__client_name=cl_name).first()
						else:
							submit_invoice = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,ship_name=ship_name,client__client_name=cl_name,voyage_no=report_id).first()
					else:
						submit_invoice = models.invoice()
				else:
					submit_invoice = models.invoice()

				if len(table_details)>1:
					invoice_noz 			  = invoice_nox
					submit_invoice.invoice_no = invoice_noz
				else:
					submit_invoice.invoice_no = invoice_noz
	
				submit_invoice.ship_name      = ship_name
				submit_invoice.proj_name      = proj_name				
				submit_invoice.cancel_invoice = 0
				submit_invoice.vm_name 		  = vm_name
				submit_invoice.invoice_date   = invoice_date
				if curr_type=='USD':
					submit_invoice.usd  	  = 'USD'
					curr_type = 'USD'
					submit_invoice.inr  	  = None
				if curr_type=='INR':
					submit_invoice.usd  	  = None
					submit_invoice.inr  	  = 'INR'
					curr_type = 'INR'
				submit_invoice.month 	 	  = invoice_date
				submit_invoice.counter 		  = 0
				submit_invoice.client_id      = db_client.id
				submit_invoice.price 		  = unit
				submit_invoice.qty 			  = qty
				submit_invoice.rate  		  = rate
				submit_invoice.heading 		  = heading

				# try:
				splitm 	= str(period_date).split('-')
				calc 	= float(qty)*float(unit)/int(splitm[2])
				# except:
				# 	calc 	= rate
				#print '------->>>>>>>>>>>',period_date#,'----',#(qty*unit)/(splitm[2])


				submit_invoice.usd_amount     = calc

				# try:
				# 	submit_invoice.disch_port       = disch_port
				# 	submit_invoice.disch_date       = disch_date
				# 	submit_invoice.voyage_no        = voyage_no
				# 	submit_invoice.nomination_date  = nominate_date
				# except:
				# 	submit_invoice.disch_port       = None
				# 	submit_invoice.disch_date       = None
				# 	submit_invoice.voyage_no        = None
				# 	submit_invoice.nomination_date  = None
				 
				#if cl_name!='Poompuhar Shipping Corporation Limited':
				#	submit_invoice.voyage_no  = '' 
				#	submit_invoice.deadwt     = 0
				#else:
				try:
					submit_invoice.voyage_no  = passage
					submit_invoice.deadwt     = report_id
				except:
					submit_invoice.voyage_no  = '' 
					submit_invoice.deadwt     = 0

				if cl_name=='Shell' or cl_name=='shell':
					submit_invoice.vessel_type  = pool_name
				else:
					submit_invoice.vessel_type  = db_client.client_name
				try:
					submit_invoice.account_type = pool_name
				except:
					submit_invoice.account_type = None

				submit_invoice.client_address   = address
				submit_invoice.month 			= period_date
				submit_invoice.remark 			= txt_remarks
				submit_invoice.save()
				#print '-----------',period_date

			else:
				pass

		update_invoice_details = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,client__id=db_client.id)
		for x in update_invoice_details:			
			x.total_amount = s
			x.save()

		app_url 		= request.path		
		split_url 		= app_url.split('/')
		url_nme 		= split_url[2].capitalize()

		IST 			= pytz.timezone('Asia/Kolkata') 
		datetime_ist 	= datetime.now(IST)		 
		now_date 		= datetime_ist.strftime('%Y-%m-%d %H:%M:%S')
		InID 			= models.log_sessions()
		InID.date  		= now_date
		InID.user_name  = login_user
		InID.url_name   = "Manual Invoice Entry"
		InID.invoice_no = invoice_nox
		InID.save()	
		return HttpResponse(json.dumps('done'))
	else:
		return HttpResponseRedirect('/it/user_login')


def get_pool_details(request):	
	if request.user.is_authenticated():		
		
		cl_name 	  	= request.GET['cl_name']	
		proj_name 		= request.GET['proj_name']
		curr_type 		= request.GET['curr_type'] 	
		
		api_url 	  	= "https://bossv2.bwesglobal.com/api/get_static_details/"	
		# api_url 	  	= "http://0.0.0.0:8004/api/get_data_boss/"
		api_method    	= "GET"		
		parameters    	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': cl_name}
		response      	= requests.get(api_url, params=parameters,verify=False)	
		client_listx  	= json.loads(response.content)
		cl_dets 	  	= client_listx['ship_details']
		
					

		ship_pool_boss 	= []
 	 	#pool_array     = models.pool_master.objects.filter(client_id=64)
 	 	#pool_array    	= models.pool_master.objects.filter(client__client_name=cl_name,client__proj_name=proj_name,client__currency_type=curr_type)
 	 	query 		= "SELECT pm.pool,pm.address,cl.proj_name,cl.currency_type FROM it_pool_master pm, it_client cl where (cl.id=pm.client_id) and cl.client_name='"+str(cl_name)+"' and cl.proj_name='"+str(proj_name)+"' and cl.currency_type='"+str(curr_type)+"' group by pool"
		cursor 		= connection.cursor()
		cursor.execute(query)
		pool_array  = cursor.fetchall()
		
		for x in pool_array:
			pool 		= x[0]			
			address 	= x[1]
			proj_name 	= x[2]
			curr_type 	= x[3]
			
						
		 	client_details_boss = models.Client.objects.filter(client_name=cl_name,proj_name='BOSS').first()
		 	if client_details_boss:
		 	 	client_rate   = client_details_boss.rate
		 	 	if client_details_boss.proj_name=='BOSS' and client_details_boss.currency_type=='INR':
		 	 		proj_name = 'BOSS'
		 	 		curr_type = 'INR'
		 	 	elif client_details_boss.proj_name=='BOSS' and client_details_boss.currency_type=='USD':
		 	 		proj_name = 'BOSS'
		 	 		curr_type = 'USD'
		 	else:
		 	 	client_rate = 1.0

		 	
			ship_pool_boss.append({
			 	'ship_class'  		: pool,
			 	'client_rate_boss' 	: client_rate,
			 	'proj_name' 		: proj_name,
			 	'curr_type' 		: curr_type,
			 	'address' 			: address,
			 	})

			#print ship_pool_boss
			
			
		return HttpResponse(json.dumps(ship_pool_boss))
	else:
		return HttpResponseRedirect('/it/user_login')


def get_address_details(request):
	if request.user.is_authenticated():
		client_name = json.loads(request.GET['cl_name'])
		proj_name   = json.loads(request.GET['proj_name'])
		curr_type   = json.loads(request.GET['curr_type'])
		pool_name   = json.loads(request.GET['pool_name'])
		get_array 	= []
		get_address = 'no'
		get_client  = models.Client.objects.filter(client_name=client_name,proj_name='BOSS').first()
		get_gst  	= models.BlueWater.objects.filter(client_name=client_name,proj_name='BOSS').first()
			
		api_url 	= "https://aboss.bwesglobal.com/api/get_static_details/"		
		api_method  = "GET"			
		parameters  = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'client_id': get_client.id,'account_tab': pool_name}
		response    = requests.get(api_url, params=parameters,verify=False)		
		boss_array  = json.loads(response.content)
		
		for x in boss_array['ship_details']:
			for c in boss_array['ship_details'][x]['clients_id']:
				#print '------',boss_array['ship_details'][x]['address']
				if c==int(get_client.id):
					if boss_array['ship_details'][x]['account_tab']==pool_name:
						get_address = boss_array['ship_details'][x]['address']

			
		pl_array    = models.pool_master.objects.filter(client__client_name=client_name,client__proj_name=proj_name,client__currency_type=curr_type,pool=pool_name).first()
		
		try:
			vm_name = pl_array.vm_name
		except:
			vm_name = ""

		try:
			address = pl_array.address
		except:
			address = get_address

		try:
			cgst = get_gst.tin_number
		except:
			cgst = ''
		#print '----',cgst,'------',vm_name,'----',client_name
		get_array.append({
			'cgst' 		: cgst,
			'vm_name' 	: vm_name, 
			'address' 	: address,
		})

		context={
			'get_array' : get_array
		}

		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')


def insert_for_new_client(request):
	e_rate 	 		= request.GET['e_rate']
	e_price 	 	= request.GET['e_price']
	e_address 	 	= request.GET['e_address']
	e_vm_name 		= request.GET['e_vm_name']
	e_proj_type 	= request.GET['e_proj_type']	
	e_client_name 	= request.GET['e_client_name']
	e_currency_type = request.GET['e_currency_type']

	CheckCl   = models.Client.objects.filter(client_name=e_client_name).count()	
	if CheckCl!=0:
		db  = models.Client.objects.filter(client_name=e_client_name).first()
	else:
		db  = models.Client()

	db.client_name  	= e_client_name
	db.proj_name 		= e_proj_type
	db.currency_type 	= e_currency_type
	db.duration_type    = 'Voyagewise'
	db.price 			= e_price
	db.rate 			= e_rate
	db.tin_number 		= None
	db.status 			= 1
	db.tax 				= None
	db.change_dollar    = 0
	try:
		db.vm_name 		= e_vm_name
	except:
		db.vm_name 		= ""
	db.save()

	Checkpm  = models.pool_master.objects.filter(client_id=db.id).count()
	if Checkpm!=0:
		pm  =  models.pool_master.objects.filter(client_id=db.id).first()
	else:
		pm 	= models.pool_master()

	pm.pool  			= e_client_name
	pm.address 			= e_address
	pm.client_id 		= int(db.id)
	pm.save()
	
	

	return HttpResponse(json.dumps('done'))