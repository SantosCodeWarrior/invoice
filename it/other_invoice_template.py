from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler
# import wget
#####################################################


def other_invoice_type(request):
	login_user 		= request.user

	context={	
		'login_user' : login_user,
	}
	return render_to_response('invoice_display/other_invoice_template.html',context)

def show_invoice_template(request):
	invoice_no 		= json.loads(request.GET['invoice_no'])[0]
	e_types 		= json.loads(request.GET['e_types'])[0]	
	invoice_details = models.invoice.objects.filter(invoice_no=invoice_no)
	proj_nme 		= invoice_details[0].proj_name
	cl_name 		= invoice_details[0].client.client_name
	currency_type   = invoice_details[0].client.currency_type
	customer_no  	= invoice_details[0].client.tin_number
	try:
		split_f 	= invoice_details[0].client.price_type.split('-')
		check_vess  = split_f[0]
	except:
		check_vess  = ''


	invoice_array   = []
	for c in invoice_details:
		disch_dt 	= c.disch_date.strftime('%d-%b-%y')
		join_disch 	= str(c.disch_port)+','+str(disch_dt)
		get_address = c.client_address#.replace(',','<br>');
		join_voy_no = str(c.ship_name)+','+str(c.voyage_no);
		voyage_no   = c.voyage_no
		try:
			remarks = c.remark
		except:
			remarks = ''

		get_aed_amt = c.price*c.qty*c.rate
		today 		= datetime.now().date()

		if cl_name=='Shell':
			due_date = c.invoice_date + timedelta(days=60)
		else:
			due_date = c.invoice_date + timedelta(days=30)

		try:
			voyID = c.voyage_id
		except:
			voyID = ''
		
		invoice_array.append({
			'vm_name' 	 	: c.vm_name,
			'address' 	 	: get_address,
			'invoice_no' 	: c.invoice_no,
			'disch_port' 	: c.disch_port,
			'disch_date' 	: (c.disch_date).strftime('%m/%d/%Y'),
			'remarks'	 	: remarks,
			'price'		 	: c.price,
			'qty'		 	: c.qty,
			'rate' 		 	: c.rate,
			'tot_amt'	 	: c.total_amount,
			'join_disch' 	: join_disch,
			'join_voy_no'	: join_voy_no,
			'voyage_no'		: voyage_no,
			'account_type' 	: c.account_type,
			'vessel_type'	: c.vessel_type,
			'ship_name'		: c.ship_name,
			'price_type'	: c.price_type,
			'invoice_date'  : (c.invoice_date).strftime('%d-%b-%Y'),
			'today'			: today.strftime('%d-%b-%Y'),
			'due_date'      : due_date.strftime('%d-%b-%Y'),
			'voyID'			: voyID,
		})
		#generate_pdf(invoice_array,proj_nme,cl_name,currency_type,customer_no,check_vess,get_aed_amt,e_types)
		if e_types=='AED':
			generate_pdf(invoice_array,proj_nme,cl_name,currency_type,customer_no,check_vess,get_aed_amt,e_types)
		elif e_types=='APP':
			generate_app_pdf(invoice_array,proj_nme,cl_name,currency_type,customer_no,check_vess,get_aed_amt,e_types,invoice_no,sums)
		else:
			print '0000000;;'
			return HttpResponseRedirect('/it/chm_tracker')
		
	
	context={
	'invoice_array' : invoice_array,
	}
	return HttpResponse(json.dumps(context))
	

def get_generate_pdf(request):	
	# from subprocess import call
	# cmd = 'scp root@bossv2.bwesglobal.com:/home/bossbwtw/Data/WeatherDatav2/wave_7_day_set1.hdf5 /home/bossbwtw/Data/WeatherData/WaveData/'
	# call(cmd.split(" "))
	# print '-------',cmd

	return HttpResponse(json.dumps('done'))


def generate_pdf(invoice_array,proj_name,client_name,currency,customer_no,check_vessel,currency_amount,e_types):
	from calendar import monthrange
	import calendar
	split_address6 	= ''
	array 			= []

	if proj_name=='BOSS':
		tag_name = 'BIM'
		img_tag  = 'boss.png'
	elif proj_name=='CHM':
		tag_name = 'HIM'
		img_tag  = 'chm.png'
	
	dir      = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/'	
	
	if not os.path.exists(dir):
		os.makedirs(dir)
	srcfile      = '/var/www/html/invoice/it/static/pdf/1.png'
	dstroot      = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/1.png'	
	copyfile(srcfile, dstroot)
	srcfile1     = '/var/www/html/invoice/it/static/pdf/rings.png'
	dstroot1     = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/rings.png'	
	copyfile(srcfile1, dstroot1)
	srcfile2     = '/var/www/html/invoice/it/static/pdf/'+str(img_tag)
	dstroot2     = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/'+img_tag	
	copyfile(srcfile2, dstroot2)
	srcfile3     = '/var/www/html/invoice/it/static/pdf/'+str(img_tag)
	dstroot3     = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/'+img_tag	
	copyfile(srcfile3, dstroot3)
		

	voy = {}
	amt = 0
	vnt = 0
	split_address11 = ''
	split_address22 = ''
	split_address33 = ''
	split_address44 = ''
	split_address55 = ''
	split_address66 = ''
	aa = 0
	tt = 0
	#print '--------',check_vessel
	check_vessel = 1
 	for x in invoice_array:
 		ship_name  = x['ship_name']
		voy_no     = x['voyage_no']		

		try:
			account_name = x['account_type']
		except:
			account_name = ''

 		remove_splash_invoice_no = x['invoice_no'].replace('/', '_')
 		if proj_name=='CHM':
 			extension_in_html 	  = remove_splash_invoice_no+'_'+x['ship_name']+'_'+x['voyage_no']+'_'+client_name+".html"
 			extension_in_pdf  	  = remove_splash_invoice_no+'_'+x['ship_name']+'_'+x['voyage_no']+'_'+client_name+".pdf"
 		else:
 			try:
 				extension_in_html = remove_splash_invoice_no+'_'+client_name+'_'+x['account_type']+".html"
 				extension_in_pdf  = remove_splash_invoice_no+'_'+client_name+'_'+x['account_type']+".pdf"
 			except:
 				extension_in_html = remove_splash_invoice_no+'_'+client_name+'_'+x['vessel_type']+".html"
 				extension_in_pdf  = remove_splash_invoice_no+'_'+client_name+'_'+x['vessel_type']+".pdf"
			
		try:
			clientaddress   = x['address'].split(',')
			split_address11 = clientaddress[0]
			split_address22 = clientaddress[1]
			split_address33 = clientaddress[2]
			split_address44 = clientaddress[3]
			split_address55 = clientaddress[4]
			split_address66 = clientaddress[5]
		except:
			split_address11 = split_address11
			split_address22 = split_address22
			split_address33 = split_address33
			split_address44 = split_address44
			split_address55 = split_address55
			split_address66 = split_address66

		


		for_invoice_dates = ''
		
		try:
			for_invoice_dates    = x['invoice_date']
		except:
			for_invoice_dates    = x['today']

		try:
			due_date = x['due_date']
		except:
			due_date = ''

		

		Html_file = open("/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_html),"w")		
		Html_file.write('<center>');		
		Html_file.write('<table style="height:1200px;width:100%; position: ; top: 0; bottom: 0; left: 0; right: 0;border-collapse: collapse;font-size:0px">');
		Html_file.write('<tbody>');
		Html_file.write('<style>');		
		Html_file.write('td{');
		Html_file.write('table { page-break-after: always !important; }');	
		Html_file.write('page-break-inside: avoid !important;');		
		Html_file.write('padding:2px;');
		Html_file.write('}');		
		Html_file.write('</style>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-top: none; border-left: none; border-right: none; width:432;"><img src="1.png"  width="280px";/></td>');
		
		if proj_name=='BOSS':
			logo = 'boss.png'
		elif proj_name=='CHM':
			logo = 'chm.png'
		
		Html_file.write('<td style="border-top: none; border-left: none; border-right: none; width: 266px;"><img style="float: right;" src="'+str(logo)+'" width="95" /></td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="padding: 10px; border-top:3px double #c4c3c2; border-left: 3px double #c4c3c2; border-right: 3px double #c4c3c2; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="1" width="100%" cellspacing="0" cellpadding="0" style="border:1px solid #c4c3c2">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		customer = "Customer"+"'s"
		Html_file.write('<td style="background: #CCCCCC;border:1px solid #c4c3c2" colspan="2"><span class="hd">'+customer+' name &amp; address:</span></td>');
		Html_file.write('<td rowspan="5" width="2%"  style="border:1px solid #c4c3c2">&nbsp;</td>');
		Html_file.write('<td style="background: #c4c3c2;border:1px solid #c4c3c2" width="22%"><span class="hd">Invoice No.</span>:</td>');
		Html_file.write('<td style="background: #CCCCCC;border:1px solid #c4c3c2" width="29%">'+x['invoice_no']+'</td>');  # 
		Html_file.write('</tr>');
		Html_file.write('<tr style="border:1px solid #c4c3c2">');
				
		Html_file.write('<td colspan="2" rowspan="2" align="left" valign="top"  style="border:1px solid #c4c3c2;border-collapse:collapse"><strong>'+str(split_address11)+'</strong><br>'+split_address22+'<br>'+split_address33+'<br>'+split_address44+'<br>'+split_address55+'<br>'+split_address66+'</td>');
		Html_file.write('<td style="border:1px solid #c4c3c2">Date:</td>');		
		try:
			if get_page=='Please see Page 1':
				gt_pge = for_invoice_dates #get_page
			else:
				gt_pge = datetime.strptime(get_page, "%Y-%m-%d").strftime('%d-%b-%Y')
		except:
			gt_pge = for_invoice_dates#'Please see Page 1'

		Html_file.write('<td style="border:1px solid #c4c3c2"><strong>'+str(gt_pge)+'</strong></td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border:1px solid #c4c3c2">Our Ref.:</td>');
		Html_file.write('<td style="border:1px solid #c4c3c2"><strong>'+x['invoice_no']+'</strong></td>'); #
		Html_file.write('</tr>');
		Html_file.write('<tr>');

		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and  currency=='INR':
			Html_file.write('<td width="21%" style="border:1px solid #c4c3c2">Customer GSTIN</td>');
			if editable!='edit':
				Html_file.write('<td width="28%" style="border:1px solid #c4c3c2">'+str(customer_no)+'</td>');
			else:
				Html_file.write('<td width="28%" style="border:1px solid #c4c3c2">'+str(customer_no)+'</td>');

			Html_file.write('<td style="border:1px solid #c4c3c2">Blue Water GSTIN:</td>');
			Html_file.write('<td  style="border:1px solid #c4c3c2"><strong>05AACCB9907G2ZQ</strong></td>');
			Html_file.write('</tr>');

		elif proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and  currency=='USD':
			print ''
		elif proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and  currency=='INR':
			Html_file.write('<td width="21%"  style="border:1px solid #c4c3c2">Customer GSTIN</td>');
			if editable!='edit':
				Html_file.write('<td width="28%"  style="border:1px solid #c4c3c2">'+str(customer_no)+'</td>');
			else:
				Html_file.write('<td width="28%"  style="border:1px solid #c4c3c2">'+str(customer_no)+'</td>');
			Html_file.write('<td style="border:1px solid #c4c3c2">Blue Water GSTIN:</td>');
			Html_file.write('<td style="border:1px solid #c4c3c2"><strong>05AACCB9907G2ZQ</strong></td>');
			Html_file.write('</tr>');

		elif proj_name=='BOSS' and inr_usd=='USD':
			Html_file.write('<td width="21%"><span class="style2">Customer GSTIN</span></td>');
			if editable!='edit':
				Html_file.write('<td width="28%">'+str(customer_no)+'</td>');
			else:
				Html_file.write('<td width="28%">'+str(customer_no)+'</td>');
			Html_file.write('<td><span class="style1">Blue Water GSTIN:</span></td>');
			Html_file.write('<td>05AACCB9907G2ZQ</td>');
			Html_file.write('</tr>');

		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #c4c3c2">Person Incharge:</td>');
		try:
			Html_file.write('<td  style="border:1px solid #c4c3c2"><strong>'+str(x['vm_name'])+'</strong></td>');
		except:
			Html_file.write('<td  style="border:1px solid #c4c3c2"><strong></strong></td>');
		
		try:
			voy_id = x['voyID']
		except:
			voy_id = ''
		
		merge_ship = str(ship_name) +',Voy: '+str(voy_no)		
		Html_file.write('<td class="hd"  style="border:1px solid #c4c3c2">Your Ref.:</td>');
		Html_file.write('<td style="border:1px solid #c4c3c2" ><b style="font-size:11px">'+str(merge_ship)+'</b><b style="color:blue;font-size:11px"> (BW ID*: '+voy_id+')</b></td>');
		Html_file.write('</tr>');
		
		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':
			print ''
		elif proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD':
			Html_file.write('<tr>');
			if proj_name=='BOSS':
				Html_file.write('<td class="hd" style="border:1px solid #c4c3c2">Remarks:</td>');
				try:
					Html_file.write('<td style="border:1px solid #c4c3c2;color:#028ffa;font-weight:bold">'+str(x['remarks'])+'</td>');
				except:
					Html_file.write('<td style="border:1px solid #c4c3c2;color:#028ffa;font-weight:bold"></td>');
				Html_file.write('<td style="border:1px solid #c4c3c2"></td>');
			else:
				Html_file.write('<td class="hd" style="border:1px solid #c4c3c2;">Remarks</td>');
				try:
					Html_file.write('<td style="border:1px solid #c4c3c2;font-size:12px">'+str(x['remarks'])+'</td>');
				except:
					Html_file.write('<td style="border:1px solid #c4c3c2">&nbsp;</td>');
					
				Html_file.write('<td style="border:1px solid #c4c3c2">&nbsp;</td>');
			if proj_name=='BOSS':
				Html_file.write('<td class="hd" style="border:1px solid #c4c3c2">Invoice Period:</td>');
			else:
				Html_file.write('<td class="hd"  style="border:1px solid #c4c3c2;font-size:13px">Disch Port, Est.Disch Date:</td>');


		if proj_name=='CHM' and currency=='USD':
			#try:
			disch_port 		= x['disch_port']
			for_disch_date 	= datetime.strptime(str(x['disch_date']), "%m/%d/%Y").strftime('%d-%b-%Y')
			#except:
			#	disch_port      = ''
			#	for_disch_date  = ''
			Html_file.write('<td style="border:1px solid #c4c3c2">'+str(disch_port)+', '+str(for_disch_date)+'</td>');
		elif proj_name=='BOSS' and currency=='USD':
			format_invoice_date = datetime.now().strftime("%d-%b-%Y")
			format_year 		= datetime.strptime(x['month'], "%Y-%m-%d").strftime('%Y')
			format_month 		= datetime.strptime(x['month'], "%Y-%m-%d").strftime('%b')
			int_month 			= datetime.strptime(x['month'], "%Y-%m-%d").strftime('%m')
			get_month 			= monthrange(int(format_year),int(int_month))[1]
			
			if get_period!='':
				remarkx = get_period  #'Sep 2022 to Dec 2022'
			else:
				remarkx = '01-'+str(format_month)+' to '+str(get_month)+'-'+str(format_month)+'-'+str(format_year)
			Html_file.write('<td style="border:1px solid #c4c3c2;font-weight:normal">'+str(remarkx)+'</td>');


		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="padding: 10px; border-left: 3px double #c4c3c2; border-right: 3px double #c4c3c2; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="1" width="100%" cellspacing="0" cellpadding="0" style="border-collapse: collapse;">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="background: #CCCCCC;border:1px solid #c4c3c2" colspan="2" align="center">Project Details</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" width="23%" style="border:1px solid #c4c3c2">Customer ID:</td>');
		service_type = 'Marine Services'
		span_type    = ''
		customerID   = ''

		if proj_name=='BOSS':
			if client_name=='Reliance':
				customerID   = 'RELIANCE/BW/BOSS  [Vendors/Business partner code : 3249511]'
				service_type = 'Other Professional, Technical And Business Services'
				span_type    = '[SAC Code: 998399]'
			elif client_name=='Shell':
				customerID   = str(client_name)+'/BW/'+str(proj_name)
				service_type = 'Marine Services'
				span_type    = ''
			elif client_name=='Apeejay':
				customerID   = str(client_name)+'/BW/'+str(proj_name)
				service_type = 'Other Professional, Technical And Business Services'
				span_type    = '[SAC Code: 998399]'				
			else:
				customerID = str(client_name)+'/BW/'+str(proj_name)

		elif proj_name=='CHM':
			if client_name=='Reliance':
				customerID   = 'RELIANCE/BW/CHM [Vendors/Business partner code : 3249511]'
				service_type = 'Other Professional, Technical And Business Services '
				span_type  	 = '[SAC Code: 998399]'			
			else:
				customerID   = str(client_name).upper()+'/BW/'+str(proj_name)
				service_type = 'Other Professional, Technical And Business Services'
				span_type    = ''


		if proj_name=='BOSS' and currency=='INR':
			displayz = ''
			col 	 = '7'
		
		elif proj_name=='CHM' and currency=='INR':
			displayz = ''
			col 	 = '7'
						
		elif proj_name=='BOSS' and currency=='USD':			
			displayz = 'none'
			col 	 = '6'
						
		elif proj_name=='CHM' and currency=='USD':
			displayz = 'none'
			col 	 = '6'
					
		Html_file.write('<td width="77%"  style="border:1px solid #c4c3c2">'+str(customerID)+'</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #c4c3c2">Service Name:</td>');
		if proj_name=='BOSS' and currency=='INR' or proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td style="border:1px solid #c4c3c2">Tech consultancy through BOSS</td>');
		elif  proj_name=='CHM' and currency=='INR' or proj_name=='CHM' and currency=='USD':
			if client_name=='Shell' or client_name=='Shell NWE':
				Html_file.write('<td style="border:1px solid #c4c3c2">Cargo Heating Management Services [Shell Contract No. DS65730] <b>'+str(account_name)+'</b> </td>')
			else:
				if client_name=='Ultranav':
					Html_file.write('<td style="border:1px solid #c4c3c2">Cargo Heating Management Services <b>['+str(x.account_type)+']</b></td>');
				else:
					Html_file.write('<td style="border:1px solid #c4c3c2">Cargo Heating Management Services</td>');

		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #c4c3c2 !important">Service Type:</td>');
		Html_file.write('<td style="border:1px solid #c4c3c2">'+str(service_type)+' <span class="style3"><b>'+str(span_type)+'</b></span></td>');
		Html_file.write('</tr>');

		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #c4c3c2 !important">Service Nature:</td>');
		Html_file.write('<td style="border:1px solid #c4c3c2">Data base, data processing charges</td>');
		Html_file.write('</tr>');

		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="padding: 10px; border-left: 3px double #c4c3c2; border-right: 3px double #c4c3c2; width: 698px;" colspan="2" align="center" valign="middle">');
		if client_name=='Ultranav' and proj_name=='BOSS':
			font_size = ''
		else:
			font_size = ''
		Html_file.write('<table border="0" width="100%" cellspacing="0" cellpadding="0" '+str(font_size)+'>');		
		Html_file.write('<tbody>');
		Html_file.write('<tr style="border-left: 1px solid #c4c3c2; border-top: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;">');		
		Html_file.write('</tr>');
		Html_file.write('<thead>');
		Html_file.write('<tr>');
		Html_file.write('<th style="background-color: #cccccc;border-left: 1px solid #c4c3c2; border-top: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 7%;" align="center" valign="middle;">S.No.</th>');
		Html_file.write('<th style="background-color: #cccccc;border-left: 1px solid #c4c3c2; border-top: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 36%;" align="center" valign="middle;">Service Details</th>');
		Html_file.write('<th style="background-color: #cccccc;border-left: 1px solid #c4c3c2; border-top: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 10%;" align="center" valign="middle;">Qty</th>');
		Html_file.write('<th style="background-color: #cccccc;border-left: 1px solid #c4c3c2; border-top: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 10%;" align="center" valign="middle;">Amount USD</th>');
		Html_file.write('<th style="background-color: #cccccc;width: 20.3102%;border-left: 1px solid #c4c3c2;border-bottom: 1px solid #c4c3c2;border-top: 1px solid #c4c3c2;" align="center" valign="middle;">Conv. Rate</th>');
		Html_file.write('<th width="14%" align="center" valign="middle" style="background-color: #cccccc;width: 20.3102%;border-left: 1px solid #c4c3c2;border-bottom: 1px solid #c4c3c2;border-top: 1px solid #c4c3c2;white-space:nowrap;">Amount '+e_types+'</th>');
		
		if proj_name=='CHM' or proj_name=='BOSS':
			if client_name=='Reliance':
				display = ''
			else:
				display = ''
		
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('</thead>');
		
		
		if currency=='USD':
			usd_tag = 'USD'
		else:
			usd_tag = ''

		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':		
			amount 	= x['price']*x['rate']*x['qty']
			amt 	= amount				
			
		
		elif proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD':			
			client_price = x['price']*x['rate']*x['qty']
			amount 		 = client_price
			amt 		 = amount

							
			try:
				client_del = models.Client.objects.filter(client_name=client_name,proj_name='CHM').first()				
				if client_del.client_name=='stena bulk':
					if check_vessel>=1 and check_vessel<=2:
						prices  = models.Client.objects.filter(price_type='1-2',proj_name='CHM',client_name='stena bulk').first()
						i_price = prices.price #_1600_____1-2'
					elif check_vessel>=3 and check_vessel<=5:
						prices  = models.Client.objects.filter(price_type='3-5',proj_name='CHM',client_name='stena bulk').first()
						i_price = prices.price # 1400 #______3-5'
					elif check_vessel>=6 and check_vessel<=8:
						prices  = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='stena bulk').first()
						i_price = prices.price # 1200#_______6-8'
					elif check_vessel>=8:
						prices  = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='stena bulk').first()
						i_price = prices.price  #1200
				elif client_del.client_name=='Stena Bulk Veg Oil':
					if check_vessel>=1 and check_vessel<=2:
						pricess = models.Client.objects.filter(price_type='1-2',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
						i_price = pricess.price # 1700 #______1-2'
					elif check_vessel>=3 and check_vessel<=5:
						pricess = models.Client.objects.filter(price_type='3-5',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
						i_price = pricess.price # 1500 # _____3-5'
					elif check_vessel>=6 and check_vessel<=8:
						pricess = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
						i_price = pricess.price # 1300# ______6-8'
					elif check_vessel>=8:
						pricess = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
						i_price = pricess.price  #1300

				elif client_del.client_name=='Ultranav':
					if check_vessel>=1 and check_vessel<=4:
						pricess 	 = models.Client.objects.filter(price_type='1-4',proj_name='CHM',client_name='Ultranav').first()
						i_price 	 = pricess.price # 1700 #______1-2'
						#price_typess = '1-4'
					elif check_vessel>=5 and check_vessel<=9:
						pricess 	 = models.Client.objects.filter(price_type='5-9',proj_name='CHM',client_name='Ultranav').first()
						i_price 	 = pricess.price # 1500 # _____3-5'
						#price_typess = '5-9'
					elif check_vessel>=10:
						pricess 	 = models.Client.objects.filter(price_type='ShortHaul',proj_name='CHM',client_name='Ultranav').first()
						i_price 	 = pricess.price # 1300# ______6-8'
						#price_typess = 'ShortHaul'

				elif client_del.client_name=='Litasco':			
					if check_vessel>=1 and check_vessel<=5:						
						pricess = models.Client.objects.filter(price_type='1-5',proj_name='CHM',client_name='Litasco').first()
						i_price = pricess.price # 1700 #______1-2'
						#price_typess = '1-5'
					elif check_vessel>=6 and check_vessel<=10:
						pricess = models.Client.objects.filter(price_type='6-10',proj_name='CHM',client_name='Litasco').first()
						i_price = pricess.price # 1500 # _____3-5'
						#price_typess = '6-10'
					elif check_vessel>=11:
						pricess = models.Client.objects.filter(price_type='11 and Up',proj_name='CHM',client_name='Litasco').first()
						i_price = pricess.price # 1300# ______6-8'
						#price_typess = '11 and Up'

				elif client_del.client_name=='Litasco_Dubai':			
					if check_vessel>=1 and check_vessel<=5:						
						pricess = models.Client.objects.filter(price_type='1-5',proj_name='CHM',client_name='Litasco_Dubai').first()
						i_price = pricess.price # 1700 #______1-2'
						price_typess = '1-5'
					elif check_vessel>=6 and check_vessel<=10:
						pricess = models.Client.objects.filter(price_type='6-10',proj_name='CHM',client_name='Litasco_Dubai').first()
						i_price = pricess.price # 1500 # _____3-5'
						price_typess = '6-10'
					elif check_vessel>=11:
						pricess = models.Client.objects.filter(price_type='11 and Up',proj_name='CHM',client_name='Litasco_Dubai').first()
						i_price = pricess.price # 1300# ______6-8'
						price_typess = '11 and Up'
				else:
					i_price = x['price']					

				client_price = i_price*x['qty']
				amount 		 = client_price
				amt 		 = amount
				total_amount = amount					
			except:				
				pass

		if proj_name=='BOSS':
			project = 'Tech consultancy through BOSS'
		elif proj_name=='CHM':
			project = 'Cargo Heating Management Services'


		amttt = 0
		total_sum = 0
		if proj_name=='CHM':
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 7%;"><center><strong>1</strong></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 36%;white-space:nowrap;"><strong>'+str(project)+'</strong></td>');
			#Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 16.9275%;white-space:nowrap;"><strong>'+x['ship_name']+'</strong></td>');
			
			client_del = models.Client.objects.filter(client_name=client_name,proj_name='CHM').first()
			if client_del.client_name=='stena bulk':
				if check_vessel>=1 and check_vessel<=2:
					prices  = models.Client.objects.filter(price_type='1-2',proj_name='CHM',client_name='stena bulk').first()
					i_price = prices.price #_1600_____1-2'
				elif check_vessel>=3 and check_vessel<=5:
					prices  = models.Client.objects.filter(price_type='3-5',proj_name='CHM',client_name='stena bulk').first()
					i_price = prices.price # 1400 #______3-5'
				elif check_vessel>=6 and check_vessel<=8:
					prices  = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='stena bulk').first()
					i_price = prices.price # 1200#_______6-8'
				elif check_vessel>=8:
					prices  = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='stena bulk').first()
					i_price = prices.price  #1200
			elif client_del.client_name=='Stena Bulk Veg Oil':
				if check_vessel>=1 and check_vessel<=2:
					pricess = models.Client.objects.filter(price_type='1-2',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
					i_price = pricess.price # 1700 #______1-2'
				elif check_vessel>=3 and check_vessel<=5:
					pricess = models.Client.objects.filter(price_type='3-5',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
					i_price = pricess.price # 1500 # _____3-5'
				elif check_vessel>=6 and check_vessel<=8:
					pricess = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
					i_price = pricess.price # 1300# ______6-8'						
				elif check_vessel>=8:
					pricess = models.Client.objects.filter(price_type='6-8',proj_name='CHM',client_name='Stena Bulk Veg Oil').first()
					i_price = pricess.price  #1300	

			elif client_del.client_name=='Ultranav':
				if check_vessel>=1 and check_vessel<=4:
					pricess 	 = models.Client.objects.filter(price_type='1-4',proj_name='CHM',client_name='Ultranav').first()
					i_price 	 = pricess.price # 1700 #______1-2'
					#price_typess = '1-4'
				elif check_vessel>=5 and check_vessel<=9:
					pricess 	 = models.Client.objects.filter(price_type='5-9',proj_name='CHM',client_name='Ultranav').first()
					i_price 	 = pricess.price # 1500 # _____3-5'
					#price_typess = '5-9'
				elif check_vessel>=10:
					pricess 	 = models.Client.objects.filter(price_type='ShortHaul',proj_name='CHM',client_name='Ultranav').first()
					i_price 	 = pricess.price # 1300# ______6-8'
					#price_typess = 'ShortHaul'

			elif client_del.client_name=='Litasco':					
				if check_vessel>=1 and check_vessel<=5:						
					pricess = models.Client.objects.filter(price_type='1-5',proj_name='CHM',client_name='Litasco').first()
					i_price = pricess.price # 1700 #______1-2'
					#price_typess = '1-5'
				elif check_vessel>=6 and check_vessel<=10:
					pricess = models.Client.objects.filter(price_type='6-10',proj_name='CHM',client_name='Litasco').first()
					i_price = pricess.price # 1500 # _____3-5'
					#price_typess = '6-10'
				elif check_vessel>=11:
					pricess = models.Client.objects.filter(price_type='11 and Up',proj_name='CHM',client_name='Litasco').first()
					i_price = pricess.price # 1300# ______6-8'
					#price_typess = '11 and Up'	

			elif client_del.client_name=='Litasco_Dubai':					
				if check_vessel>=1 and check_vessel<=5:						
					pricess = models.Client.objects.filter(price_type='1-5',proj_name='CHM',client_name='Litasco_Dubai').first()
					i_price = pricess.price # 1700 #______1-2'
					#price_typess = '1-5'
				elif check_vessel>=6 and check_vessel<=10:
					pricess = models.Client.objects.filter(price_type='6-10',proj_name='CHM',client_name='Litasco_Dubai').first()
					i_price = pricess.price # 1500 # _____3-5'
					#price_typess = '6-10'
				elif check_vessel>=11:
					pricess = models.Client.objects.filter(price_type='11 and Up',proj_name='CHM',client_name='Litasco_Dubai').first()
					i_price = pricess.price # 1300# ______6-8'
					#price_typess = '11 and Up'				
			else:
				i_price = x['price']	
			
			

			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 10%;"><center><strong>'+str(x['qty'])+'</strong></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; width: 17%;"><center><strong>'+str(float(x['price']))+'</strong></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;><center><strong>'+str(float(x['rate']))+'</strong></center></td>');
			Html_file.write('<td style="border-left: solid #c4c3c2 1px; border-top: solid #c4c3c2 0px; border-bottom: solid  #c4c3c2 1px; border-right: solid  #c4c3c2 1px;width: 17%;""><center><strong>'+str(x['rate'])+'</strong></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2;border-right: 1px solid #c4c3c2;width: 17%;border-top: 0px solid #c4c3c2;"><b style="white-space:nowrap"><center>'+str(float(round(currency_amount,3)))+'</center></b></td>');

			Html_file.write('</tr>');


			for c in range(1,10):
				Html_file.write('<tr>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 7%;white-space:nowrap">&nbsp;</td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 16.9275%;white-space:nowrap"></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 10%;"><center></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 17%;"><center></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 17%;"><center></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2;border-right: 1px solid #c4c3c2"><center></center></td>');
				if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':
					Html_file.write('<td style="border-top: solid #c4c3c2 1px;border-left: solid #c4c3c2 1px; border-right: solid #c4c3c2 1px;"><center></center></td>');
				else:
					print ''
				Html_file.write('</tr>');
			
			round_off 	= round(currency_amount,0)
			in_words 	= convert_money_to_text(round_off)

		
		total_taxable_amount = currency_amount
		


		if proj_name=='BOSS' and currency=='INR':
			a 			= 'style="width: 7%;"'
			b 			= 'style="width:36%;"'
			c 			= 'style="width:16.9275%"'
			d 			= 'style="width:10%;"'
			e 			= ''
			f 			= 'style="width:0.60423%;"'
			hide    	= ''
			colspan 	= 'colspan="2"'
			currency 	= 'INR'
			gst_label   = 'IGST Amount @18.0%'
			tot_invoice = 'Total Invoice Amount'
			taxable     = 'Taxable Amount (Rs.)'
		if proj_name=='CHM' and currency=='INR':
			a 			= 'style="width:7%;"'
			b 			= 'style="width:36%;"'
			c 			= 'style="width:16.9275%"'
			d 			= 'style="width:10%;"'
			e 			= 'style="display:''"'
			f 			= 'style="width:0.60423%;"'
			hide  		= ''
			colspan 	= 'colspan="2"'
			currency 	= 'INR'
			gst_label   = 'IGST Amount @18.0%'
			tot_invoice = 'Total Invoice Amount'
			taxable     = 'Taxable Amount (Rs.)'
		if proj_name=='BOSS' and currency=='USD':
			a 			= ''
			b 			= ''
			c 			= ''
			d 			= ''
			e 			= 'style="display:''"'
			f 			= ''
			hide 		= 'none'
			# colspan 	= 'colspan="2"'
			colspan 	= ''
			currency 	= 'USD'
			gst_label   = ''
			tot_invoice = ''
			taxable     = 'Total'


		if proj_name=='CHM' and currency=='USD':
			a 			= ''
			b 			= ''
			c 			= ''
			d 			= ''
			e 			= 'style="display:none"'
			f 			= ''
			hide 		= 'none'
			colspan 	= ''
			currency 	= 'USD'
			gst_label   = ''
			tot_invoice = ''
			taxable     = 'Total'

		Html_file.write('<tr style="text-align: right;">');
		Html_file.write('<td '+str(a)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');		

		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(c)+'>&nbsp;</td>');
		else:
			Html_file.write('<td '+str(c)+'>&nbsp;</td>');



		
		usd_tag = e_types		

		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2"><b>'+str(taxable)+'</b></td>');  # colspan="2" for CHM
		elif  proj_name=='CHM' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2"><b>'+str(taxable)+'</b></td>');  # colspan="2" for CHM
		elif proj_name=='CHM' and currency=='USD':
			Html_file.write('<td '+str(d)+' colspan="2"><b>'+str(taxable)+'</b></td>');  # colspan="2" for CHM
			#print '=======================',taxable,'=====',amttt
			Html_file.write('<td '+str(e)+'></td>');
		elif proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' '+str(colspan)+'><b>'+str(taxable)+'</b></td>');  # colspan="2" for CHM

		


		
		Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2;border-right: 1px solid #c4c3c2;border-top:1px solid #c4c3c2"><center><strong style="white-space:nowrap">'+str(usd_tag)+' '+str(float(round(currency_amount,2)))+'</strong></center></td>'); # aa
			
		Html_file.write('<td style="display:'+str(hide)+'">&nbsp;</td>');
		Html_file.write('</tr>');

		Html_file.write('<tr style="text-align: right;">');
		Html_file.write('<td '+str(a)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');
		Html_file.write('<td '+str(c)+'>&nbsp;</td>');

		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(gst_label)+'</td>'); # CHM colspan="2"
		elif  proj_name=='CHM' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(gst_label)+'</td>'); # CHM colspan="2"
		elif proj_name=='CHM' and currency=='USD':
			Html_file.write('<td '+str(d)+' colspan="2"></td>');  # colspan="2" for CHM
			Html_file.write('<td '+str(e)+'></td>');
		elif proj_name=='BOSS' and currency=='USD':

			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(gst_label)+'</td>');  # colspan="2" for CHM

		

		Html_file.write('<tr style="text-align: right;">');
		Html_file.write('<td '+str(a)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');
		Html_file.write('<td '+str(c)+'>&nbsp;</td>');
		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM
		if proj_name=='CHM' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM
		elif proj_name=='CHM' and currency=='USD':
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM
			Html_file.write('<td '+str(e)+'></td>');
		elif proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM

		
		Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2; border-right: 1px solid #c4c3c2;display:'+str(hide)+'"><strong>'+str(float(round(total_taxable_amount,0)))+'</strong></td>'); #net_taxable_amount
		Html_file.write('<td style="display:'+str(hide)+'">&nbsp;</td>');
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: 3px double #c4c3c2; border-right: 3px double #c4c3c2; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="0" width="98%" cellspacing="0" cellpadding="0">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		Html_file.write('<td style="; border-left: solid  #c4c3c2  1px; border-top: double #c4c3c2 1px;" width="86%">Total Invoice Amount Due (Rounded Off):</td>');
		Html_file.write('<td style="border-left: solid  #c4c3c2  1px; ; border-right: solid  #c4c3c2  1px; border-top: solid #c4c3c2 1px;" width="14%"><strong style="white-space:nowrap">'+str(usd_tag)+' '+str(float(round(currency_amount,0)))+'</strong></td>'); #amount
			
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="background: #CCCCCC; border: solid  #c4c3c2  1px;" colspan="2"> <strong style="text-transform: capitalize;">'+str(usd_tag)+' '+str(in_words)+'</strong></td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: solid  #c4c3c2  1px; border-bottom: solid  #c4c3c2  1px;" align="right" valign="middle">Payment Due Date</td>');
		try:
			if get_page=='Please see Page 1':
				get_pge = due_date #get_page
			else:
				get_pge = due_date
		except:
			get_pge = due_date

		Html_file.write('<td style="border-left: solid  #c4c3c2  1px; border-right: solid  #c4c3c2  1px; border-bottom: solid  #c4c3c2  1px;"><b style="font-size:14px">'+str(get_pge)+'</b></td>'); #'+str(due_date_format)+'
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: 3px double #c4c3c2; border-right: 3px double #c4c3c2; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="0" width="99%" cellspacing="0" cellpadding="0">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		Html_file.write('<td>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<td>Terms of payment:</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td>By wire transfer to our account "<b>BlueWater Trade Winds Pvt Ltd</b>" with-</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td>');
		Html_file.write('<table border="0" width="50%" cellspacing="0" cellpadding="0">');
		Html_file.write('<tbody>');
		
		if client_name=='Stena Bulk Veg Oil':
			#print '-------------->>',client_name
			if proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD' or proj_name=='BOSS' and currency=='INR':
				Html_file.write('<tr>');
				Html_file.write('<td><b>HDFC Bank</b></td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');
				Html_file.write('<td>56 Rajpur Road, Dehradun (UK), India</td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');
				Html_file.write('<td>SWIFT Code: <b>HDFCINBB</b></td>');
				Html_file.write('</tr>');

				if currency=='INR':
					Html_file.write('<tr>');
					Html_file.write('<td>RTGS/NEFT IFSC Code: <b>HDFC0000893</td>');
					Html_file.write('</tr>');
				else:
					pass
				
				Html_file.write('<tr>');
				if client_name=='Oldendorff':
					Html_file.write('<td>Account Number: <b>50200056415893</b></td>');
				else:
					Html_file.write('<td>Account Number: <b>02252560001213</b></td>');
				Html_file.write('</tr>');
		else:
			
			if proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD' or proj_name=='BOSS' and currency=='INR':
				Html_file.write('<tr>');
				Html_file.write('<td><b>HDFC Bank</b></td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');
				Html_file.write('<td>32, Arhat Bazar, Dehradun - 248001, Uttarakhand, INDIA</td>');
				Html_file.write('</tr>');				
				Html_file.write('<tr>');
				Html_file.write('<td>SWIFT Code: <b>HDFCINBB</b></td>');
				Html_file.write('</tr>');				
				if currency=='INR':
					Html_file.write('<tr>');
					Html_file.write('<td>RTGS/NEFT IFSC Code: <b>HDFC0000893</b></td>');
					Html_file.write('</tr>');	
				else:
					pass	

				Html_file.write('<tr>');
				Html_file.write('<td>Account Name: <b>BlueWater Trade Winds Pvt Ltd</b></td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');				
				if client_name=='Oldendorff':
					Html_file.write('<td>Account Number: <b>50200056415893</b></td>');
				else:
					Html_file.write('<td>Account Number: <b>02252560001213</b></td>');					

		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: 3px double #c4c3c2; border-right: 3px double #c4c3c2; border-bottom: 3px double #c4c3c2; width: 698px;" colspan="2" align="center" valign="middle">');
		
		Html_file.write('<table border="0" width="100%" cellspacing="0" cellpadding="0" style="margin-top:0px">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');

		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':
			Html_file.write('<td style="border: solid #999999 1px;">Note: GST rates in this invoice is based on current applicable rate. In case of revision of GST rates and policy in the current financial year, arrears arising due to such revision will be settled at the end of current financial year.</td>');
		elif proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td></td>');
		elif proj_name=='CHM' and currency=='USD':
			#Html_file.write('<td></td>');
			Html_file.write('<td><div style="font-size:13px;text-align:left;color:blue;width:140%">* BW ID is the Voyage ID created when vessel/voyage is setup in BlueWater system. Please use BW ID in correspondence along with invoice number.</div></td>');
		
		if proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD' or proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td align="center" width="52%"><img src="/var/www/html/invoice/it/static/pdf/stamp.png" style="margin-top: -224px;margin-left:215px" width="185" height="195">');
		else:
			Html_file.write('<td align="center" width="52%">');
		
		Html_file.write('');
		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':
			
			Html_file.write('');
		elif proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD':
			
			Html_file.write('');

		
		Html_file.write('');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</center>');

		if proj_name=='CHM':		
		 	Html_file.write('<div style="margin-top:10px;border-top:0px solid grey;"><center style="font-family: arial; font-size: 9pt;">BlueWater Trade Winds Pvt. Ltd.</center><center style="font-family: arial; font-size: 8pt;">4, Siddarth Enclave, GMS Road, Ballupur, Dehradun - 248001 Uttarkhand INDIA<br /> Tel:+91-135-2649301, 2649464 Corporate Email: accounts@bwesglobal.com Website:www.bwesglobal.com</center><div> <div style="padding:6px;margin-bottom:0px;display:''" colspan="6" align="right"><img src="rings.png" width="95" height="36" style="float:right;margin-bottom:50px;margin-top:-44px;;margin-right:-26px;margin-right:27px" /><br /><span style="font-family: Arial, Helvetica, sans-serif; font-size: 7pt;float:right;margin-top:-29px;margin-right:-119px;">Visit us at : www.bwesglobal.com</span></div>');
		else:
			Html_file.write('<div style="margin-top:10px;"><center style="font-family: arial; font-size: 9pt;">BlueWater Trade Winds Pvt. Ltd.</center><center style="font-family: arial; font-size: 8pt;">4, Siddarth Enclave, GMS Road, Ballupur, Dehradun - 248001 Uttarkhand INDIA<br /> Tel:+91-135-2649301, 2649464 Corporate Email: accounts@bwesglobal.com Website:www.bwesglobal.com</center><div> <div style="padding:6px;margin-bottom:0px;display:''" colspan="6" align="right"><img src="rings.png" width="95" height="36" style="float:right;margin-bottom:50px;margin-top:-44px;;margin-right:-26px;margin-right:27px" /><br /><span style="font-family: Arial, Helvetica, sans-serif; font-size: 7pt;float:right;margin-top:-29px;margin-right:-119px;">Visit us at : www.bwesglobal.com</span></div>');

		Html_file.close()

	 	html  		= "/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_html)
	 	pdf 		= "/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_pdf)	

	 	if proj_name=='BOSS':
			options = {
				'page-width'     			: '216mm',
				'page-height'     			: '279mm',
				'dpi'             			: '96',		
				'margin-top'				: '0.1cm',
				'margin-bottom'				: '1.1cm',
				'margin-left'				: '0.1cm',
				'margin-right' 				: '0.1cm',
				'encoding'					: "UTF-8",
				'footer-html'				: '/var/www/html/invoice/it/templates/footer.html',
				'no-outline'				: None,						
				'enable-local-file-access' 	: None,		

			}
		else:
			options = {
				'page-size' 				:'A4',	
				'enable-local-file-access' 	: None,		
				
			}

	 	pdfkit.from_file([html],pdf,options=options)
	 	path_save  			= models.invoice.objects.filter(invoice_no=x['invoice_no']).first()
		path_save.pdf_path 	= pdf
		path_save.save()
		username   	= "BWTW059"
		password   	= "sandeep@123"
		clientname  = "OHMSERVER"
		servername  = "OHMSERVER"
		domain 	   	= 'WORKGROUP'
		ipaddress  	= "172.16.5.100"
		conn 	   	= SMBConnection(username,password,clientname,servername,domain,use_ntlm_v2=True, sign_options=2, is_direct_tcp=True)
		conn.connect(ipaddress,445)
		Shares 		= conn.listShares()
		share_name  = Shares[0].name
		
		path_name  = 'Finance/Current/Finance/'+str(tag_name)+'/Online/'+str(client_name)+'/Online/'		
		sharedfiles = conn.listPath(share_name,path_name)	

		with open("/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_pdf), 'rb') as file:	
			conn.storeFile('Finance',"Finance/Current/Finance/"+str(tag_name)+"/Online/"+str(client_name)+"/Online/"+extension_in_pdf, file)	 
			print '---...4',client_name			
		conn.close()		
	return array
			
	
def generate_app_pdf(invoice_array,proj_name,client_name,currency,customer_no,check_vessel,currency_amount,e_types,invoice_no,sums):
	arrays 			= []
	from calendar import monthrange
	import calendar
	split_address6 	= ''
	

	if proj_name=='BOSS':
		tag_name = 'BIM'
		img_tag  = 'boss.png'
	elif proj_name=='CHM':
		tag_name = 'HIM'
		img_tag  = 'chm.png'
	
	dir      = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/'	
	
	if not os.path.exists(dir):
		os.makedirs(dir)
	srcfile      = '/var/www/html/invoice/it/static/pdf/1.png'
	dstroot      = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/1.png'	
	copyfile(srcfile, dstroot)
	srcfile1     = '/var/www/html/invoice/it/static/pdf/rings.png'
	dstroot1     = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/rings.png'	
	copyfile(srcfile1, dstroot1)
	srcfile2     = '/var/www/html/invoice/it/static/pdf/'+str(img_tag)
	dstroot2     = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/'+img_tag	
	copyfile(srcfile2, dstroot2)
	srcfile3     = '/var/www/html/invoice/it/static/pdf/'+str(img_tag)
	dstroot3     = '/var/www/html/invoice/it/static/pdf/'+str(tag_name)+'/'+client_name+'/Online/'+img_tag	
	copyfile(srcfile3, dstroot3)

	voy = {}
	amt = 0
	vnt = 0
	split_address11 = ''
	split_address22 = ''
	split_address33 = ''
	split_address44 = ''
	split_address55 = ''
	split_address66 = ''
	aa = 0
	tt = 0
	#print '--------',check_vessel
	check_vessel = 1

	remove_splash_invoice_no = invoice_no.replace('/', '_')
	extension_in_html = remove_splash_invoice_no+'_'+client_name+".html"
	extension_in_pdf  = remove_splash_invoice_no+'_'+client_name+".pdf"
	
 	
 	for x in invoice_array:
		try:
			clientaddress   = x['address'].split(',')			
			for e in range(0,len(x['address'].split('\n'))):
				split_address11 = clientaddress[e]
		# 	split_address11 = clientaddress[0]
		# 	split_address22 = clientaddress[1]
		# 	split_address33 = clientaddress[2]
		# 	split_address44 = clientaddress[3]
		# 	split_address55 = clientaddress[4]
		# 	split_address66 = clientaddress[5]
		except:
			pass
		#print '------>>>',split_address11
		# 	split_address11 = split_address11
		# 	split_address22 = split_address22
		# 	split_address33 = split_address33
		# 	split_address44 = split_address44
		# 	split_address55 = split_address55
		# 	split_address66 = split_address66

		for_invoice_dates = ''
		
		try:
			for_invoice_dates    = x['invoice_date']
		except:
			for_invoice_dates    = x['today']

		try:
			due_date = x['due_date']
		except:
			due_date = ''

		

		

		Html_file = open("/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_html),"w")		
		Html_file.write('<center>');		
		Html_file.write('<table style="height:1200px;width:100%; position: ; top: 0; bottom: 0; left: 0; right: 0;border-collapse: collapse;font-size:0px">');
		Html_file.write('<tbody>');
		Html_file.write('<style>');		
		Html_file.write('td{');
		Html_file.write('table { page-break-after: always !important; }');	
		Html_file.write('page-break-inside: avoid !important;');		
		Html_file.write('padding:2px;');
		Html_file.write('}');		
		Html_file.write('</style>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-top: none; border-left: none; border-right: none; width:432;"><img src="1.png"  width="280px";/></td>');
		
		if proj_name=='BOSS':
			logo = 'boss.png'
		elif proj_name=='CHM':
			logo = 'chm.png'

		Html_file.write('<td style="border-top: none; border-left: none; border-right: none; width: 266px;"><img style="float: right;" src="'+str(logo)+'" width="95" /></td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="padding: 10px; border-top:3px double #009CD8; border-left: 3px double #009CD8; border-right: 3px double #009CD8; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="1" width="100%" cellspacing="0" cellpadding="0" style="border:1px solid #009CD8">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		customer = "Customer"+"'s"
		Html_file.write('<td style="background: #009CD8;border:1px solid #009CD8;font-size:12px;color:white;font-weight:bold" colspan="2"><span class="hd">'+customer+' name &amp; address:</span></td>');
		Html_file.write('<td rowspan="5" width="2%"  style="border:1px solid #009CD8">&nbsp;</td>');
		Html_file.write('<td style="background: #009CD8;border:1px solid #009CD8;font-size:12px;color:white;font-weight:bold" width="22%"><span class="hd">Invoice No.</span>:</td>');
		Html_file.write('<td style="background: #009CD8;border:1px solid #009CD8;font-size:12px;color:white;font-weight:bold" width="29%">'+x['invoice_no']+'</td>');  # 
		Html_file.write('</tr>');
		Html_file.write('<tr style="border:1px solid #009CD8">');
				
		Html_file.write('<td colspan="2" rowspan="2" align="left" valign="top"  style="border:1px solid #009CD8;font-size:12px"><strong>'+str(split_address11)+'</strong></td>');
		Html_file.write('<td style="border:1px solid #009CD8;font-size:12px">Date:</td>');	
		#print '------kitna aya',x['ship_name']
		try:
			if get_page=='Please see Page 1':
				gt_pge = for_invoice_dates #get_page
			else:
				gt_pge = datetime.strptime(get_page, "%Y-%m-%d").strftime('%d-%b-%Y')
		except:
			gt_pge = for_invoice_dates#'Please see Page 1'

		Html_file.write('<td style="border:1px solid #009CD8;font-size:12px"><strong>'+str(gt_pge)+'</strong></td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border:1px solid #009CD8;font-size:12px">Our Ref.:</td>');
		Html_file.write('<td style="border:1px solid #009CD8;font-size:12px"><strong>'+x['invoice_no']+'</strong></td>'); #
		Html_file.write('</tr>');
		Html_file.write('<tr>');

		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and  currency=='INR':
			Html_file.write('<td width="21%" style="border:1px solid #009CD8">Customer GSTIN</td>');
			if editable!='edit':
				Html_file.write('<td width="28%" style="border:1px solid #009CD8">'+str(customer_no)+'</td>');
			else:
				Html_file.write('<td width="28%" style="border:1px solid #009CD8">'+str(customer_no)+'</td>');

			Html_file.write('<td style="border:1px solid #009CD8">Blue Water GSTIN:</td>');
			Html_file.write('<td  style="border:1px solid #009CD8"><strong>05AACCB9907G2ZQ</strong></td>');
			Html_file.write('</tr>');

		elif proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and  currency=='USD':
			print ''
		elif proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and  currency=='INR':
			Html_file.write('<td width="21%"  style="border:1px solid #009CD8">Customer GSTIN</td>');
			if editable!='edit':
				Html_file.write('<td width="28%"  style="border:1px solid #009CD8">'+str(customer_no)+'</td>');
			else:
				Html_file.write('<td width="28%"  style="border:1px solid #009CD8;font-size:12px">'+str(customer_no)+'</td>');
			Html_file.write('<td style="border:1px solid #009CD8">Blue Water GSTIN:</td>');
			Html_file.write('<td style="border:1px solid #009CD8"><strong>05AACCB9907G2ZQ</strong></td>');
			Html_file.write('</tr>');

		elif proj_name=='BOSS' and inr_usd=='USD':
			Html_file.write('<td width="21%"><span class="style2">Customer GSTIN</span></td>');
			if editable!='edit':
				Html_file.write('<td width="28%">'+str(customer_no)+'</td>');
			else:
				Html_file.write('<td width="28%">'+str(customer_no)+'</td>');
			Html_file.write('<td><span class="style1">Blue Water GSTIN:</span></td>');
			Html_file.write('<td>05AACCB9907G2ZQ</td>');
			Html_file.write('</tr>');

		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #009CD8;font-size:12px">Person Incharge:</td>');
		try:
			Html_file.write('<td  style="border:1px solid #009CD8;font-size:12px"><strong>'+str(x['vm_name'])+'</strong></td>');
		except:
			Html_file.write('<td  style="border:1px solid #009CD8"><strong></strong></td>');
		
		
		#merge_ship = str(x['ship_name']) +',Voy: '+str(x['voy_no'])		
		Html_file.write('<td class="hd"  style="border:1px solid #009CD8;font-size:12px">Your Ref.:</td>');
		Html_file.write('<td style="border:1px solid #009CD8" ></td>');
		Html_file.write('</tr>');
		#print '------kitna aya',x['ship_name']
		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':
			print ''
		elif proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD':
			Html_file.write('<tr>');
			if proj_name=='BOSS':
				Html_file.write('<td class="hd" style="border:1px solid #009CD8;font-size:12px">Remarks:</td>');
				try:
					Html_file.write('<td style="border:1px solid #009CD8;color:#028ffa;font-weight:bold">'+str(x['remarks'])+'</td>');
				except:
					Html_file.write('<td style="border:1px solid #009CD8;color:#028ffa;font-weight:bold"></td>');
				Html_file.write('<td style="border:1px solid #009CD8"></td>');
			else:
				Html_file.write('<td class="hd" style="border:1px solid #009CD8;font-size:12px">Remarks</td>');
				try:
					Html_file.write('<td style="border:1px solid #009CD8;font-size:12px">'+str(x['remarks'])+'</td>');
				except:
					Html_file.write('<td style="border:1px solid #009CD8">&nbsp;</td>');
					
				Html_file.write('<td style="border:1px solid #009CD8">&nbsp;</td>');
			#if proj_name=='BOSS':
			Html_file.write('<td class="hd" style="border:1px solid #009CD8;font-size:12px">Invoice Period:</td>');
			#else:
			#	Html_file.write('<td class="hd"  style="border:1px solid #009CD8">Disch Port,Disch Date:</td>');


		#if proj_name=='CHM' and currency=='USD':			
		#	disch_port 		= x['disch_port']
		#	for_disch_date 	= datetime.strptime(str(x['disch_date']), "%m/%d/%Y").strftime('%d-%b-%Y')			
		#	Html_file.write('<td style="border:1px solid #009CD8">'+str(disch_port)+', '+str(for_disch_date)+'</td>');
		#elif proj_name=='BOSS' and currency=='USD':
		format_invoice_date = datetime.now().strftime("%d-%b-%Y")
		format_year 		= datetime.strptime(x['month'], "%Y-%m-%d").strftime('%Y')
		format_month 		= datetime.strptime(x['month'], "%Y-%m-%d").strftime('%b')
		int_month 			= datetime.strptime(x['month'], "%Y-%m-%d").strftime('%m')
		get_month 			= monthrange(int(format_year),int(int_month))[1]
		
		# if get_period!='':
		# 	remarkx = get_period  #'Sep 2022 to Dec 2022'
		# else:
		remarkx = '01-'+str(format_month)+' to '+str(get_month)+'-'+str(format_month)+'-'+str(format_year)
		Html_file.write('<td style="border:1px solid #009CD8;font-weight:normal;font-size:12px">'+str(remarkx)+'</td>');


		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="padding: 10px; border-left: 3px double #009CD8; border-right: 3px double #009CD8; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="1" width="100%" cellspacing="0" cellpadding="0" style="border-collapse: collapse;">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="background: #009CD8;border:1px solid #009CD8;font-size:12px;color:white;font-weight:bold" colspan="2" align="center">Project Details</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" width="23%" style="border:1px solid #009CD8;font-size:12px;font-weight:bold;color:#009CD8">Customer ID:</td>');
		service_type = 'Marine Services'
		span_type    = ''
		customerID   = ''

		if proj_name=='BOSS':
			if client_name=='Reliance':
				customerID   = 'RELIANCE/BW/BOSS  [Vendors/Business partner code : 3249511]'
				service_type = 'Other Professional, Technical And Business Services'
				span_type    = '[SAC Code: 998399]'
			elif client_name=='Shell':
				customerID   = str(client_name)+'/BW/'+str(proj_name)
				service_type = 'Marine Services'
				span_type    = ''
			elif client_name=='Apeejay':
				customerID   = str(client_name)+'/BW/'+str(proj_name)
				service_type = 'Other Professional, Technical And Business Services'
				span_type    = '[SAC Code: 998399]'				
			else:
				customerID = str(client_name)+'/BW/'+str(proj_name)

		elif proj_name=='CHM':
			if client_name=='Reliance':
				customerID   = 'RELIANCE/BW/CHM [Vendors/Business partner code : 3249511]'
				service_type = 'Other Professional, Technical And Business Services '
				span_type  	 = '[SAC Code: 998399]'			
			else:
				customerID   = str(client_name).upper()+'/BW/'+str(proj_name)
				service_type = 'Other Professional, Technical And Business Services'
				span_type    = ''


		if proj_name=='BOSS' and currency=='INR':
			displayz = ''
			col 	 = '7'
		
		elif proj_name=='CHM' and currency=='INR':
			displayz = ''
			col 	 = '7'
						
		elif proj_name=='BOSS' and currency=='USD':			
			displayz = 'none'
			col 	 = '6'
						
		elif proj_name=='CHM' and currency=='USD':
			displayz = 'none'
			col 	 = '6'
					
		Html_file.write('<td width="77%"  style="border:1px solid #009CD8;font-size:12px">'+str(customerID)+'</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #009CD8;font-size:12px;font-weight:bold;color:#009CD8">Service Name:</td>');
		if proj_name=='BOSS' and currency=='INR' or proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td style="border:1px solid #009CD8">Tech consultancy through BOSS</td>');
		elif  proj_name=='CHM' and currency=='INR' or proj_name=='CHM' and currency=='USD':
			if client_name=='Shell' or client_name=='Shell NWE':
				Html_file.write('<td style="border:1px solid #009CD8">Cargo Heating Management Services [Shell Contract No. DS65730] <b>'+str(account_name)+'</b> </td>')
			else:
				if client_name=='Ultranav':
					Html_file.write('<td style="border:1px solid #009CD8;font-size:12px">Cargo Heating Management Services <b>['+str(x.account_type)+']</b></td>');
				else:
					Html_file.write('<td style="border:1px solid #009CD8;font-size:12px">Cargo Heating Management Services</td>');

		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #009CD8 !important;font-size:12px;font-weight:bold;color:#009CD8">Service Type:</td>');
		Html_file.write('<td style="border:1px solid #009CD8;font-size:12px">'+str(service_type)+' <span class="style3"><b>'+str(span_type)+'</b></span></td>');
		Html_file.write('</tr>');

		Html_file.write('<tr>');
		Html_file.write('<td class="hd" style="border:1px solid #009CD8 !important;font-size:12px;font-weight:bold;color:#009CD8">Service Nature:</td>');
		Html_file.write('<td style="border:1px solid #009CD8;font-size:12px">Data base, data processing charges</td>');
		Html_file.write('</tr>');

		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="padding: 10px; border-left: 3px double #009CD8; border-right: 3px double #009CD8; width: 698px;" colspan="2" align="center" valign="middle">');
		if client_name=='Ultranav' and proj_name=='BOSS':
			font_size = ''
		else:
			font_size = ''
		Html_file.write('<table border="0" width="100%" cellspacing="0" cellpadding="0" '+str(font_size)+'>');		
		Html_file.write('<tbody>');
		Html_file.write('<tr style="border-left: 1px solid #009CD8; border-top: 1px solid #009CD8; border-bottom: 1px solid #009CD8;">');		
		Html_file.write('</tr>');
		Html_file.write('<thead>');
		Html_file.write('<tr>');
		Html_file.write('<th style="background-color: #009CD8;border-left: 1px solid #009CD8; border-top: 1px solid #009CD8; border-bottom: 1px solid #009CD8;font-size:12px;color:white; width: 7%;" align="center" valign="middle;">S.No.</th>');
		Html_file.write('<th style="background-color: #009CD8;border-left: 1px solid #009CD8; border-top: 1px solid #009CD8; border-bottom: 1px solid #009CD8;font-size:12px;color:white; width: 10%;" align="center" valign="middle;">Vessel</th>');
		Html_file.write('<th style="background-color: #009CD8;border-left: 1px solid #009CD8; border-top: 1px solid #009CD8; border-bottom: 1px solid #009CD8;font-size:12px;color:white; width: 7%;" align="center" valign="middle;">Voy No.</th>');
		Html_file.write('<th style="background-color: #009CD8;border-left: 1px solid #009CD8; border-top: 1px solid #009CD8; border-bottom: 1px solid #009CD8;font-size:12px;color:white; width: 10%;" align="center" valign="middle;">Load.Port</th>');
		Html_file.write('<th style="background-color: #009CD8;width: 20.3102%;border-left: 1px solid #009CD8;border-bottom: 1px solid #009CD8;font-size:12px;color:white;border-top: 1px solid #009CD8;" align="center" valign="middle;">Date</th>');
		Html_file.write('<th style="background-color: #009CD8;width: 20.3102%;border-left: 1px solid #009CD8;border-bottom: 1px solid #009CD8;font-size:12px;color:white;border-top: 1px solid #009CD8;" align="center" valign="middle;">Disch.Port</th>');
		Html_file.write('<th style="background-color: #009CD8;width: 20.3102%;border-left: 1px solid #009CD8;border-bottom: 1px solid #009CD8;font-size:12px;color:white;border-top: 1px solid #009CD8;" align="center" valign="middle;">Date</th>');
		Html_file.write('<th style="background-color: #009CD8;width: 20.3102%;border-left: 1px solid #009CD8;border-bottom: 1px solid #009CD8;font-size:12px;color:white;border-top: 1px solid #009CD8;white-space:nowrap" align="center" valign="middle;">Qty.(Days)</th>');
		Html_file.write('<th style="background-color: #009CD8;width: 20.3102%;border-left: 1px solid #009CD8;border-bottom: 1px solid #009CD8;font-size:12px;color:white;border-top: 1px solid #009CD8;" align="center" valign="middle;">Rate</th>');
		Html_file.write('<th width="14%" align="center" valign="middle" style="background-color: #009CD8;width: 20.3102%;border-left: 1px solid #009CD8;font-size:12px;color:white;border-bottom: 1px solid #009CD8;border-top: 1px solid #009CD8;white-space:nowrap;">Amount</th>');
		
		if proj_name=='CHM' or proj_name=='BOSS':
			if client_name=='Reliance':
				display = ''
			else:
				display = ''
		
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('</thead>');
		
		
		if currency=='USD':
			usd_tag = 'USD'
		else:
			usd_tag = ''


		if proj_name=='BOSS':
			project = 'Tech consultancy through BOSS'
		elif proj_name=='CHM':
			project = ''


		amttt = 0
		total_sum = 0
		s_n=0
		for t in invoice_array:
			if proj_name=='CHM':

				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px; width: 7%;"><center><strong>'+str(s_n+1)+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px; width: 10%;white-space:nowrap;"><strong>'+(t['ship_name'])+'</strong></td>');			
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px;white-space:nowrap;"><center><strong>'+str(t['voy_age_no'])+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px;"><center><strong>'+str(t['load_port'])+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px;"><center><strong>'+str(t['load_date'])+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px;"><center><strong>'+str(t['disch_port'])+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px"><center><strong>'+str(t['disch_date'])+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px"><center><strong>'+str(t['qty'])+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px"><center><strong>'+str(t['price'])+'</strong></center></td>');
				Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px;border-right: 1px solid #c4c3c2;"><center><strong>'+str(float(t['qty']*t['price']))+'</strong></center></td>');
				Html_file.write('</tr>');
				s_n+=1
			
			
			round_off 	= round(sums,0)
			in_words 	= convert_money_to_text(round_off)

		
		total_taxable_amount = sums


		for c in range(1,9):
			Html_file.write('<tr>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 7%;white-space:nowrap">&nbsp;</td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 10%;white-space:nowrap"></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 10%;"><center></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 17%;"><center></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 17%;"><center></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2;"><center></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 17%;"><center></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 17%;"><center></center></td>');	
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom:1px solid #c4c3c2; width: 17%;"><center></center></td>');
			Html_file.write('<td style="border-left: 1px solid #c4c3c2; border-bottom: 1px solid #c4c3c2;font-size:12px;border-right: 1px solid #c4c3c2;"><center></center></td>');	
			Html_file.write('</tr>');
		
		


		if proj_name=='BOSS' and currency=='INR':
			a 			= 'style="width: 7%;"'
			b 			= 'style="width:36%;"'
			c 			= 'style="width:16.9275%"'
			d 			= 'style="width:10%;"'
			e 			= ''
			f 			= 'style="width:0.60423%;"'
			hide    	= ''
			colspan 	= 'colspan="2"'
			currency 	= 'INR'
			gst_label   = 'IGST Amount @18.0%'
			tot_invoice = 'Total Invoice Amount'
			taxable     = 'Taxable Amount (Rs.)'
		if proj_name=='CHM' and currency=='INR':
			a 			= 'style="width:7%;"'
			b 			= 'style="width:36%;"'
			c 			= 'style="width:16.9275%"'
			d 			= 'style="width:10%;"'
			e 			= 'style="display:''"'
			f 			= 'style="width:0.60423%;"'
			hide  		= ''
			colspan 	= 'colspan="2"'
			currency 	= 'INR'
			gst_label   = 'IGST Amount @18.0%'
			tot_invoice = 'Total Invoice Amount'
			taxable     = 'Taxable Amount (Rs.)'
		if proj_name=='BOSS' and currency=='USD':
			a 			= ''
			b 			= ''
			c 			= ''
			d 			= ''
			e 			= 'style="display:''"'
			f 			= ''
			hide 		= 'none'
			# colspan 	= 'colspan="2"'
			colspan 	= ''
			currency 	= 'USD'
			gst_label   = ''
			tot_invoice = ''
			taxable     = 'Total'


		if proj_name=='CHM' and currency=='USD':
			a 			= ''
			b 			= ''
			c 			= ''
			d 			= ''
			e 			= 'style="display:none"'
			f 			= ''
			hide 		= 'none'
			colspan 	= ''
			currency 	= 'USD'
			gst_label   = ''
			tot_invoice = ''
			taxable     = 'Total'

		Html_file.write('<tr style="text-align: right;">');
		Html_file.write('<td '+str(a)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');		
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');		
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');

		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(c)+'>&nbsp;</td>');
		else:
			Html_file.write('<td '+str(c)+'>&nbsp;</td>');
		
		usd_tag = e_types

		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2"><b style="font-size:12px;font-weight:bold">'+str(taxable)+'</b></td>');  # colspan="2" for CHM
		elif  proj_name=='CHM' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2"><b style="font-size:12px;font-weight:bold">'+str(taxable)+'</b></td>');  # colspan="2" for CHM
		elif proj_name=='CHM' and currency=='USD':
			Html_file.write('<td '+str(d)+' colspan="2"><b style="font-size:12px;font-weight:bold">'+str(taxable)+'</b></td>');  # colspan="2" for CHM
			#print '=======================',taxable,'=====',amttt
			Html_file.write('<td '+str(e)+'></td>');
		elif proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' '+str(colspan)+'><b style="font-size:12px;font-weight:bold">'+str(taxable)+'</b></td>');  # colspan="2" for CHM

		


		
		Html_file.write('<td style="border-left: 1px solid #CCCCCC; border-bottom:1px solid #CCCCCC;border-right: 1px solid #CCCCCC;border-top:1px solid #CCCCCC;font-size:12px"><center><strong style="white-space:nowrap">'+str(float(round(sums,2)))+'</strong></center></td>'); # aa
			
		Html_file.write('<td style="display:'+str(hide)+'">&nbsp;</td>');
		Html_file.write('</tr>');

		Html_file.write('<tr style="text-align: right;">');
		Html_file.write('<td '+str(a)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');
		Html_file.write('<td '+str(c)+'>&nbsp;</td>');

		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(gst_label)+'</td>'); # CHM colspan="2"
		elif  proj_name=='CHM' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(gst_label)+'</td>'); # CHM colspan="2"
		elif proj_name=='CHM' and currency=='USD':
			Html_file.write('<td '+str(d)+' colspan="2"></td>');  # colspan="2" for CHM
			Html_file.write('<td '+str(e)+'></td>');
		elif proj_name=='BOSS' and currency=='USD':

			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(gst_label)+'</td>');  # colspan="2" for CHM

		

		Html_file.write('<tr style="text-align: right;">');
		Html_file.write('<td '+str(a)+'>&nbsp;</td>');
		Html_file.write('<td '+str(b)+'>&nbsp;</td>');
		Html_file.write('<td '+str(c)+'>&nbsp;</td>');
		if proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM
		if proj_name=='CHM' and currency=='INR':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM
		elif proj_name=='CHM' and currency=='USD':
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM
			Html_file.write('<td '+str(e)+'></td>');
		elif proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td '+str(e)+'></td>');
			Html_file.write('<td '+str(d)+' colspan="2">'+str(tot_invoice)+'</td>'); # colspan="2" for CHM

		
		Html_file.write('<td style="border-left: 1px solid #009CD8; border-bottom: 1px solid #009CD8; border-right: 1px solid #009CD8;display:'+str(hide)+'"><strong>'+str(float(round(total_taxable_amount,0)))+'</strong></td>'); #net_taxable_amount
		Html_file.write('<td style="display:'+str(hide)+'">&nbsp;</td>');
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: 3px double #009CD8; border-right: 3px double #009CD8; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="0" width="98%" cellspacing="0" cellpadding="0">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		Html_file.write('<td style="; border-left: solid  #009CD8  1px; border-top: double #009CD8 1px;font-size:12px" width="86%">Total Invoice Amount Due (Rounded Off):</td>');
		Html_file.write('<td style="border-left: solid  #009CD8  1px; ; border-right: solid  #009CD8  1px; border-top: solid #009CD8 1px;font-size:12px" width="14%"><strong style="white-space:nowrap">USD '+str(float(round(sums,0)))+'</strong></td>'); #amount
			
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="background: #CCCCCC; border: solid  #009CD8  1px;font-size:12px" colspan="2"> <strong style="text-transform: capitalize;">USD '+str(in_words)+'</strong></td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: solid  #009CD8  1px; border-bottom: solid  #009CD8  1px;font-size:12px" align="right" valign="middle">Payment Due Date</td>');
		try:
			if get_page=='Please see Page 1':
				get_pge = due_date #get_page
			else:
				get_pge = due_date
		except:
			get_pge = due_date

		Html_file.write('<td style="border-left: solid  #009CD8  1px; border-right: solid  #009CD8  1px; border-bottom: solid  #009CD8  1px;"><b style="font-size:12px">'+str(get_pge)+'</b></td>'); #'+str(due_date_format)+'
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: 3px double #009CD8; border-right: 3px double #009CD8; width: 698px;" colspan="2" align="center" valign="middle">');
		Html_file.write('<table border="0" width="99%" cellspacing="0" cellpadding="0">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');
		Html_file.write('<td>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<td style="font-size:12px">Terms of payment:</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="font-size:12px">By wire transfer to our account "<b>BlueWater Trade Winds Pvt Ltd</b>" with-</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td>');
		Html_file.write('<table border="0" width="50%" cellspacing="0" cellpadding="0">');
		Html_file.write('<tbody>');
		
		if client_name=='Stena Bulk Veg Oil':
			#print '-------------->>',client_name
			if proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD' or proj_name=='BOSS' and currency=='INR':
				Html_file.write('<tr>');
				Html_file.write('<td><b>HDFC Bank</b></td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');
				Html_file.write('<td>56 Rajpur Road, Dehradun (UK), India</td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');
				Html_file.write('<td>SWIFT Code: <b>HDFCINBB</b></td>');
				Html_file.write('</tr>');

				if currency=='INR':
					Html_file.write('<tr>');
					Html_file.write('<td>RTGS/NEFT IFSC Code: <b>HDFC0000893</td>');
					Html_file.write('</tr>');
				else:
					pass
				
				Html_file.write('<tr>');
				if client_name=='Oldendorff':
					Html_file.write('<td>Account Number: <b>50200056415893</b></td>');
				else:
					Html_file.write('<td>Account Number: <b>02252560001213</b></td>');
				Html_file.write('</tr>');
		else:
			
			if proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD' or proj_name=='BOSS' and currency=='INR':
				Html_file.write('<tr>');
				Html_file.write('<td style="font-size:12px"><b>HDFC Bank</b></td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');
				Html_file.write('<td style="font-size:12px">32, Arhat Bazar, Dehradun - 248001, Uttarakhand, INDIA</td>');
				Html_file.write('</tr>');				
				Html_file.write('<tr>');
				Html_file.write('<td style="font-size:12px">SWIFT Code: <b>HDFCINBB</b></td>');
				Html_file.write('</tr>');				
				if currency=='INR':
					Html_file.write('<tr>');
					Html_file.write('<td style="font-size:12px">RTGS/NEFT IFSC Code: <b>HDFC0000893</b></td>');
					Html_file.write('</tr>');	
				else:
					pass	

				Html_file.write('<tr>');
				Html_file.write('<td style="font-size:12px">Account Name: <b>BlueWater Trade Winds Pvt Ltd</b></td>');
				Html_file.write('</tr>');
				Html_file.write('<tr>');				
				if client_name=='Oldendorff':
					Html_file.write('<td style="font-size:12px">Account Number: <b>50200056415893</b></td>');
				else:
					Html_file.write('<td style="font-size:12px">Account Number: <b>02252560001213</b></td>');					

		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('<tr>');
		Html_file.write('<td style="border-left: 3px double #009CD8; border-right: 3px double #009CD8; border-bottom: 3px double #009CD8; width: 698px;" colspan="2" align="center" valign="middle">');
		
		Html_file.write('<table border="0" width="100%" cellspacing="0" cellpadding="0" style="margin-top:0px">');
		Html_file.write('<tbody>');
		Html_file.write('<tr>');

		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':
			Html_file.write('<td style="border: solid #999999 1px;">Note: GST rates in this invoice is based on current applicable rate. In case of revision of GST rates and policy in the current financial year, arrears arising due to such revision will be settled at the end of current financial year.</td>');
		elif proj_name=='BOSS' and currency=='USD':
			Html_file.write('<td></td>');
		elif proj_name=='CHM' and currency=='USD':
			Html_file.write('<td></td>');
		
		if proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD' or proj_name=='BOSS' and currency=='INR':
			Html_file.write('<td align="center" width="52%"><img src="/var/www/html/invoice/it/static/pdf/stamp.png" style="margin-top: -224px;margin-left:215px" width="185" height="195">');
		else:
			Html_file.write('<td align="center" width="52%">');
		
		Html_file.write('');
		if proj_name=='BOSS' and currency=='INR' or proj_name=='CHM' and currency=='INR':
			
			Html_file.write('');
		elif proj_name=='BOSS' and currency=='USD' or proj_name=='CHM' and currency=='USD':
			
			Html_file.write('');

		
		Html_file.write('');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</td>');
		Html_file.write('</tr>');
		Html_file.write('</tbody>');
		Html_file.write('</table>');
		Html_file.write('</center>');

		if proj_name=='CHM':		
		 	Html_file.write('<div style="margin-top:10px;border-top:0px solid grey;"><center style="font-family: arial; font-size: 9pt;">BlueWater Trade Winds Pvt. Ltd.</center><center style="font-family: arial; font-size: 8pt;">4, Siddarth Enclave, GMS Road, Ballupur, Dehradun - 248001 Uttarkhand INDIA<br /> Tel:+91-135-2649301, 2649464 Corporate Email: info@bwesglobal.com Website:www.bwesglobal.com</center><div> <div style="padding:6px;margin-bottom:0px;display:''" colspan="6" align="right"><img src="rings.png" width="95" height="36" style="float:right;margin-bottom:50px;margin-top:-44px;;margin-right:-26px;margin-right:27px" /><br /><span style="font-family: Arial, Helvetica, sans-serif; font-size: 7pt;float:right;margin-top:-29px;margin-right:-119px;">Visit us at : www.bwesglobal.com</span></div>');
		else:
			Html_file.write('<div style="margin-top:10px;"><center style="font-family: arial; font-size: 9pt;">BlueWater Trade Winds Pvt. Ltd.</center><center style="font-family: arial; font-size: 8pt;">4, Siddarth Enclave, GMS Road, Ballupur, Dehradun - 248001 Uttarkhand INDIA<br /> Tel:+91-135-2649301, 2649464 Corporate Email: info@bwesglobal.com Website:www.bwesglobal.com</center><div> <div style="padding:6px;margin-bottom:0px;display:''" colspan="6" align="right"><img src="rings.png" width="95" height="36" style="float:right;margin-bottom:50px;margin-top:-44px;;margin-right:-26px;margin-right:27px" /><br /><span style="font-family: Arial, Helvetica, sans-serif; font-size: 7pt;float:right;margin-top:-29px;margin-right:-119px;">Visit us at : www.bwesglobal.com</span></div>');

		Html_file.close()

	 	html  		= "/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_html)
	 	pdf 		= "/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_pdf)	

	 	if proj_name=='BOSS':
			options = {
				'page-width'     			: '216mm',
				'page-height'     			: '279mm',
				'dpi'             			: '96',		
				'margin-top'				: '0.1cm',
				'margin-bottom'				: '1.1cm',
				'margin-left'				: '0.1cm',
				'margin-right' 				: '0.1cm',
				'encoding'					: "UTF-8",
				'footer-html'				: '/var/www/html/invoice/it/templates/footer.html',
				'no-outline'				: None,						
				'enable-local-file-access' 	: None,		

			}
		else:
			options = {
				'page-size' 				:'A4',	
				'enable-local-file-access' 	: None,		
				
			}

	 	pdfkit.from_file([html],pdf,options=options)
	 	path_save  			= models.invoice.objects.filter(invoice_no=x['invoice_no']).first()		
		path_save.pdf_path 	= pdf	
		path_save.save()	
		username   	= "BWTW059"
		password   	= "sandeep@123"
		clientname  = "OHMSERVER"
		servername  = "OHMSERVER"
		domain 	   	= 'WORKGROUP'
		ipaddress  	= "172.16.5.100"
		conn 	   	= SMBConnection(username,password,clientname,servername,domain,use_ntlm_v2=True, sign_options=2, is_direct_tcp=True)
		conn.connect(ipaddress,445)
		Shares 		= conn.listShares()
		share_name  = Shares[0].name
		
		path_name  = 'Finance/Current/Finance/'+str(tag_name)+'/Online/'+str(client_name)+'/Online/'		
		sharedfiles = conn.listPath(share_name,path_name)	

		with open("/var/www/html/invoice/it/static/pdf/"+str(tag_name)+"/"+str(client_name)+"/Online/"+str(extension_in_pdf), 'rb') as file:	
			conn.storeFile('Finance',"Finance/Current/Finance/"+str(tag_name)+"/Online/"+str(client_name)+"/Online/"+extension_in_pdf, file)	 
			print '---...4',client_name			
		conn.close()		

	return arrays




def convert_money_to_text(money):
    try:
        money_number = int(money)
    except ValueError:
        return "You must enter Number"
    if money_number == 0:
        return None
    positions 		= [None for i in range(4)]
    key_range 		= 0
    one_place 		= ["", "One ", "Two ", "Three ", "Four ","Five ", "Six ", "Seven ", "Eight ", "Nine "]
    one_ten_place   = ["Ten ", "Eleven ", "Twelve ", "Thirteen ", "Fourteen ","Fifteen ", "Sixteen ", "Seventeen ", "Eighteen ","Nineteen "]
    ten_place 		= ["Twenty ", "Thirty ", "Forty ", "Fifty ","Sixty ", "Seventy ", "Eighty ", "Ninety "]
    name_of_number  = ["Thousand ", "Lakh ", "Crore "]
    money_number_money_text = ''
    positions[0] = money_number % 1000  # units
    positions[1] = money_number / 1000
    positions[2] = money_number / 100000
    positions[1] = positions[1] - 100 * positions[2]  # thousands
    positions[3] = money_number / 10000000  # crores
    positions[2] = positions[2] - 100 * positions[3]  # lakhs
    for counter in range(3, 0, -1):
        if positions[counter] != 0:
            key_range = counter
            break
    for i in range(key_range, -1, -1):
        if positions[i] == 0:
            continue
        ones = positions[i] % 10  # ones
        tens = positions[i] / 10
        hundreds = positions[i] / 100  # hundreds
        tens = tens - 10 * hundreds  # tens
        if (hundreds > 0):
            money_number_money_text += one_place[hundreds] + "Hundred "
        if (ones > 0 or tens > 0):
            if (hundreds > 0 and i == 0):
                money_number_money_text += "and "
            if (tens == 0):
                money_number_money_text += one_place[ones]
            elif (tens == 1):
                money_number_money_text += one_ten_place[ones]
            else:
                money_number_money_text += ten_place[tens - 2] + \
                    one_place[ones]
        if (i != 0):
            money_number_money_text += name_of_number[i - 1]
    return money_number_money_text

def update_rate_for_pdf(request):
	e_rate 			= json.loads(request.GET['e_rate'])
	e_invoice_no 	= json.loads(request.GET['invoice_no'])[0]
	e_price 		= json.loads(request.GET['e_price'])
	e_remarks 		= json.loads(request.GET['e_remarks'])
	
	db 				= models.invoice.objects.filter(invoice_no=e_invoice_no).first()
	db.rate 		= e_rate
	db.price 		= e_price
	db.remark 		= e_remarks
	db.save()
	return HttpResponse(json.dumps('done'))


