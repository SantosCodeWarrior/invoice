from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection
##############################################
##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt

from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
import calendar

#####################################################
def get_fetch_api(request):
	if request.user.is_authenticated():
		get_client_details 	= models.Client.objects.filter(status=1,proj_name='BOSS').order_by('client_name')
		login_user 			= request.user
		now 				= datetime.now() - timedelta(days=30)		
		month_name 			= now.strftime('%B')
		gt_month 			= str(now).split('-')
		get_year 			= now.strftime('%Y')
		try:
			get_number 		= int(gt_month[1])+1 #for displaying 
			next_month 		= calendar.month_name[get_number]
		except:
			get_number 		= 01
			next_month 		= "Januray"

		cl_array 			= []
		todays  			= datetime.now().date()
		format_year 		= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%Y')
		format_month 		= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%b')
		int_month 			= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%m')
		try:
			get_month 		= monthrange(int(format_year),int(int_month)-1)[1]
		except calendar.IllegalMonthError:
			get_month 		= str(get_year)+'-02-01' 

		prev 				= str(int(int_month)-1)
		if prev<="9":
			prev_month = "0"+str(prev)
		else:
			prev_month = prev

		st_date = str(format_year)+'-'+(prev_month)+'-01'
		et_date = str(format_year)+'-'+(prev_month)+str(get_month)


		start_date_format 	= st_date
		end_date_format 	= et_date
		api_x 				= "https://bossv2.bwesglobal.com/api/get_static_details/"		
		api_method_x     	= "GET"			
		parameters_x     	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_date_format),'end_date': str(end_date_format)}
		response_x       	= requests.get(api_x, params=parameters_x,verify=False)		
		client_list_x    	= json.loads(response_x.content)
		cl_list 			= client_list_x['client_details']
		for sh in cl_list:  ## Live API		
			cl_array.append({
				'id' 		  	: sh,
				'client_name' 	: cl_list[sh]['name'],	
				'join_client_n' : str(cl_list[sh]['name'])+' ('+str(sh)+')'
			})

		context={
			'cl_array'   : cl_array,
			'login_user' : login_user,
			'next_month' : next_month,
			'month_name' : month_name
		}
		
		return render_to_response("invoice_display/get_fetch_api.html",context)
	else:
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
		return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def client_fetch_data(request):
	if request.user.is_authenticated():
		fetch_type 		   = request.POST['ft_data']
		client_id 		   = request.POST['client_id']
		get_ft_date 	   = request.POST['get_ft_date']		

		if client_id=='20011':
			get_client_Id = 135	
		
		elif client_id=='15576':
			get_client_Id= 113

		elif client_id=='8520':
			get_client_Id= 59

		elif client_id=='20002':
			get_client_Id = 136

		elif client_id=='15578':
			get_client_Id = 120

		elif client_id=='20007':
			get_client_Id = 131

		elif client_id=='15588':
			get_client_Id = 119

		elif client_id=='15586':
			get_client_Id = 127
		
		elif client_id=='42':
			get_client_Id == 17
		
		else:
			get_client_Id = client_id

		get_date 		   = get_ft_date #datetime.strptime(start_date,"%m/%d/%Y").strftime('%Y-%m-%d')	
		_client_id 	   	   = get_client_Id
		bill_type 		   = fetch_type
		
		api_url 		   = "https://bossv2.bwesglobal.com/api/get_ship_to_be_billed/"		
		api_method         = "POST"			
		parameters         = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'client_id': _client_id,'billing_type':bill_type,'billing_date':(get_date)}
		response           = requests.get(api_url, params=parameters,verify=False)		
		boss_array         = json.loads(response.content)
		voyage_array	   = []
		_start_date 	   = get_date
		
		for c in boss_array:			
			get_account 	= models.api_client_data.objects.filter(client_id=_client_id,ship_name=c['Ship Name']).first()			
			try:
				get_account = get_account.account_tab
			except:
				get_account = ""		
			
			if fetch_type=='monthly_prorated':
				voyage_array.append({
					'id'			: c['Client ID'],
					'ship_name' 	: c['Ship Name'],
					'no_of_day'		: c['No. of Days'],
					'start_date'	: c['Start Date'],
					'end_date'  	: c['End Date'],
					'first_port'	: c['First Port'],
					'last_port'		: c['Last Port'],
					'account'		: c['voyage__priority'],
					'client'		: c['Client Name'],
					'voyage_no'     : c['voyage__voyage_no'],
					'report_id'		: '',
					})
			
			if fetch_type=='monthly_flat':
				voyage_array.append({
					'id'			: c['Client ID'],
					'ship_name' 	: c['Ship Name'],
					'no_of_day'		: 0,
					'start_date'	: 0,
					'end_date'  	: 0,
					'first_port'	: 0,
					'last_port'		: 0,
					'account'		: get_account,
					'client'		: c['Client Name'],
					'report_id'		: '',
					'voyage_no'     : '',
					})

			if fetch_type=='voyage_eop':
				voyage_array.append({
					'id'			: c['Client ID'],
					'client'		: c['Client Name'],
					'account'		: get_account,
					'end_date'  	: c['End Date'],
					'no_of_day'		: c['No. of Days'],
					'last_port'		: c['Last Port'],
					'ship_name' 	: c['Ship Name'],
					'voyage_no'     : '',
					'report_id'		: c['voyage_id'],
					'start_date'	: c['Start Date'],
					'first_port'	: c['First Port'],
					})

		context={
			'voyage_array' : voyage_array,
		}		
		return HttpResponse(json.dumps(context))
	else:
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
		return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def export_fetch_api_data(request):
	if request.user.is_authenticated():
		get_csv 	  	= json.loads(request.POST['csv_data'])
		gt_ft_date 	  	= json.loads(request.POST['gt_ft_date'])
		select_dir    	= os.path.dirname(__file__)
		srcfile       	= select_dir+'/static/API DATA.xlsx'
		dstroot       	= select_dir+'/API DATA.xlsx'	 
		copyfile(srcfile,dstroot)
		my_dir 	      	= os.path.dirname(__file__)		
		location      	= (my_dir+'/API DATA.xlsx')
		wb 	   	      	= load_workbook(my_dir+'/API DATA.xlsx')
		ws_usd        	= wb.get_sheet_by_name("Sheet1")
		thin_border   	= Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style     	= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
		thin_border   	= Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style      	= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
		_row_style    	= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
		dt_split 	  	= gt_ft_date.split('-')
		get_splt_date 	= dt_split[1]
		get_splt_year   = dt_split[0]
		get_month 		= int(get_splt_date) #int(get_splt_date)-1
		if get_month<=9:
			gt_mnth = "0"+str(get_month)
		else:
			gt_mnth = get_month

		get_prev_month 	= str(get_splt_year)+'-'+str(gt_mnth)+'-01'
		get_month 	  	= datetime.strptime(str(get_prev_month), "%Y-%m-%d").strftime('%b %Y')
		ws_usd 		  	= wb.get_sheet_by_name('Sheet1')

		if get_csv[0]['client']=='Poompuhar Shipping Corporation Limited':
			ws_usd.title = 'Poompuhar'+' ('+str(get_month)+')'		
		elif get_csv[0]['client']=='Adhart Shipping Pte. Ltd':
			ws_usd.title = 'Adhart'+' ('+str(get_month)+')'		
		elif get_csv[0]['client']=='Tankers International':
			ws_usd.title = 'Sheet2'		
		else:
			ws_usd.title = get_csv[0]['client']

		ii   		  = 2
		jj 	 		  = 1
		current_date  = datetime.now().date()
		for x in get_csv:
			ws_usd.cell('A'+str(ii)).value = jj
			ws_usd.cell('B'+str(ii)).value = x['ship_name']
			ws_usd.cell('C'+str(ii)).value = x['voyage_no']
			ws_usd.cell('D'+str(ii)).value = x['no_of_day']
			ws_usd.cell('E'+str(ii)).value = x['first_port']
			ws_usd.cell('F'+str(ii)).value = x['start_date']		
			ws_usd.cell('G'+str(ii)).value = x['last_port']
			ws_usd.cell('H'+str(ii)).value = x['end_date']
			ws_usd.cell('I'+str(ii)).value = x['account']
			ws_usd.cell('J'+str(ii)).value = x['report_id']			
			ws_usd.cell('A'+str(ii)).style = _row_style	 
			ws_usd.cell('B'+str(ii)).style = _row_style	 
			ws_usd.cell('C'+str(ii)).style = _row_style	 
			ws_usd.cell('D'+str(ii)).style = _row_style	 
			ws_usd.cell('E'+str(ii)).style = _row_style	 
			ws_usd.cell('F'+str(ii)).style = _row_style		
			ws_usd.cell('G'+str(ii)).style = _row_style	
			ws_usd.cell('H'+str(ii)).style = _row_style	 
			ws_usd.cell('I'+str(ii)).style = _row_style	 
			ws_usd.cell('J'+str(ii)).style = _row_style	 			
			ii+=1
			jj+=1				
			wb.save(my_dir+'/API DATA'+str(current_date)+'.xlsx')
		return HttpResponse(json.dumps('done'))
	else:		
		return HttpResponseRedirect('/it/user_login')

def export_api_data(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	current_date  					= datetime.now().date()	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/API DATA'+str(current_date)+'.xlsx')
	filename 	   					= my_dir+'/API DATA'+str(current_date)+'.xlsx'
	download_name  					= "API DATA"+str(current_date)+".xlsx"	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response

@csrf_exempt
def save_client_fetch_data(request):
	todays  			= datetime.now().date()
	format_year 		= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%Y')
	format_month 		= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%b')
	int_month 			= datetime.strptime(str(todays), "%Y-%m-%d").strftime('%m')
	get_month 			= monthrange(int(format_year),int(int_month)-1)[1]
	prev 				= str(int(int_month)-1)
	if prev<="9":
		prev_month = "0"+str(prev)
	else:
		prev_month = prev
	st_date 			= str(format_year)+'-'+(prev_month)+'-01'
	et_date 			= str(format_year)+'-'+(prev_month)+str(get_month)
	start_date_format 	= st_date
	end_date_format 	= et_date
	api_urlx 			= "https://bossv2.bwesglobal.com/api/get_static_details/"		
	api_methodx     	= "GET"			
	parametersx     	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_date_format),'end_date': str(end_date_format)}
	responsex       	= requests.get(api_urlx, params=parametersx,verify=False)		
	client_listx    	= json.loads(responsex.content)
	cl_dets 			= client_listx['ship_details']
	error 				= 'Done'
	for key in cl_dets:
		if cl_dets[key]["clients_id"][0]==63:
			account_tab = cl_dets[key]["account_tab"]				
		else:
			account_tab = cl_dets[key]["account_tab"]
		patn 			= re.sub(r"[\([{})\]]", "", str(cl_dets[key]["clients_id"]))			
		ship_id     	= cl_dets[key]["id"]
		client_id   	= patn	
		ship_name   	= cl_dets[key]["ship_name"]		
		db1				= models.api_client_data()
		db1.account_tab = account_tab			 				 	
		db1.ship_name   = ship_name
		db1.address 	= cl_dets[key]["address"]
		db1.client_id 	= str(patn)
		db1.ship_id     = ship_id	
		db1.curr_date 	= datetime.now().date()	 	
		#db1.save()
	return HttpResponse(json.dumps('done'))