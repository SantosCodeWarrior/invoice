from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

from openpyxl.styles import PatternFill
#####################################################


@csrf_exempt
def export_summary_report(request):
	track_details = json.loads(request.POST['track_details'])	
	client_name   = json.loads(request.POST['client_name'])
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/summary_report.xlsx'
	dstroot       = select_dir+'/summary_report.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/summary_report.xlsx')
	wb 	   	      = load_workbook(my_dir+'/summary_report.xlsx')
	ws     	      = wb.get_sheet_by_name("Sheet1")	
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	ii   		  = 7
	jj 	 		  = 1
	single 		  = []
	tt     		  = 0
	vt 			  = 0	
	e 			  = 1
	
	
	get_ship_details = models.vessel_billing_day.objects.all().order_by('ship_name')
	for c in get_ship_details:	
	 	thin_border			= Border(left=Side(style='thin',color='808080'),right=Side(style='thin',color='808080'),top=Side(style='thin',color='808080'),bottom=Side(style='thin',color='808080'))
	 	my_style			= Style(font=Font(name='Arial', size=8, bold=False),border=thin_border)
	 	_row_style			= Style(font=Font(name='Arial', size=8, bold=False),border=thin_border)
	 	_styles2 			= Style(font=Font(name='Arial', size=8, italic=False,bold=False),alignment=Alignment(horizontal='center',vertical='center'),border=thin_border)
		get_total_day 		= models.merge_billing_day.objects.filter(ship_name=c.ship_name).first()
		
		_styles 			= Style(font=Font(name='Arial', size=8, bold=False),fill=PatternFill(patternType='solid', fgColor=Color(rgb='ebebe0')),border=thin_border)
		_styles3 			= Style(font=Font(name='Arial', size=8, bold=False),fill=PatternFill(patternType='solid', fgColor=Color(rgb='ebebe0')),alignment=Alignment(horizontal='center',vertical='center'),border=thin_border)
		
		if c.status!=None:
	 		ws.cell('B'+str(ii)).value 	= c.ship_name
	 		addx1 						= int(ii)+int(c.status)	
	 		addx2 						= int(ii)+int(c.status)	
	 		addx3 						= int(ii)+int(c.status)			
	 		merge1 						= ('B'+str(ii)+':B'+str(addx1))	
	 		merge2 						= ('A'+str(ii)+':A'+str(addx2))	
	 		merge3 						= ('H'+str(ii)+':H'+str(addx3))		 	 		
			ws.merge_cells(merge1)
			ws.merge_cells(merge2)
			ws.merge_cells(merge3)
			ws.cell('A'+str(ii)).value  = e
			ws.cell('H'+str(ii)).value = get_total_day.no_of_day
			e+=1
	 	else:
	 		ws.cell('B'+str(ii)).value  = ""
	 		ws.cell('A'+str(ii)).value  = ""	
	 		ws.cell('H'+str(ii)).value  = ""

 		
	 	
	
		ws.cell('C'+str(ii)).value = c.first_port
		ws.cell('D'+str(ii)).value = c.last_port
		ws.cell('E'+str(ii)).value = int(c.no_of_day)
		ws.cell('F'+str(ii)).value = str(c.start_date)				
	 	ws.cell('G'+str(ii)).value = str(c.end_Date)
	 	if e%2==1:
			ws.cell('A'+str(ii)).style 	= _styles3
			ws.cell('B'+str(ii)).style 	= _styles3
			ws.cell('C'+str(ii)).style 	= _styles
			ws.cell('D'+str(ii)).style 	= _styles
			ws.cell('E'+str(ii)).style 	= _styles3
			ws.cell('F'+str(ii)).style 	= _styles3
			ws.cell('G'+str(ii)).style 	= _styles3
			ws.cell('A'+str(ii)).style 	= _styles3
			ws.cell('B'+str(ii)).style 	= _styles3
			ws.cell('H'+str(ii)).style 	= _styles3
		else:
			ws.cell('A'+str(ii)).style 	= _row_style
			ws.cell('B'+str(ii)).style 	= _row_style
			ws.cell('C'+str(ii)).style 	= _row_style
			ws.cell('D'+str(ii)).style 	= _row_style
			ws.cell('E'+str(ii)).style 	= _styles2
			ws.cell('F'+str(ii)).style 	= _styles2
			ws.cell('G'+str(ii)).style 	= _styles2
			ws.cell('A'+str(ii)).style 	= _styles2
			ws.cell('B'+str(ii)).style 	= _styles2
			ws.cell('H'+str(ii)).style 	= _styles2
	

		ii+=1
		jj+=1
		tt+=1
		
		vt+=1	
		ws.cell('A7').value  = 1
		wb.save(my_dir+'/summary_report.xlsx')

	return HttpResponse(json.dumps('done'))

def export_summary_invoice(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)
	location 	   					= (my_dir+'/summary_report.xlsx')
	filename 	   					= my_dir+'/summary_report.xlsx'
	download_name  					= "Summary Report.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response