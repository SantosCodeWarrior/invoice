from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login

import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection

from it import apj_template as apj
from it import clearlake_template as ct
from it import shell_template as sh
from it import vessel_template as vt

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler
import json
from operator import itemgetter

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
#####################################################


def get_api_from_boss(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		name 			   = ''
		ship_type 		   = ''
		start_date_format  = ''
		end_date_format    = ''
		login_user 		   = request.user	
		api_url 		   = "https://bossv2.bwesglobal.com/api/get_data_boss/"	
		#api_url 	 	   = "http://0.0.0.0:8004/api/get_data_boss/"
		api_method         = "GET"			
		parameters         = {'key':"938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': name,'ship_type':ship_type,'start_date':str(start_date_format),'end_date': str(end_date_format)}
		response           = requests.get(api_url,params=parameters,verify=False)
		boss_array     	   = json.loads(response.content)
		client_array       = boss_array['cl_array']
		client_list 	   = []

		for cl in client_array:
			client_list.append({
				'clID'			: cl['id'],
				'client_name'	: cl['client_name'],			
				})

		client_wise = sorted(client_list, key=itemgetter('client_name'))

		context={
		'client_list' : client_wise,
		'login_user'  : login_user
		}
		return render_to_response("invoice_display/front_api.html",context)
	else:
		return HttpResponseRedirect('/it/user_login')

def get_vessel_list(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':	
		name 			   = request.GET['clientID']
		ship_type 		   = request.GET['acc_list']
		start_date_format  = ''
		end_date_format    = ''	
		start_date 		   = request.GET['fin_date']
		month_name         = datetime.strptime(start_date,"%Y-%m-%d").strftime('%b')
		year_name 		   = datetime.strptime(start_date,"%Y-%m-%d").strftime('%Y')
		month_number       = datetime.strptime(start_date,"%Y-%m-%d").strftime('%m')	
		start_format 	   = datetime.strptime(start_date,"%Y-%m-%d").strftime('%B')
		start_date_format  = start_date
		get_month 	 	   = monthrange(int(year_name),int(month_number))[1]
		end_date_format    = str(year_name)+'-'+str(month_number)+'-'+str(get_month)
		
		
		api_url 		   = "https://bossv2.bwesglobal.com/api/get_data_boss/"	
		#api_url 	 	   = "http://0.0.0.0:8004/api/get_data_boss/"
		api_method         = "GET"			
		parameters         = {'key':"938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': name,'ship_type':ship_type,'start_date':str(start_date_format),'end_date': str(end_date_format)}
		response           = requests.get(api_url,params=parameters,verify=False)
		boss_array     	   = json.loads(response.content)
		client_array       = boss_array['cl_array']	
		i = 0
		
		for cl in client_array:	
		 	if cl['client_name']==name:
		 		api_url1 	 = "https://bossv2.bwesglobal.com/api/get_data_boss/"	
				#api_url1 	 	   = "http://0.0.0.0:8004/api/get_data_boss/"
				api_method1  = "GET"			
				parameters1  = {'key':"938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': name,'ship_type':ship_type,'start_date':str(start_date_format),'end_date': str(end_date_format)}
				response1    = requests.get(api_url1,params=parameters1,verify=False)
				boss_array1  = json.loads(response.content)			
				all_array 	 = boss_array1['all_array']
				acc_ae 		 = []
				for c in boss_array1['all_array']:
					acc_ae.append({
						'ship_name' : c['ship_name'],
						'acc_tab'   : c['account_tab'],					
					})	

		return HttpResponse(json.dumps(acc_ae))
	else:
		return HttpResponseRedirect('/it/user_login')
	
def get_account_list(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':	
		cl_name   = request.GET['cl_namez']
		cl_tab    = models.pool_master.objects.filter(client__client_name=cl_name)
		acc_list  = []
		
		for c in cl_tab:
			acc_list.append({
				'id'          : c.id,
				'account_tab' : c.pool,			
				})	
		
		context={
			'acc_list' :  acc_list
		}
		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')



@csrf_exempt
def export_from_api(request):
	track_details = json.loads(request.POST['handsondata'])	
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/CLIENT-EXPORT.xlsx'
	dstroot       = select_dir+'/CLIENT-EXPORT.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/CLIENT-EXPORT.xlsx')
	wb 	   	      = load_workbook(my_dir+'/CLIENT-EXPORT.xlsx')
	ws     	      = wb.get_sheet_by_name("Sheet1")	
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	ii   		  = 2
	jj 	 		  = 1
	
	for t in track_details:
	  	thin_border    				   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style       				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
		_row_style     				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)	
		ws.cell('A'+str(ii)).value = jj
		ws.cell('B'+str(ii)).value = t['ship_name']
		try:
			acc_tab = t['acc_tab']
		except:
			acc_tab = ''
		ws.cell('C'+str(ii)).value = str(acc_tab)
		ws.cell('A'+str(ii)).style = _row_style	 
		ws.cell('B'+str(ii)).style = _row_style	 	
		ws.cell('C'+str(ii)).style = _row_style		
	 	ii+=1
	 	jj+=1
			
	 	wb.save(my_dir+'/CLIENT-EXPORT.xlsx')
	return HttpResponse(json.dumps('done'))

def api_invoice_export(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)
	location 	   					= (my_dir+'/CLIENT-EXPORT.xlsx')
	filename 	   					= my_dir+'/CLIENT-EXPORT.xlsx'
	download_name  					= "CLIENT-EXPORT.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response