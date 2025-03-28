from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

#####################################################


def chm_entry(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':		
		login_user = request.user	
		context={
		'login_user' :  login_user
		}			
		return render_to_response("invoice_display/chm_entry.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')

@csrf_exempt
def submit_chm_invoice_alag(request):
	table_details = json.loads(request.POST['table_details'])
	invoice_nox	  = request.POST['invoice_no']
	proj_name	  = request.POST['proj_name']
	curr_type 	  = request.POST['curr_type']
	vm_name       = request.POST['vm_name']
	in_date 	  = request.POST['in_date']
	cl_name       = request.POST['cl_name']
	gtin_no       = request.POST['gtin_no']
	address 	  = request.POST['address']
	pool_name     = request.POST['pool_name']
	period_dat 	  = request.POST['period_dat']
	subtotal 	  = request.POST['subtotal']	

	invoice_date  = datetime.strptime(in_date, "%m/%d/%Y").strftime('%Y-%m-%d')	
	period_date   = datetime.strptime(period_dat, "%m/%d/%Y").strftime('%Y-%m-%d')			
	s     		  = 0
	check_client  = models.Client.objects.filter(client_name=cl_name,currency_type=curr_type,proj_name=proj_name).count()		
	if check_client!=0:
		db_client =  models.Client.objects.filter(client_name=cl_name,currency_type=curr_type,proj_name=proj_name).first()
	else:
		db_client =  models.Client()

	db_client.client_name   = cl_name
	db_client.proj_name     = proj_name
	db_client.currency_type = curr_type
	db_client.rate 			= 1
	db_client.price 		= 1
	db_client.price_type    = None
	db_client.duration_type = 'Voyagewise'
	db_client.vm_name       = cl_name
	db_client.tin_number    = gtin_no		
	db_client.status 		= 1
	db_client.save()

	check_dbx = models.pool_master.objects.filter(client_id=db_client.id,pool=cl_name).count()
	if check_dbx!=0:
		save_db = models.pool_master.objects.filter(client_id=db_client.id,pool=cl_name).first()
	else:
		save_db = models.pool_master()
	save_db.address   = address
	save_db.pool      = cl_name
	save_db.client_id = db_client.id
	save_db.vm_name   = vm_name		
	save_db.save()


	for f in table_details:		
		if f['vessel']!=None:
			print '-----------',f['vessel']
			ship_name  		= f['vessel']
			voy_no 	   		= f['voy_no']
			disch_port      = f['disch_port']	
			disch_date 	   	= f['disch_date']
			load_port  		= f['load_port']
			load_date    	= f['load_date']
			qty    			= f['qty']
			rate    		= f['rate']
			amount  		= f['amount']
			CHK  			= models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,ship_name=ship_name,voyage_no=voy_no).count()
			if CHK>1:
				submit_invoice = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,ship_name=ship_name,voyage_no=voy_no).first()
			else:
				submit_invoice = models.invoice()

			# if len(table_details)>1:
			# 	invoice_noz 			  = invoice_nox
			# 	submit_invoice.invoice_no = invoice_noz
			# else:
			# 	submit_invoice.invoice_no = invoice_noz

			disch_dates = datetime.strptime(disch_date, "%d/%b/%Y").strftime('%Y-%m-%d')	
			load_dates = datetime.strptime(load_date, "%d/%m/%Y").strftime('%Y-%m-%d')	

			submit_invoice.invoice_no 	  = invoice_nox
			submit_invoice.ship_name      = ship_name
			submit_invoice.proj_name      = proj_name				
			submit_invoice.cancel_invoice = 0
			submit_invoice.vm_name 		  = vm_name
			submit_invoice.invoice_date   = invoice_date
			if curr_type=='USD':
				submit_invoice.usd  	  = 'USD'
				curr_type = 'USD'
				submit_invoice.inr  	  = None
			if curr_type=='INR':
				submit_invoice.usd  	  = None
				submit_invoice.inr  	  = 'INR'
				curr_type = 'INR'
			submit_invoice.month 	 	  = invoice_date
			submit_invoice.counter 		  = 0
			submit_invoice.client_id      = db_client.id
			submit_invoice.price 		  = rate
			submit_invoice.qty 			  = qty
			submit_invoice.rate  		  = rate
			submit_invoice.amount 		  = subtotal
			submit_invoice.load_port 	  = load_port
			submit_invoice.load_date  	  = load_dates
			submit_invoice.disch_date  	  = disch_dates
			submit_invoice.disch_port 	  = disch_port
			submit_invoice.voyage_no 	  = voy_no
			
			
			submit_invoice.vessel_type 	  = db_client.client_name
			try:
				submit_invoice.account_type = pool_name
			except:
				submit_invoice.account_type = None

			submit_invoice.client_address   = address
			submit_invoice.month 			= period_date
			submit_invoice.save()

			
	return HttpResponse(json.dumps('done'))



