from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################


def export_invoice_template(request):
	invoice_no 		= json.loads(request.GET['invoice_no'])[0]
	export_details 	= models.invoice.objects.filter(invoice_no=invoice_no)
	select_dir    	= os.path.dirname(__file__)
	srcfile       	= select_dir+'/static/invoice_summary_report.xlsx'
	dstroot       	= select_dir+'/invoice_summary_report.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      	= os.path.dirname(__file__)		
	location      	= (my_dir+'/invoice_summary_report.xlsx')
	wb 	   	      	= load_workbook(my_dir+'/invoice_summary_report.xlsx')
	ws     	      	= wb.get_sheet_by_name("Sheet1")	
	thin_border   	= Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      	= Style(font=Font(name='Calibri', size=12, bold=False),border=thin_border)
	ii   		  	= 5
	jj 	 		  	= 1
	single 		  	= []
	tt     		  	= 0
	vt 			  	= 0	
	e 			  	= 0
	ws.cell('A3').value = 'Invoice No. '+str(export_details[0].invoice_no)
	ws.cell('A1').value = export_details[0].client.client_name	
	get_month_name  	= datetime.strptime(export_details[0].month,"%Y-%m-%d").strftime('%B')
	get_year_name  		= datetime.strptime(export_details[0].month,"%Y-%m-%d").strftime('%Y')	
	ws.cell('A2').value = "Summary of Billing for the month of "+str(get_month_name)+' '+str(get_year_name)
	get_splt 			= export_details[0].month
	split_month 		= get_splt.split('-')
	get_month 			= monthrange(int(split_month[0]),int(split_month[1]))[1]
	
	if export_details[0].client.id==83 or export_details[0].client.id==113:	
		ws.cell('B4').value = "Vessel"
		ws.cell('C4').value = "Passage"
		ws.cell('D4').value = "Report ID"
		ws.cell('D4').value = "Billing Amount"
		
	if export_details[0].client.id==112:		
		ws.cell('B4').value = "Vessel"
		ws.cell('C4').value = "Total Days"
		ws.cell('D4').value = "Billing Amount"
		ws.cell('E4').value = ""

		


	s=0
	for x in export_details:
		thin_border			= Border(left=Side(style='thin',color='808080'),right=Side(style='thin',color='808080'),top=Side(style='thin',color='808080'),bottom=Side(style='thin',color='808080'))
	 	my_style			= Style(font=Font(name='Arial', size=10, bold=False),border=thin_border)
	 	_row_style			= Style(font=Font(name='Arial', size=10, bold=False),border=thin_border)
	 	_styles2 			= Style(font=Font(name='Arial', size=10, italic=False,bold=False),alignment=Alignment(horizontal='center',vertical='center'),border=thin_border)
		
		ws.cell('A'+str(ii)).value 	= e+1
		ws.cell('B'+str(ii)).value 	= x.ship_name
		if export_details[0].client.id==112:
			ws.cell('C'+str(ii)).value 	= x.qty
			calcs  = float(x.qty*x.price)/int(get_month) 
			if calcs!=x.price:
				try:
					ws.cell('D'+str(ii)).value 	= round(float(x.qty*x.price)/int(get_month),1)
				except:
					ws.cell('D'+str(ii)).value 	= ''
			else:
				ws.cell('D'+str(ii)).value 	= round(calcs,1)
			

		if export_details[0].client.id==83 or export_details[0].client.id==113:
			ws.cell('C'+str(ii)).value 	= x.voyage_no
			ws.cell('E'+str(ii)).value 	= x.price		
			ws.cell('D'+str(ii)).value 	= x.deadwt
		

		ws.cell('A'+str(ii)).style 	= _styles2
		ws.cell('B'+str(ii)).style 	= _row_style
		ws.cell('C'+str(ii)).style 	= _styles2
		ws.cell('D'+str(ii)).style 	= _row_style
		ws.cell('E'+str(ii)).style 	= _row_style
		if export_details[0].client.id==112:
			s+=calcs
		elif export_details[0].client.id==83 or export_details[0].client.id==113:			
			s+=x.price
		else:
			s+=x.price


		ii+=1
		jj+=1
		tt+=1
		e+=1
		vt+=1

	_border = Border(top=Side(style='thin'))
	style_border  = Style(font=Font(name='Arial', size=10, italic=False,bold=True,color=colors.BLACK),alignment=Alignment(horizontal='right'))			
	_styles 	  = Style(font=Font(name='Calibri', size=11, bold=True),fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFFF00')),border=Border(top=Side(style='thin'),bottom=Side(style='thin')))
	_row_count 							 = ws.max_row
	
	if export_details[0].client.id==83 or export_details[0].client.id==113:
		ws.cell('E'+str(_row_count+1)).value = s
	elif export_details[0].client.id==112:
		ws.cell('D'+str(_row_count+1)).value = round(s,1)

	merges1 = ('C'+str(_row_count+2)+':D'+str(_row_count+2))
	ws.merge_cells(merges1)	
	if export_details[0].client.id==112:
		ws.cell('C'+str(_row_count+2)).value = ""
		total_amount 		= 0
		amount_with_words 	= convert_money_to_text(s)
		merges = ('B'+str(_row_count+4)+':D'+str(_row_count+4))
		ws.merge_cells(merges)
		ws.cell('B'+str(_row_count+4)).value = '(Rupees '+str(amount_with_words)+')'
	else:
		ws.cell('C'+str(_row_count+2)).value = 'Add: IGST @18%'		
		detect_amount 						 = ((s*18)/100)
		total_amount 						 = (s+detect_amount)
		ws.cell('E'+str(_row_count+2)).value = ((s*18)/100)	
		ws.cell('E'+str(_row_count+3)).value = total_amount
		amount_with_words 					 = convert_money_to_text(total_amount)

		merges = ('C'+str(_row_count+4)+':E'+str(_row_count+4))
		ws.merge_cells(merges)
		ws.cell('C'+str(_row_count+4)).value = '(Rupees '+str(amount_with_words)+')'

	if export_details[0].client.id==83 or export_details[0].client.id==113:
		ws.cell('E'+str(_row_count+1)).style  = style_border
		ws.cell('E'+str(_row_count+2)).style  = style_border
		ws.cell('E'+str(_row_count+3)).style  = _styles
		ws.cell('C'+str(_row_count+4)).style  = style_border
		ws.cell('C'+str(_row_count+2)).style  = style_border
	elif export_details[0].client.id==112:
		ws.cell('D'+str(_row_count+1)).style  = style_border
		ws.cell('D'+str(_row_count+2)).style  = style_border
		ws.cell('D'+str(_row_count+1)).style  = _styles
		ws.cell('C'+str(_row_count+4)).style  = style_border
		ws.cell('C'+str(_row_count+1)).style  = style_border
		ws.cell('B'+str(_row_count+1)).style  = style_border

	else:
		ws.cell('E'+str(_row_count+1)).style  = style_border
		ws.cell('E'+str(_row_count+2)).style  = style_border
		ws.cell('E'+str(_row_count+3)).style  = _styles
		ws.cell('C'+str(_row_count+4)).style  = style_border
		ws.cell('C'+str(_row_count+2)).style  = style_border

	

	wb.save(my_dir+'/invoice_summary_report.xlsx')

	return HttpResponse(json.dumps('done'))



def other_export_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/invoice_summary_report.xlsx')
	filename 	   					= my_dir+'/invoice_summary_report.xlsx'
	download_name  					= "Summary of Billing.xlsx"	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response


def convert_money_to_text(money):
    try:
        money_number = int(money)
    except ValueError:
        return "You must enter Number"
    if money_number == 0:
        return None
    positions 		= [None for i in range(4)]
    key_range 		= 0
    one_place 		= ["", "One ", "Two ", "Three ", "Four ","Five ", "Six ", "Seven ", "Eight ", "Nine "]
    one_ten_place   = ["Ten ", "Eleven ", "Twelve ", "Thirteen ", "Fourteen ","Fifteen ", "Sixteen ", "Seventeen ", "Eighteen ","Nineteen "]
    ten_place 		= ["Twenty ", "Thirty ", "Forty ", "Fifty ","Sixty ", "Seventy ", "Eighty ", "Ninety "]
    name_of_number  = ["Thousand ", "Lakh ", "Crore "]
    money_number_money_text = ''
    positions[0] = money_number % 1000  # units
    positions[1] = money_number / 1000
    positions[2] = money_number / 100000
    positions[1] = positions[1] - 100 * positions[2]  # thousands
    positions[3] = money_number / 10000000  # crores
    positions[2] = positions[2] - 100 * positions[3]  # lakhs
    for counter in range(3, 0, -1):
        if positions[counter] != 0:
            key_range = counter
            break
    for i in range(key_range, -1, -1):
        if positions[i] == 0:
            continue
        ones = positions[i] % 10  # ones
        tens = positions[i] / 10
        hundreds = positions[i] / 100  # hundreds
        tens = tens - 10 * hundreds  # tens
        if (hundreds > 0):
            money_number_money_text += one_place[hundreds] + "Hundred "
        if (ones > 0 or tens > 0):
            if (hundreds > 0 and i == 0):
                money_number_money_text += "and "
            if (tens == 0):
                money_number_money_text += one_place[ones]
            elif (tens == 1):
                money_number_money_text += one_ten_place[ones]
            else:
                money_number_money_text += ten_place[tens - 2] + \
                    one_place[ones]
        if (i != 0):
            money_number_money_text += name_of_number[i - 1]
    return money_number_money_text
