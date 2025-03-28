from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection
##############################################
##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt

from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

#####################################################
def other_invoice(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'gmadam':
		
		return render_to_response("invoice_display/other_invoice.html")
	else:
		return HttpResponseRedirect('/it/user_login')


def get_other_invoice(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'gmadam':
		typex     = request.GET['type']
		proj_name = request.GET['proj_name']
		
		if typex=='USD':
			invoice_details = models.invoice.objects.filter(proj_name=proj_name,usd='USD').order_by('-invoice_no').first()
		elif typex=='INR':
			invoice_details = models.invoice.objects.filter(proj_name=proj_name,inr='INR').order_by('-invoice_no').first()

		get_invoice = []
		slp         = invoice_details.invoice_no.split('/')
		first_no    = int(slp[0])+1
		second_no   = str(first_no)+'/'+str(slp[1])		

		get_invoice.append({
			'invoice_no' : second_no,
		})

		context={
			'get_invoice' : get_invoice
		}		
		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def submit_other_invoice(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'gmadam':
		table_details = json.loads(request.POST['table_details'])
		invoice_nox	  = request.POST['invoice_no']
		proj_name	  = request.POST['proj_name']
		curr_type 	  = request.POST['curr_type']
		vm_name       = request.POST['vm_name']
		in_date 	  = request.POST['in_date']
		cl_name       = request.POST['cl_name']
		gtin_no       = request.POST['gtin_no']
		address 	  = request.POST['address']

		invoice_date  = datetime.strptime(in_date, "%m/%d/%Y").strftime('%Y-%m-%d')		
		s     		  = 0

		check_client = models.Client.objects.filter(client_name=cl_name,currency_type=curr_type,proj_name=proj_name).count()		
		if check_client!=0:
			db_client =  models.Client.objects.filter(client_name=cl_name,currency_type=curr_type,proj_name=proj_name).first()
		else:
			db_client =  models.Client()

		db_client.client_name   = cl_name
		db_client.proj_name     = proj_name
		db_client.currency_type = curr_type
		db_client.rate 			= 1
		db_client.price_type    = None
		db_client.duration_type = 'Voyagewise'
		db_client.vm_name       = cl_name
		db_client.tin_number    = gtin_no
		db_client.status 		= 1
		db_client.save()

		check_dbx = models.pool_master.objects.filter(client_id=db_client.id,pool=cl_name).count()
		if check_dbx!=0:
			save_db = models.pool_master.objects.filter(client_id=db_client.id,pool=cl_name).first()
		else:
			save_db = models.pool_master()
		save_db.address   = address
		save_db.pool      = cl_name
		save_db.client_id = db_client.id
		save_db.vm_name   = vm_name
		save_db.save()


		for f in table_details:
			ship_name  = f['Description']
			unit 	   = f['unit_cost']
			qty        = f['qty']	
			rate 	   = f['total_cost']			
			s+=float(unit)*float(qty)
			check_db = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name).count()
			if check_db>1:
				submit_invoice = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name).first()
			else:
				submit_invoice = models.invoice()

			if len(table_details)>1:
				invoice_noz 			      = invoice_nox			
				submit_invoice.invoice_no     = invoice_noz
			else:
				submit_invoice.invoice_no = invoice_noz

			submit_invoice.ship_name      = ship_name
			submit_invoice.proj_name      = proj_name
			submit_invoice.cancel_invoice = 0
			submit_invoice.vm_name 		  = vm_name
			submit_invoice.invoice_date   = invoice_date
			if curr_type=='USD':
				submit_invoice.usd  	  = 'USD'
				curr_type = 'USD'
				submit_invoice.inr  	  =  None
			if curr_type=='INR':
				submit_invoice.usd  	  = None
				submit_invoice.inr  	  = 'INR'
				curr_type = 'INR'
			submit_invoice.month 	 	  = invoice_date
			submit_invoice.counter 		  = 0
			submit_invoice.client_id      = db_client.id
			submit_invoice.price 		  = unit
			submit_invoice.qty 			  = qty
			submit_invoice.rate  		  = rate
			submit_invoice.vessel_type    = db_client.client_name
			submit_invoice.client_address = address
			submit_invoice.save()

		update_invoice_details = models.invoice.objects.filter(invoice_no=invoice_nox,proj_name=proj_name,client__id=db_client.id)
		for x in update_invoice_details:			
			x.total_amount = s
			x.save()

		context={
		'proj_name'    : proj_name,
		'curr_type'	   : curr_type,
		'invoice_date' : str(datetime.date().date()),
		'msg' 		   : 'done'
		}

		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')