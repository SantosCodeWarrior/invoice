from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
#####################################################
import calendar


def invoice_raise_summary(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		login_user = request.user	
		client_list  = models.Client.objects.filter(proj_name='CHM',currency_type='USD',status=0).order_by('client_name').exclude(client_name="")
		cl_array     = []
		single_array = []
		for cl in client_list:
			if cl.client_name not in single_array:
				single_array.append(cl.client_name)

		sums 		= 0
		fin_date    = '2020'
		fin_start   = '2020-04-01'
		fin_end     = '2021-03-31'
		curr_date   = datetime.now().date()
		
		month_array = list(calendar.month_name[1:])
		i    	     = 0
		j            = 1
		#print '----',(fin_start)		
		for v in single_array:
			#if j<=12:
			cl_name    = v.encode('utf-8').strip()		
			client_na  = models.Client.objects.filter(client_name=cl_name,currency_type='USD').order_by('client_name').first()		
			qry        = 'SELECT (inv.invoice_amount) FROM it_invoice inv,it_client cl where (inv.client_id=cl.id) and cl.id="1" and inv.usd="USD" and (inv.invoice_date)>="'+str(fin_start)+'" and inv.invoice_date<="'+str(fin_end)+'" and MONTH(inv.invoice_date)="1" and  cancel_invoice=0 group by MONTH(inv.invoice_date) order by MONTH(inv.invoice_date)'
			cursor     = connection.cursor()
			cursor.execute(qry)
			curr_deta  = cursor.fetchall()
			
			try:
				invoice_raised = curr_deta[i][0]				
			except:
				invoice_raised = 0

			print '----------',	qry
			
			try:
				dt_date = str(curr_date)
			except:
				dt_date = ''		

			cl_array.append({
				'client_name' : v,
				'id' 		  : client_na.id,
				'type' 		  : client_na.currency_type,
				'service'     : client_na.proj_name,
				'inv_date' 	  : dt_date,
				'invoice_amt' : invoice_raised,
				})				
			i+=1
			j+=1

		context={
		'client_array' : cl_array,
		'login_user'   : login_user,		
		}
		return render_to_response('invoice_display/summary_report.html',context)
	else:
		return HttpResponseRedirect('/it/user_login')