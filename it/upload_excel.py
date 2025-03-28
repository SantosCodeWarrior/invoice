from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
#####################################################
from it.forms import DocumentForm
import xlrd

from openpyxl import load_workbook

@csrf_exempt
def upload_excel(request):	
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		print ''
	else:
		if request.user.is_authenticated():
			return HttpResponseRedirect('/it/user_login')	
	return render_to_response("list.html")
	#else:
	#	return HttpResponseRedirect('/it/user_login')

def list(request):

	return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def submit_upload_invoice(request):
	models.uploader.objects.all().delete()
	if request.method == 'POST':
		i=24
		j=1
		for f in request.FILES:
			form 		= DocumentForm(request.POST, request.FILES) 
			upload_file = request.FILES[f]					
			newdoc 		 		  = models.uploader()				
			newdoc.upload_date    = datetime.now().date()  
			newdoc.invoice_excel  = upload_file
			newdoc.file_name 	  = upload_file
			newdoc.save()
			excel_files_details = models.uploader.objects.filter(id=str(newdoc))
			for e in excel_files_details:
				get_the_file_name = '/var/www/html/invoice/it/static/excel/'+str(e.file_name)				
				wb   = load_workbook(get_the_file_name, data_only=True)
				sh   = wb["1458-1"]
				x    = 0
				sums = 0
				for c in range(26,230):
					
					if sh.cell('C'+str(c)).value != 'Service Details' and sh.cell('C'+str(c)).value != None and sh.cell('C'+str(c)).value != 0:
						ship_name 		= sh.cell('C'+str(c)).value
						rate 			= sh.cell('I'+str(c)).value
						qty 			= sh.cell('H'+str(c)).value
						invoice_no 	 	= sh.cell('H5').value
						fleet_type  	= 'Petredec' #sh.cell('H12').value#####
						vm_name     	= sh.cell('D12').value
						invoice_period 	= sh.cell('H13').value
						invoice_date 	= sh.cell('H6').value						
						sums+=qty*rate
						#print '--------',invoice_period
						 
						client_s 	 	= sh.cell('D17').value #sh.cell('B6').value
						split_client 	= client_s.split('/')
						client_name  	= split_client[0]						
						client_details 	= models.Client.objects.filter(client_name=client_name,proj_name='BOSS',currency_type='USD').first()
						pool_details    = models.pool_master.objects.filter(client_id=client_details.id).first()
						#print '-----------',client_name

						#check_invoice   = models.invoice.objects.filter(invoice_no=invoice_no,proj_name='BOSS',usd='USD',client_id=client_details.id).count()						
						#if check_invoice>0:
						#	db_invoice   = models.invoice.objects.filter(invoice_no=invoice_no,proj_name='BOSS',usd='USD',client_id=client_details.id).first()
						#else:
						db_invoice   = models.invoice()					

						db_invoice.invoice_no 	  = invoice_no
						db_invoice.vm_name    	  = vm_name
						db_invoice.invoice_date   = invoice_date			
						db_invoice.vessel_type 	  = fleet_type
						db_invoice.qty 			  = qty		
						db_invoice.rate 		  = 1
						db_invoice.ship_name      = ship_name
						db_invoice.client_id      = client_details.id
						db_invoice.counter 		  = 0
						db_invoice.deadwt		  = 0						
						db_invoice.price 		  = rate
						split_date 				  = invoice_period.split(' ')
					
						# if client_name=='Clearlake' or client_name=='CLEARLAKE': 
						# 	get_month 			  = split_date[4] # Clearlake = 4
						# 	get_year 			  = split_date[6] # Clearlake = 6
						#else:
						get_month 			  = split_date[1] # shell 1
						get_year 			  = split_date[6] # shell 5

						format_period 			  = '01-'+str(get_month)+'-'+str(get_year)
						#print '--------',get_month,'-----',get_year,'-----',split_date
						format_end_date			  = datetime.strptime(str(format_period), "%d-%b-%Y").strftime('%Y-%m-%d')						
						db_invoice.month 		  = format_end_date
						db_invoice.client_address = pool_details.address
						db_invoice.url            = '/static/pdf/Tristar/Online/1340_2021_Tristar.pdf'
						db_invoice.proj_name 	  = 'BOSS'
						db_invoice.usd 			  = 'USD'	
						db_invoice.total_amount   = 0
						db_invoice.voyage_no 	  = ""
						db_invoice.remark 		  = ""
						#print '---------------',format_end_date
						
						#if client_name=='Clearlake':
						#	db_invoice.account_type  = 0
						#else:
						db_invoice.account_type  = fleet_type						
						db_invoice.save()				
						x+=1

		return render_to_response('list.html',{'form': form},context_instance=RequestContext(request))

	return HttpResponse(json.dumps('Done'))


##################################
# Note: Date format -> %d-%B-%Y
#					-> %d-%b-%Y
# For Clearlake's excel