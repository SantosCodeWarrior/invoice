from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection
##############################################
##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt

from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

#####################################################
def get_client_api(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		get_client_details 	= models.Client.objects.filter(status=1,proj_name='BOSS',currency_type='USD').exclude(id=15581)
		cl_array 			= []
		for x in get_client_details:
			cl_array.append({
				'id' 			: x.id,
				'client_name' 	: x.client_name,
				})

		context={
			'cl_array' : cl_array,
		}
		
		return render_to_response("invoice_display/get_client_api.html",context)
	else:
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
		return HttpResponseRedirect('/it/user_login')


# @csrf_exempt
# def client_fetch_data(request):
# 	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
# 		fetch_type 		   = request.POST['ft_data']
# 		client_id 		   = request.POST['client_id']
# 		get_ft_date 	   = request.POST['get_ft_date']
# 		#print '------',get_ft_date
# 		get_date 		   = get_ft_date #datetime.strptime(start_date,"%m/%d/%Y").strftime('%Y-%m-%d')	
# 		get_client_id 	   = client_id
# 		bill_type 		   = fetch_type
# 		api_url 		   = "https://aboss.bwesglobal.com/api/get_ship_to_be_billed/"		
# 		api_method         = "POST"			
# 		parameters         = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'client_id': get_client_id,'billing_type':bill_type,'billing_date':(get_date)}
# 		response           = requests.get(api_url, params=parameters,verify=False)		
# 		boss_array         = json.loads(response.content)
# 		voyage_array	   = []

# 		for c in boss_array:
# 			get_account = models.api_client_data.objects.filter(client_id=get_client_id,ship_name=c['Ship Name']).first()
# 			try:
# 				get_account = get_account.account_tab
# 			except:
# 				get_account = ""
				
# 			if fetch_type=='monthly_prorated':
# 				voyage_array.append({
# 					'id'			: c['Client ID'],
# 					'ship_name' 	: c['Ship Name'],
# 					'no_of_day'		: c['No. of Days'],
# 					'start_date'	: c['Start Date'],
# 					'end_date'  	: c['End Date'],
# 					'first_port'	: c['First Port'],
# 					'last_port'		: c['Last Port'],
# 					'account'		: get_account
# 					})
# 			if fetch_type=='monthly_flat':
# 				voyage_array.append({
# 					'id'			: c['Client ID'],
# 					'ship_name' 	: c['Ship Name'],
# 					'no_of_day'		: 0,
# 					'start_date'	: 0,
# 					'end_date'  	: 0,
# 					'first_port'	: 0,
# 					'last_port'		: 0,
# 					'account'		: get_account
# 					})

# 		context={
# 			'voyage_array' : voyage_array,
# 		}
	
# 		return HttpResponse(json.dumps(context))
# 	else:
# 		if user.is_anonymous():
# 			return HttpResponseRedirect('/it/user_login')
# 		return HttpResponseRedirect('/it/user_login')