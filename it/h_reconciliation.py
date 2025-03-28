from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################

@csrf_exempt
def h_reconciliation(request):
	if request.user.is_authenticated():
		login_user 	= request.user

		context={
			'login_user' : login_user,	
		}		
		
		return render_to_response("invoice_display/h_reconciliation.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def inserting_remittance_data(request):
	datas = json.loads(request.POST['datas'])
	for c in datas:
		if c['client_name']!=None:
			print '---------',c['net_inward_amt_ref']

	return HttpResponse(json.dumps('done'))



def calc_remittance_datas(request):	
	
	try:	
		e_n 		= json.loads(request.GET['n'])
		e_service 	= json.loads(request.GET['service'])
		e_invoice 	= json.loads(request.GET['invoice_no'])	
		amountss    = 0
		
		get_calc 	= models.invoice.objects.filter(proj_name=e_service,invoice_no=e_invoice).first()
		get_amount 	= models.invoice.objects.filter(proj_name=e_service,invoice_no=e_invoice).aggregate(Sum('price'))
		#curr_type   = get_calc.client.currency_type

		try:
			month_select = datetime.strptime(str(get_calc.month),"%Y-%m-%d").strftime('%m')
			year_select  = datetime.strptime(str(get_calc.month),"%Y-%m-%d").strftime('%Y')
			get_month    = monthrange(int(year_select),int(month_select))[1]
		except:
			get_month    = 0

		if get_calc.client_id!=20001:
			if e_service=='CHM':			
				query1 		= "SELECT sum(price*qty) FROM invoice.it_invoice where invoice_no='"+str(e_invoice)+"';"	
				cursor 		= connection.cursor()
				cursor.execute(query1)
				totamts 	= cursor.fetchall()
				amountss 	= round(totamts[0][0],0)

			elif e_service=='BOSS':			
				query1 		= "SELECT sum(price*qty) FROM invoice.it_invoice where invoice_no='"+str(e_invoice)+"';"	
				cursor 		= connection.cursor()
				cursor.execute(query1)
				totamts 	= cursor.fetchall()
				amountss 	= round(totamts[0][0],0)

			else:
				query1 		= "SELECT sum(price*qty)/"+str(get_month)+" FROM invoice.it_invoice where invoice_no='"+str(e_invoice)+"';"	
				cursor 		= connection.cursor()
				cursor.execute(query1)
				totamts 	= cursor.fetchall()
				amountss 	= round(totamts[0][0],0)			
		else:
			query1 		= "SELECT sum(price*qty*rate) FROM invoice.it_invoice where invoice_no='"+str(e_invoice)+"';"	
			cursor 		= connection.cursor()
			cursor.execute(query1)
			totamts 	= cursor.fetchall()
			amountss 	= round(totamts[0][0],0)
			#print '-------',query1
		
		if get_calc.client_id==112:
			amountss = round(get_calc.usd_amount,0)
			print '------wala hoga'
		elif get_calc.client_id==20017:
			amountss = get_calc.usd_amount #get_amount['price__sum']
			print '--------22 hoga'
		elif get_calc.client_id==115:
			amountss = get_amount['price__sum']
		elif get_calc.client_id==20035:
			amountss = get_calc.usd_amount
		elif get_calc.client_id==15588:
			amountss = get_calc.usd_amount
		elif get_calc.client_id==20034:
			amountss = get_calc.usd_amount
		elif get_calc.client_id==20044:
			amountss = get_calc.usd_amount
		elif get_calc.client_id==20047:
			amountss = get_calc.usd_amount
		elif get_calc.client_id==20046:
			amountss = get_calc.usd_amount
		# elif get_calc.client_id==20040:
		# 	amountss = get_calc.usd_amount

		elif get_calc.client_id==64 or get_calc.client_id==20040:
			query1 		= "SELECT sum(price*qty*rate) FROM invoice.it_invoice where invoice_no='"+str(e_invoice)+"';"	
			cursor 		= connection.cursor()
			cursor.execute(query1)
			totamts 	= cursor.fetchall()
			amountss 	= round(totamts[0][0],0)


		
		print '--------',amountss
		# else:
		# 	if get_calc.qty>1:
		# 		amt_qty = get_amount['price__sum']*float(get_calc.qty)
		# 		amountss  = amt_qty
		# 		print '--------usd wala hoga'
		# 	else:
		# 		if e_currency=='INR':
		# 			amountss		= get_amount['price__sum']*float(get_calc.rate)
		# 			print '------inr wala hoga'
		# 		else:
		# 			in_amt = get_amount['price__sum']
		# 			amountss		= (get_amount['price__sum']*float(get_calc.rate))
		# 		print '-------yahi tha kya'
		context={
			'sno' 		: e_n,
			'amount' 	: amountss,
		}



	except:
		context={
			'sno' 		: 0,
			'amount' 	: 0,
		}

	return HttpResponse(json.dumps(context))


def submitting_remittance_data(request):
	value_date 			= json.loads(request.GET['value_date'])
	client_name 		= json.loads(request.GET['client_name'])
	remitttance_date 	= json.loads(request.GET['remitttance_date'])
	inward_number 		= json.loads(request.GET['inward_number'])
	inward_amount 		= json.loads(request.GET['inward_amount'])
	currency_type 		= json.loads(request.GET['currency_type'])
	#market_rate 		= json.loads(request.GET['market_rate'])
	rate 				= json.loads(request.GET['rate'])
	amount_inr 			= json.loads(request.GET['amount_inr'])
	net_amt_ref 		= json.loads(request.GET['net_amt_ref'])
	bank_deduct 		= json.loads(request.GET['bank_deduct'])
	tds 				= json.loads(request.GET['tds'])
	service1 			= json.loads(request.GET['service1'])
	invoice_n1 			= json.loads(request.GET['invoice_n1'])
	amount1 			= json.loads(request.GET['amount1'])
	service2 			= json.loads(request.GET['service2'])
	invoice_n2 			= json.loads(request.GET['invoice_n2'])
	amount2 			= json.loads(request.GET['amount2'])
	service3 			= json.loads(request.GET['service3'])
	invoice_n3 			= json.loads(request.GET['invoice_n3'])
	amount3 			= json.loads(request.GET['amount3'])
	service4 			= json.loads(request.GET['service4'])
	invoice_n4 			= json.loads(request.GET['invoice_n4'])
	amount4 			= json.loads(request.GET['amount4'])
	service5 			= json.loads(request.GET['service5'])
	invoice_n5 			= json.loads(request.GET['invoice_n5'])
	amount5 			= json.loads(request.GET['amount5'])
	service6 			= json.loads(request.GET['service6'])
	invoice_n6 			= json.loads(request.GET['invoice_n6'])
	amount6 			= json.loads(request.GET['amount6'])
	service7 			= json.loads(request.GET['service7'])
	invoice_n7 			= json.loads(request.GET['invoice_n7'])
	amount7 			= json.loads(request.GET['amount7'])
	service8 			= json.loads(request.GET['service8'])
	invoice_n8 			= json.loads(request.GET['invoice_n8'])
	amount8 			= json.loads(request.GET['amount8'])
	service9 			= json.loads(request.GET['service9'])
	invoice_n9 			= json.loads(request.GET['invoice_n9'])
	amount9 			= json.loads(request.GET['amount9'])
	service10 			= json.loads(request.GET['service10'])
	invoice_n10 		= json.loads(request.GET['invoice_n10'])
	amount10 			= json.loads(request.GET['amount10'])
	service11 			= json.loads(request.GET['service11'])
	invoice_n11 		= json.loads(request.GET['invoice_n11'])
	amount11 			= json.loads(request.GET['amount11'])
	service12 			= json.loads(request.GET['service12'])
	invoice_n12 		= json.loads(request.GET['invoice_n12'])
	amount12 			= json.loads(request.GET['amount12'])
	service13 			= json.loads(request.GET['service13'])
	invoice_n13 		= json.loads(request.GET['invoice_n13'])
	amount13 			= json.loads(request.GET['amount13'])
	service14 			= json.loads(request.GET['service14'])
	invoice_n14 		= json.loads(request.GET['invoice_n14'])
	amount14 			= json.loads(request.GET['amount14'])
	service15 			= json.loads(request.GET['service15'])
	invoice_n15 		= json.loads(request.GET['invoice_n15'])
	amount15 			= json.loads(request.GET['amount15'])

	service16 			= json.loads(request.GET['service16'])
	invoice_n16 		= json.loads(request.GET['invoice_n16'])
	amount16 			= json.loads(request.GET['amount16'])

	service17 			= json.loads(request.GET['service17'])
	invoice_n17 		= json.loads(request.GET['invoice_n17'])
	amount17 			= json.loads(request.GET['amount17'])

	service18 			= json.loads(request.GET['service18'])
	invoice_n18 		= json.loads(request.GET['invoice_n18'])
	amount18 			= json.loads(request.GET['amount18'])

	service19 			= json.loads(request.GET['service19'])
	invoice_n19 		= json.loads(request.GET['invoice_n19'])
	amount19 			= json.loads(request.GET['amount19'])

	service20 			= json.loads(request.GET['service20'])
	invoice_n20 		= json.loads(request.GET['invoice_n20'])
	amount20 			= json.loads(request.GET['amount20'])

	service21 			= json.loads(request.GET['service21'])
	invoice_n21 		= json.loads(request.GET['invoice_n21'])
	amount21 			= json.loads(request.GET['amount21'])
	
	service22 			= json.loads(request.GET['service22'])
	invoice_n22 		= json.loads(request.GET['invoice_n22'])
	amount22 			= json.loads(request.GET['amount22'])
	
	service23 			= json.loads(request.GET['service23'])
	invoice_n23 		= json.loads(request.GET['invoice_n23'])
	amount23 			= json.loads(request.GET['amount23'])
	
	service24 			= json.loads(request.GET['service24'])
	invoice_n24 		= json.loads(request.GET['invoice_n24'])
	amount24 			= json.loads(request.GET['amount24'])
	
	service25 			= json.loads(request.GET['service25'])
	invoice_n25 		= json.loads(request.GET['invoice_n25'])
	amount25 			= json.loads(request.GET['amount25'])

	service26 			= json.loads(request.GET['service26'])
	invoice_n26 		= json.loads(request.GET['invoice_n26'])
	amount26 			= json.loads(request.GET['amount26'])

	service27 			= json.loads(request.GET['service27'])	
	invoice_n27 		= json.loads(request.GET['invoice_n27'])
	amount27 			= json.loads(request.GET['amount27'])

	service28 			= json.loads(request.GET['service28'])
	invoice_n28 		= json.loads(request.GET['invoice_n28'])
	amount28 			= json.loads(request.GET['amount28'])

	service29 			= json.loads(request.GET['service29'])
	invoice_n29 		= json.loads(request.GET['invoice_n29'])
	amount29 			= json.loads(request.GET['amount29'])

	service30 			= json.loads(request.GET['service30'])
	invoice_n30 		= json.loads(request.GET['invoice_n30'])
	amount30 			= json.loads(request.GET['amount30'])
	

	remarks 			= json.loads(request.GET['remarks'])

	CHK 				= models.remittance_data.objects.filter(certificate_no=inward_number.strip()).count()
	if CHK!=0:
		db = models.remittance_data.objects.filter(certificate_no=inward_number.strip()).first()
		msg = 'Already Submitted'
	else:
		db = models.remittance_data()
		msg = 'Done'

	split_client 		= client_name.split('-')
	cl_name 			= split_client[0]
	proj_name 			= split_client[1]
	curr_types      	= split_client[2]
	
	f_client_name   	= cl_name.encode('ascii',errors='ignore')
	f_proj_name 		= proj_name.encode('ascii',errors='ignore')
	f_curr_types 		= curr_types.encode('ascii',errors='ignore')


	db.certificate_no   = inward_number.strip()
	db.remittance_date  = (remitttance_date).encode('ascii',errors='ignore')
	db.inward_amount    = inward_amount
	db.client_name      = f_client_name.strip()
	db.currency         = (f_curr_types).strip()
	db.rate             = rate
	db.bank_charges     = bank_deduct
	db.invoice_amount   = inward_amount
	db.inr_amount       = amount_inr
	db.net_amount_ref   = net_amt_ref
	db.tds_deducted     = tds
	db.entry_date       = datetime.now().date()
	db.remarks          = remarks
	db.value_date 	    = value_date.encode('ascii',errors='ignore')
	

	try:
		db.services1   	= service1.encode('ascii',errors='ignore').strip()
		db.invoice_no1 	= invoice_n1
		db.amount1  	= amount1
	except:
		db.amount1  	= None

	try:
		db.services2  	= service2.encode('ascii',errors='ignore').strip()
		db.invoice_no2 	= invoice_n2
		db.amount2  	= amount2
	except:
		db.amount2  	= None

	try:
		db.services3 	= service3.encode('ascii',errors='ignore').strip()
		db.invoice_no3 	= invoice_n3
		db.amount3 		= amount3
	except:
		db.amount3 		= None

	try:
		db.services4 	= service4.encode('ascii',errors='ignore').strip()
		db.invoice_no4 	= invoice_n4
		db.amount4 		= amount4
	except:
		db.amount4 		= None

	try:
		db.services5 	= service5.encode('ascii',errors='ignore').strip()
		db.invoice_no5 	= invoice_n5
		db.amount5 		= amount5
	except:
		db.amount5 		= None

	try:
		db.services6 	= service6.encode('ascii',errors='ignore').strip()
		db.invoice_no6 	= invoice_n6
		db.amount6 		= amount6
	except:
		db.amount6 		= None

	try:
		db.services7 	= service7.encode('ascii',errors='ignore').strip()
		db.invoice_no7 	= invoice_n7
		db.amount7 		= amount7
	except:
		db.amount7 		= None

	try:
		db.services8 	= service8.encode('ascii',errors='ignore').strip()
		db.invoice_no8 	= invoice_n8
		db.amount8 		= amount8
	except:
		db.amount8 		= None

	try:
		db.services9 	= service9.encode('ascii',errors='ignore').strip()
		db.invoice_no9 	= invoice_n9
		db.amount9 		= amount9
	except:
		db.amount9 		= None

	try:
		db.services10 	= service10.encode('ascii',errors='ignore').strip()
		db.invoice_no10 = invoice_n10
		db.amount10 	=  amount10
	except:
		db.amount10 	=  None

	try:
		db.services11 	= service11.encode('ascii',errors='ignore').strip()
		db.invoice_no11 = invoice_n11
		db.amount11 	= amount11
	except:
		db.amount11 	= None

	try:
		db.services12 	= service12.encode('ascii',errors='ignore').strip()
		db.invoice_no12 = invoice_n12
		db.amount12 	= amount12
	except:
		db.amount12 	= None

	try:
		db.services13 	= service13.encode('ascii',errors='ignore').strip()
		db.invoice_no13 = invoice_n13
		db.amount13 	= amount13
	except:
		db.amount13 	= None

	try:
		db.services14 	= service14.encode('ascii',errors='ignore').strip()
		db.invoice_no14 = invoice_n14
		db.amount14 	= amount14
	except:
		db.amount14 	= None

	try:
		db.services15 	= service15.encode('ascii',errors='ignore').strip()
		db.invoice_no15 = invoice_n15
		db.amount15 	= amount15
	except:
		db.amount15 	= None

	try:
		db.services16 	= service16.encode('ascii',errors='ignore').strip()
		db.invoice_no16 = invoice_n16
		db.amount16 	= amount16
	except:
		db.amount16 	= None


	try:
		db.services17 	= service17.encode('ascii',errors='ignore').strip()
		db.invoice_no17 = invoice_n17
		db.amount17 	= amount17
	except:
		db.amount17 	= None


	try:
		db.services18 	= service18.encode('ascii',errors='ignore').strip()
		db.invoice_no18 = invoice_n18
		db.amount18 	= amount18
	except:
		db.amount18 	= None


	try:
		db.services19 	= service19.encode('ascii',errors='ignore').strip()
		db.invoice_no19 = invoice_n19
		db.amount19 	= amount19
	except:
		db.amount19 	= None


	try:
		db.services20 	= service20.encode('ascii',errors='ignore').strip()
		db.invoice_no20 = invoice_n20
		db.amount20 	= amount20
	except:
		db.amount20 	= None

	try:
		db.services21 	= service21.encode('ascii',errors='ignore').strip()
		db.invoice_no21 = invoice_n21
		db.amount21 	= amount21
	except:
		db.amount21 	= None

	try:
		db.services22 	= service22.encode('ascii',errors='ignore').strip()
		db.invoice_no22 = invoice_n22
		db.amount22 	= amount22
	except:
		db.amount22 	= None

	try:
		db.services23 	= service23.encode('ascii',errors='ignore').strip()
		db.invoice_no23 = invoice_n23
		db.amount23 	= amount23
	except:
		db.amount23 	= None

	try:
		db.services24 	= service24.encode('ascii',errors='ignore').strip()
		db.invoice_no24 = invoice_n24
		db.amount24 	= amount24
	except:
		db.amount24 	= None

	try:
		db.services25 	= service25.encode('ascii',errors='ignore').strip()
		db.invoice_no25 = invoice_n25
		db.amount25 	= amount25
	except:
		db.amount25 	= None

	try:
		db.services26 	= service26.encode('ascii',errors='ignore').strip()
		db.invoice_no26 = invoice_n26
		db.amount26 	= amount26
	except:
		db.amount26 	= None

	try:
		db.services27 	= service27.encode('ascii',errors='ignore').strip()
		db.invoice_no27 = invoice_n27
		db.amount27 	= amount27
	except:
		db.amount27 	= None

	try:
		db.services28 	= service28.encode('ascii',errors='ignore').strip()
		db.invoice_no28 = invoice_n28
		db.amount28 	= amount28
	except:
		db.amount28 	= None

	try:
		db.services29 	= service29.encode('ascii',errors='ignore').strip()
		db.invoice_no29 = invoice_n29
		db.amount29 	= amount29
	except:
		db.amount29 	= None

	try:
		db.services30 	= service30.encode('ascii',errors='ignore').strip()
		db.invoice_no30 = invoice_n30
		db.amount30 	= amount30
	except:
		db.amount30 	= None

	db.save()

	return HttpResponse(json.dumps(msg))