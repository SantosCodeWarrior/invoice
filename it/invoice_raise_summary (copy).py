from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import NamedStyle
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
#####################################################


@csrf_exempt
def invoice_raise_summary(request):
	#track_details = json.loads(request.POST['track_details'])
	#e_start_date  = request.POST['sdate']
	#e_end_date    = request.POST['edate']
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/Template-InvoiceRaised Summary(Running List).xlsx'
	dstroot       = select_dir+'/Tracker_Export/Template-InvoiceRaised Summary(Running List).xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/Tracker_Export/Template-InvoiceRaised Summary(Running List).xlsx')
	wb 	   	      = load_workbook(my_dir+'/Tracker_Export/Template-InvoiceRaised Summary(Running List).xlsx')	
	ws	      	  = wb.get_sheet_by_name("IRS")	
	#thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	#my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	i    		  = 3
	j 	  		  = 1
	#start_date    = datetime.strptime(str(e_start_date), "%m/%d/%Y").strftime('%Y-%m-%d')
	#end_date      = datetime.strptime(str(e_end_date), "%m/%d/%Y").strftime('%Y-%m-%d')

	QRY1 		  = "SELECT invoice_date FROM invoice.it_invoice where proj_name='CHM' and usd='USD' and cancel_invoice='0' and payment_status is null group by MONTH(invoice_date);"
	cursor  	  = connection.cursor()
	cursor.execute(QRY1)
	qry1_details  = cursor.fetchall()
	for c in qry1_details:		
		print '_____',c[0]
	
		

	query 		  = "SELECT * FROM invoice.it_invoice WHERE cancel_invoice='0' and invoice_date>='2019-05-01'  GROUP BY client_id;"
	cursor  	  = connection.cursor()
	cursor.execute(query)
	track_details = cursor.fetchall()	
	
	vv=0
	for t in track_details:			
		thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		#my_style    = Style(font=Font(name='Calibri', size=9, bold=False),border=thin_border)
		#_row_style  = Style(font=Font(name='Calibri', size=9, bold=False),border=thin_border,alignment=Alignment(horizontal='center'))	
		invoice_no 	= t[1]
		if t[18]=='USD' and t[9]=='CHM':
			invoice_no   				= invoice_no
			invoice_date 				= t[2]
			client_details 				= models.Client.objects.filter(id=t[24]).first()
			ws.cell('A'+str(i+1)).value = str(client_details.client_name).strip()
			sum_invoice_amount 			= models.invoice.objects.filter(proj_name='CHM',usd='USD',cancel_invoice=0,client_id=t[24],invoice_date=t[2]).aggregate(Sum('invoice_amount'))		
			#ws.cell(column=i+vv, row=i+1, value=client_details.client_name)
			ws.cell('B'+str(i+1)).value = 'USD'
			ws.cell('C'+str(i+1)).value = 'CHM'			
			ws.cell('E'+str(i+1)).value = sum_invoice_amount['invoice_amount__sum']
			# ws.cell('F'+str(i+1)).value = invoice_no
			# ws.cell('G'+str(i+1)).value = t[2]
			#ws.cell(column=i+1, row=vv+1, value=sum_invoice_amount['invoice_amount__sum'])
			#vv+=4
			
			i+=1			
		j+=1
		wb.save(my_dir+'/Tracker_Export/Template-InvoiceRaised Summary(Running List).xlsx')
	return HttpResponse(json.dumps('done'))









def invoice_raise_export(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)
	location 	   					= (my_dir+'/Tracker_Export/Template-InvoiceRaised Summary(Running List).xlsx')
	filename 	   					= my_dir+'/Tracker_Export/Template-InvoiceRaised Summary(Running List).xlsx' #Select your file here.
	download_name  					= "Template-InvoiceRaised Summary(Running List).xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response


def tracker_summary(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'gmadam':
		QRY1 		  = "SELECT invoice_date FROM invoice.it_invoice where proj_name='CHM' and usd='USD' and cancel_invoice='0' and received_date is null group by year(invoice_date),month(invoice_date);"
		cursor  	  = connection.cursor()
		cursor.execute(QRY1)
		qry1_details  = cursor.fetchall()
		inv_date 	  = []
		for c in qry1_details:		
			inv_date.append({
				'invoice_date' : (c[0]),
				})

		query 		  = "SELECT * FROM invoice.it_invoice WHERE cancel_invoice='0' and invoice_date>='2019-05-01'  GROUP BY client_id;"
		cursor  	  = connection.cursor()
		cursor.execute(query)
		track_details = cursor.fetchall()
		all_invoice   = []		
		for t in track_details:
			if t[18]=='USD' and t[9]=='CHM':	
				client_details 		= models.Client.objects.filter(id=t[24]).first()
				sum_invoice_amount 	= models.invoice.objects.filter(proj_name='CHM',usd='USD',cancel_invoice=0,client_id=t[24],invoice_date=t[2]).aggregate(Sum('invoice_amount'))		
				
				
				all_invoice.append({
					'invoice_date'  : t[2],
					'currency_type' : 'USD',
					'proj_name'     : 'CHM',
					'tot_amount'    : sum_invoice_amount['invoice_amount__sum'],
					'client_name'   : str(client_details.client_name).strip(),					
				})

		context={
		'inv_date'    : inv_date,
		'all_invoice' : all_invoice
		}
		
		return render_to_response("invoice_display/tracker_summary.html",context)
	else:
		return HttpResponseRedirect('/it/user_login')