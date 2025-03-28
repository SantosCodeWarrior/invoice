from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################
def reconciliation(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		login_user 			= request.user
		invoice_array 		= []
		remitt_details 		= models.remittance_data.objects.all().aggregate(Min('remittance_date'))
		remittance_details 	= models.remittance_data.objects.all().order_by('-id')
		min_date  			= remitt_details['remittance_date__min']
		#print '-------',str(min_date).split('-')
		currency_list    	= []

		cl_array 			= []
		query 				= "SELECT id,client_name,proj_name,status FROM invoice.it_client where status='1' and currency_type='USD' group by client_name,proj_name order by client_name,proj_name"
		cursor 				= connection.cursor()
		cursor.execute(query)
		gt_client_det   	= cursor.fetchall()
		for t in gt_client_det:			
			cl_array.append({
				'id' 		  : t[0],
				'client_name' : t[1],
				'client_list' : str(t[1])+' ('+str(t[2])+')'
			})


		currency_details = models.currency_data.objects.all().order_by('currency')
		for c in currency_details:
			currency_list.append({
				'id' 			: c.id,
				'currency_name' : c.currency
				})


		for x in remittance_details:
			qrty 				= "SELECT reference_no,amount_inr,referencez,txn_date FROM invoice.it_bank_statement where descs like '%"+str(x.certificate_no)+"%'"
			cursor 				= connection.cursor()
			cursor.execute(qrty)
			get_reference_no  	= cursor.fetchall()
			try:
				gt_reference_no = get_reference_no[0][0]
			except:
				gt_reference_no = 0

			if x.tds_deducted!=None:
				try:
					amt_deduct = ((x.amount1+x.amount2+x.amount3+x.amount4+x.amount5+x.amount6+x.amount7+x.amount8+x.amount9+x.amount10+x.amount11+x.amount12+x.amount13+x.amount14+x.amount15)-x.tds_deducted)
				except:
					amt_deduct = 0
			else:
				try:
					amt_deduct = ((x.amount1+x.amount2+x.amount3+x.amount4+x.amount5+x.amount6+x.amount7+x.amount8+x.amount9+x.amount10+x.amount11+x.amount12+x.amount13+x.amount14+x.amount15)-x.bank_charges)
				except:
					amt_deduct = 0

			#print '-----------',amt_deduct
			
			if x.net_amount_ref==amt_deduct:
				status 		= "Reconciled"
				bg_color 	= "white"
			else:
				status 		= "Not Reconciled"
				bg_color 	= "#f58787"


			if x.bank_charges!=0:
				chargs_color   	= '#f7aba6'
				chrgs 			= x.bank_charges
			else:
				chargs_color 	= ''
				chrgs 			= ''

			try:
				if x.approved=='yes':
					bg_collors = '#c0f7b7'
				else:
					bg_collors = ''
			except:
				pass

			min_date = str(min_date) #'2022-01-01'

			try:
				tdsc  = x.tds_deducted
			except:
				tdsc = 0



			

			invoice_array.append({				
				'get_id'			 : x.id,
				'remittance_date'	 : (x.remittance_date), 					
				'certificate_no' 	 : x.certificate_no,
				'inward_amount' 	 : x.inward_amount,
				'currency' 			 : x.currency,
				'rate' 				 : x.rate,
				'amount_inr' 		 : x.inr_amount,
				'status' 			 : status,
				'bank_deduction' 	 : chrgs, #x.bank_charges,			
				'net_inward_amt_ref' : x.net_amount_ref,
				'total_amount'		 : amt_deduct,
				'tds_deducted'		 : tdsc,

				'service1'			 : x.services1,
				'invoice_no1'		 : x.invoice_no1,
				'amount1'			 : x.amount1,
				'service2'			 : x.services2,
				'invoice_no2'		 : x.invoice_no2,
				'amount2'			 : x.amount2,
				'service3'			 : x.services3,
				'invoice_no3'		 : x.invoice_no3,
				'amount3'			 : x.amount3,
				'service4'			 : x.services4,
				'invoice_no4'		 : x.invoice_no4,
				'amount4'			 : x.amount4,
				'service5'			 : x.services5,
				'invoice_no5'		 : x.invoice_no5,
				'amount5'			 : x.amount5,
				'service6'			 : x.services6,
				'invoice_no6'		 : x.invoice_no6,
				'amount6'			 : x.amount6,
				'service7'			 : x.services7,
				'invoice_no7'		 : x.invoice_no7,
				'amount7'			 : x.amount7,
				'service8'			 : x.services8,
				'invoice_no8'		 : x.invoice_no8,
				'amount8'			 : x.amount8,
				'service9'			 : x.services9,
				'invoice_no9'		 : x.invoice_no9,
				'amount9'			 : x.amount9,				
				'service10'			 : x.services10,
				'invoice_no10'		 : x.invoice_no10,
				'amount10'			 : x.amount10,
				'service11'			 : x.services11,
				'invoice_no11'		 : x.invoice_no11,
				'amount11'			 : x.amount11,	
				'service12'			 : x.services12,
				'invoice_no12'		 : x.invoice_no12,
				'amount12'			 : x.amount12,	
				'service13'			 : x.services13,
				'invoice_no13'		 : x.invoice_no13,
				'amount13'			 : x.amount13,	
				'service14'			 : x.services14,
				'invoice_no14'		 : x.invoice_no14,
				'amount14'			 : x.amount14,	
				'service15'			 : x.services15,
				'invoice_no15'		 : x.invoice_no15,
				'amount15'			 : x.amount15,
				

				'client_name'   	 : x.client_name,	
				'bg_color' 			 : bg_color,
				'remit_date'		 : str(x.remittance_date), 
				'gt_reference_no'    : gt_reference_no,		
				'remarks'			 : x.remarks,
				'chargs_color'		 : chargs_color,
				'bg_collors' 		 : bg_collors,

			})

		context={
			'invoice_array' : invoice_array,
			'login_user'	: login_user,
			'cl_array'		: cl_array,
			'min_date' 		: min_date,
			'currency_list' : currency_list
		}
		
		return render_to_response("invoice_display/reconciliation.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')

@csrf_exempt
def reconciliation_details(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		invoice_array 		= []
		remittance_details 	= models.remittance_data.objects.all().order_by('id')
		

		for x in remittance_details:			
			invoice_array.append({
				'id'				 : '',
				'get_id'			 : x.id,
				'remittance_date'	 : str(x.remittance_date), 					
				'certificate_no' 	 : x.certificate_no,
				'inward_amount' 	 : x.inward_amount,
				'currency' 			 : x.currency,
				'rate' 				 : x.rate,
				'amount_inr' 		 : x.inr_amount,
				'status' 			 : x.status,
				'bank_deduction' 	 : x.bank_charges,			
				'net_inward_amt_ref' : x.net_amount_ref,
				'total_amount'		 : x.total_amount,

				'service1'			 : x.services1,
				'invoice_no1'		 : x.invoice_no1,
				'amount1'			 : x.amount1,
				'service2'			 : x.services2,
				'invoice_no2'		 : x.invoice_no2,
				'amount2'			 : x.amount2,
				'service3'			 : x.services3,
				'invoice_no3'		 : x.invoice_no3,
				'amount3'			 : x.amount3,
				'service4'			 : x.services4,
				'invoice_no4'		 : x.invoice_no4,
				'amount4'			 : x.amount4,
				'service5'			 : x.services5,
				'invoice_no5'		 : x.invoice_no5,
				'amount5'			 : x.amount5,
				'service6'			 : x.services6,
				'invoice_no6'		 : x.invoice_no6,
				'amount6'			 : x.amount6,
				'service7'			 : x.services7,
				'invoice_no7'		 : x.invoice_no7,
				'amount7'			 : x.amount7,
				'service8'			 : x.services8,
				'invoice_no8'		 : x.invoice_no8,
				'amount8'			 : x.amount8,
				'service9'			 : x.services9,
				'invoice_no9'		 : x.invoice_no9,
				'amount9'			 : x.amount9,
				'service10'			 : x.services10,
				'invoice_no10'		 : x.invoice_no10,
				'amount10'			 : x.amount10,	
				'client_name'   	 : x.client_name,					
			})

		context={
			'invoice_array' : invoice_array
		}
		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')

def get_total_amount(request):
	amt1   = json.loads(request.GET['amt1'])
	amt2   = json.loads(request.GET['amt2'])
	amt3   = json.loads(request.GET['amt3'])
	amt4   = json.loads(request.GET['amt4'])
	amt5   = json.loads(request.GET['amt5'])
	amt6   = json.loads(request.GET['amt6'])
	amt7   = json.loads(request.GET['amt7'])
	amt8   = json.loads(request.GET['amt8'])
	amt9   = json.loads(request.GET['amt9'])
	amt10  = json.loads(request.GET['amt10'])	
	chrgs  = json.loads(request.GET['chrgs'])
	#if amt1!=None or amt2!=None or amt3!=None or amt4!=None or amt5!=None or amt6!=None or amt7!=None or amt8!=None or amt9!=None or amt10!=None:
	try:
		tot_amt = ((amt1)+(amt2)+(amt3)+(amt4)+(amt5)+(amt6)+(amt7)+(amt8)+(amt9)+(amt10)+(amt11)+(amt12)+(amt13)+(amt14)+(amt15))-(chrgs)
	except:
		tot_amt = 0	
	
	return HttpResponse(json.dumps(tot_amt))


def get_amount_value1(request):
	service1 	= request.GET['service1']
	invoice1 	= request.GET['invoice1']	
	#print '-----------',invoice1
	if service1!=None or invoice1!=None:
		get_client 	= models.invoice.objects.filter(proj_name=service1,invoice_no=invoice1).first()
		get_amount 	= models.invoice.objects.filter(proj_name=service1,invoice_no=invoice1).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service1,invoice_no=invoice1).aggregate(Sum('qty'))		
		
		try:
			if get_client.client_id==112:
				amount1 = get_client.usd_amount
			#elif get_client.client_id==112:
			else:
				amount1		= get_amount['price__sum']
		except:
			amount1 = 0
		
	return HttpResponse(json.dumps(amount1))

def get_amount_value2(request):
	service2 	= request.GET['service2']
	invoice2 	= request.GET['invoice2']	
	if service2!=None or invoice2!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service2,invoice_no=invoice2).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service2,invoice_no=invoice2).aggregate(Sum('qty'))		
		amount2		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount2))		



def get_amount_value3(request):
	service3 	= request.GET['service3']
	invoice3 	= request.GET['invoice3']	
	if service3!=None or invoice3!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service3,invoice_no=invoice3).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service3,invoice_no=invoice3).aggregate(Sum('qty'))		
		amount3		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount3))		

def get_amount_value4(request):
	service4 	= request.GET['service4']
	invoice4 	= request.GET['invoice4']	
	if service4!=None or invoice4!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service4,invoice_no=invoice4).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service4,invoice_no=invoice4).aggregate(Sum('qty'))	
		amount4		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount4))

def get_amount_value5(request):
	service5 	= request.GET['service5']
	invoice5 	= request.GET['invoice5']	
	if service5!=None or invoice5!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service5,invoice_no=invoice5).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service5,invoice_no=invoice5).aggregate(Sum('qty'))	
		amount5		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount5))

def get_amount_value6(request):
	service6 	= request.GET['service6']
	invoice6 	= request.GET['invoice6']	
	if service6!=None or invoice6!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service6,invoice_no=invoice6).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service6,invoice_no=invoice6).aggregate(Sum('qty'))		
		amount6		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount6))


def get_amount_value7(request):
	service7 	= request.GET['service7']
	invoice7 	= request.GET['invoice7']	
	if service7!=None or invoice7!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service7,invoice_no=invoice7).aggregate(Sum('price'))
		get_qty 	= models.invoice.objects.filter(proj_name=service7,invoice_no=invoice7).aggregate(Sum('qty'))
		amount7		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount7))


def get_amount_value8(request):
	service8 	= request.GET['service8']
	invoice8 	= request.GET['invoice8']	
	if service8!=None or invoice8!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service8,invoice_no=invoice8).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service8,invoice_no=invoice8).aggregate(Sum('qty'))	
		amount8		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount8))


def get_amount_value9(request):
	service9 	= request.GET['service9']
	invoice9 	= request.GET['invoice9']	
	if service9!=None or invoice8!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service9,invoice_no=invoice9).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service9,invoice_no=invoice9).aggregate(Sum('qty'))
		amount9		= get_amount['price__sum']

	return HttpResponse(json.dumps(amount9))

def get_amount_value10(request):
	service10 	= request.GET['service10']
	invoice10 	= request.GET['invoice10']
	if service10!=None or invoice10!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service10,invoice_no=invoice10).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service10,invoice_no=invoice10).aggregate(Sum('qty'))			
		amount10	= get_amount['price__sum']

	return HttpResponse(json.dumps(amount10))

def get_amount_value11(request):
	service11 	= request.GET['service11']
	invoice11 	= request.GET['invoice11']
	if service11!=None or invoice11!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service11,invoice_no=invoice11).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service11,invoice_no=invoice11).aggregate(Sum('qty'))			
		amount11	= get_amount['price__sum']

	return HttpResponse(json.dumps(amount11))

def get_amount_value12(request):
	service12 	= request.GET['service12']
	invoice12 	= request.GET['invoice12']
	if service12!=None or invoice12!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service12,invoice_no=invoice12).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service12,invoice_no=invoice12).aggregate(Sum('qty'))			
		amount12	= get_amount['price__sum']
	return HttpResponse(json.dumps(amount12))

def get_amount_value13(request):
	service13 	= request.GET['service13']
	invoice13 	= request.GET['invoice13']
	if service13!=None or invoice13!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service13,invoice_no=invoice13).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service13,invoice_no=invoice13).aggregate(Sum('qty'))			
		amount13	= get_amount['price__sum']

	return HttpResponse(json.dumps(amount13))

def get_amount_value14(request):
	service14 	= request.GET['service14']
	invoice14 	= request.GET['invoice14']
	if service14!=None or invoice14!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service14,invoice_no=invoice14).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service14,invoice_no=invoice14).aggregate(Sum('qty'))			
		amount14	= get_amount['price__sum']

	return HttpResponse(json.dumps(amount14))

def get_amount_value15(request):
	service15 	= request.GET['service15']
	invoice15 	= request.GET['invoice15']
	if service15!=None or invoice15!=None:
		get_amount 	= models.invoice.objects.filter(proj_name=service15,invoice_no=invoice15).aggregate(Sum('price'))	
		get_qty 	= models.invoice.objects.filter(proj_name=service15,invoice_no=invoice15).aggregate(Sum('qty'))			
		amount15	= get_amount['price__sum']

	return HttpResponse(json.dumps(amount15))

def get_status(request):
	amt1   	= json.loads(request.GET['amt1'])
	amt2   	= json.loads(request.GET['amt2'])
	amt3   	= json.loads(request.GET['amt3'])
	amt4   	= json.loads(request.GET['amt4'])
	amt5   	= json.loads(request.GET['amt5'])
	amt6   	= json.loads(request.GET['amt6'])
	amt7   	= json.loads(request.GET['amt7'])
	amt8   	= json.loads(request.GET['amt8'])
	amt9   	= json.loads(request.GET['amt9'])
	amt10  	= json.loads(request.GET['amt10'])	

	amt11  	= json.loads(request.GET['amt11'])	
	amt12  	= json.loads(request.GET['amt12'])	
	amt13  	= json.loads(request.GET['amt13'])	
	amt14  	= json.loads(request.GET['amt14'])	
	amt15  	= json.loads(request.GET['amt15'])	
	

	chrgs  	= json.loads(request.GET['chrgs'])
	amt_ref = json.loads(request.GET['amt_ref'])
	
	try:
		tot_amt = (float(amt1)+float(amt2)+float(amt3)+float(amt4)+float(amt5)+float(amt6)+float(amt7)+float(amt8)+float(amt9)+float(amt10)+float(amt11)+float(amt12)+float(amt13)+float(amt14)+float(amt15)-float(chrgs))	
	except:
		tot_amt = 0
	#print '--------',tot_amt,'-----',amt_ref
	
	if amt_ref==tot_amt:
		msg 	= ""
		colors 	= ''
	else:
		msg 	= ""
		colors 	= "red"
	
	context={
	'msg'		: msg,
	'colors' 	: colors,
	}
	
	return HttpResponse(json.dumps(context))


def edit_remittance_list(request):
	get_id 				= json.loads(request.GET['get_id'])
	remittance_array	= []
	remittance_details 	= models.remittance_data.objects.filter(id=get_id)
	gt_reference_no 	= ''
	#print '---------',get_id
	for x in remittance_details:
		qrty 				= "SELECT reference_no,amount_inr,referencez,txn_date FROM invoice.it_bank_statement where descs like '%"+str(x.certificate_no)+"%'"
		cursor 				= connection.cursor()
		cursor.execute(qrty)
		get_reference_no  	= cursor.fetchall()
		gt_reference_no 	= get_reference_no[0][0]

		if x.services1!=None:
			service1 = x.services1
		else:
			service1 = ''
	
		if x.invoice_no1!=None:
			invoice_no1 = x.invoice_no1
		else:
			invoice_no1 = ''

		if  x.services2!=None:
			service2 = x.services2
		else:
			service2 = ''

		if  x.services3!=None:
			service3 = x.services3
		else:
			service3 = ''

		if x.services4!=None:
			service4 = x.services4
		else:
			service4 = ''

		if x.services5!=None:
			service5 = x.services5
		else:
			service5 = ''

		if x.services6!=None:
			service6 = x.services6
		else:
			service6 = ''

		if x.services7!=None:
			service7 = x.services7
		else:
			service7 = ''

		if x.services8!=None:
			service8 = x.services8
		else:
			service8 = ''

		if x.services9!=None:
			service9 = x.services9
		else:
			service9 = ''

		if x.services10!=None:
			service10 = x.services10
		else:
			service10 = ''

		if x.services11!=None:
			service11 = x.services11
		else:
			service11 = ''


		if x.services12!=None:
			service12 = x.services12
		else:
			service12 = ''


		if x.services13!=None:
			service13 = x.services13
		else:
			service13 = ''


		if x.services14!=None:
			service14 = x.services14
		else:
			service14 = ''


		if x.services15!=None:
			service15 = x.services15
		else:
			service15 = ''


		if x.invoice_no2!=None:
			invoice_no2 = x.invoice_no2
		else:
			invoice_no2 = ''

		if x.invoice_no3!=None:
			invoice_no3 = x.invoice_no3
		else:
			invoice_no3 = ''

		if x.invoice_no4!=None:
			invoice_no4 = x.invoice_no4
		else:
			invoice_no4 = ''

		if x.invoice_no5!=None:
			invoice_no5 = x.invoice_no5
		else:
			invoice_no5 = ''

		if x.invoice_no6!=None:
			invoice_no6 = x.invoice_no6
		else:
			invoice_no6 = ''

		if x.invoice_no7!=None:
			invoice_no7 = x.invoice_no7
		else:
			invoice_no7 = ''

		if x.invoice_no8!=None:
			invoice_no8 = x.invoice_no8
		else:
			invoice_no8 = ''

		if x.invoice_no9!=None:
			invoice_no9 = x.invoice_no9
		else:
			invoice_no9 = ''

		if x.invoice_no10!=None:
			invoice_no10 = x.invoice_no10
		else:
			invoice_no10 = ''

		if x.invoice_no11!=None:
			invoice_no11 = x.invoice_no11
		else:
			invoice_no11 = ''

		if x.invoice_no12!=None:
			invoice_no12 = x.invoice_no12
		else:
			invoice_no12 = ''

		if x.invoice_no13!=None:
			invoice_no13 = x.invoice_no13
		else:
			invoice_no13 = ''


		if x.invoice_no14!=None:
			invoice_no14 = x.invoice_no14
		else:
			invoice_no14 = ''


		if x.invoice_no15!=None:
			invoice_no15 = x.invoice_no15
		else:
			invoice_no15 = ''

		remittance_array.append({			
			'get_id'			 : x.id,
			'remittance_date'	 : str(x.remittance_date), 					
			'certificate_no' 	 : x.certificate_no,
			'inward_amount' 	 : x.inward_amount,
			'currency' 			 : x.currency,
			'rate' 				 : x.rate,
			'amount_inr' 		 : x.inr_amount,
			'status' 			 : x.status,
			'bank_deduction' 	 : x.bank_charges,			
			'net_inward_amt_ref' : x.net_amount_ref,
			'total_amount'		 : x.total_amount,

			'service1'			 : service1,
			'invoice_no1'		 : invoice_no1,
			'amount1'			 : x.amount1,
			'service2'			 : service2,
			'invoice_no2'		 : invoice_no2,
			'amount2'			 : x.amount2,
			'service3'			 : service3,
			'invoice_no3'		 : invoice_no3,
			'amount3'			 : x.amount3,
			'service4'			 : service4,
			'invoice_no4'		 : invoice_no4,
			'amount4'			 : x.amount4,
			'service5'			 : service5,
			'invoice_no5'		 : invoice_no5,
			'amount5'			 : x.amount5,
			'service6'			 : service6,
			'invoice_no6'		 : invoice_no6,
			'amount6'			 : x.amount6,
			'service7'			 : service7,
			'invoice_no7'		 : invoice_no7,
			'amount7'			 : x.amount7,
			'service8'			 : service8,
			'invoice_no8'		 : invoice_no8,
			'amount8'			 : x.amount8,
			'service9'			 : service9,
			'invoice_no9'		 : invoice_no9,
			'amount9'			 : x.amount9,
			'service10'			 : service10,
			'invoice_no10'		 : invoice_no10,
			'amount10'			 : x.amount10,	

			'service11'			 : service11,
			'invoice_no11'		 : invoice_no11,
			'amount11'			 : x.amount11,	
			'service12'			 : service12,
			'invoice_no12'		 : invoice_no12,
			'amount12'			 : x.amount12,	
			'service13'			 : service13,
			'invoice_no13'		 : invoice_no13,
			'amount13'			 : x.amount13,	
			'service14'			 : service14,
			'invoice_no14'		 : invoice_no14,
			'amount14'			 : x.amount14,	
			'service15'			 : service15,
			'invoice_no15'		 : invoice_no15,
			'amount15'			 : x.amount15,	
			

			'client_name'   	 : x.client_name,
			'gt_reference_no2'	 : gt_reference_no,			
		})


	context={
		'remittance_array' : remittance_array
	}

	return HttpResponse(json.dumps(context))

@csrf_exempt
def update_remittance_list(request):	
	e_get_id        		 = json.loads(request.POST['e_get_id'])	
	e_reference_no 			 = json.loads(request.POST['e_reference_no'])
	e_get_client_name        = json.loads(request.POST['e_get_client_name'])
	e_remittance_date        = json.loads(request.POST['e_remittance_date'])
	e_certificate_no         = json.loads(request.POST['e_certificate_no'])
	e_inward_amount        	 = json.loads(request.POST['e_inward_amount'])
	e_currency        		 = json.loads(request.POST['e_currency'])
	e_rate        			 = json.loads(request.POST['e_rate'])
	e_amount_inr        	 = json.loads(request.POST['e_amount_inr'])
	e_get_net_inward_amt_ref = json.loads(request.POST['e_get_net_inward_amt_ref'])
	e_get_status        	 = json.loads(request.POST['e_get_status'])
	e_get_total_amt        	 = json.loads(request.POST['e_get_total_amt'])
	e_get_bank_deduction     = json.loads(request.POST['e_get_bank_deduction'])
	e_service_1        		 = json.loads(request.POST['e_service_1'])
	e_invoice_no_1        	 = json.loads(request.POST['e_invoice_no_1'])
	e_amount_1        		 = json.loads(request.POST['e_amount_1'])
	e_service_2        		 = json.loads(request.POST['e_service_2'])
	e_invoice_no_2        	 = json.loads(request.POST['e_invoice_no_2'])
	e_amount_2        		 = json.loads(request.POST['e_amount_2'])
	e_service_3        		 = json.loads(request.POST['e_service_3'])
	e_invoice_no_3        	 = json.loads(request.POST['e_invoice_no_3'])
	e_amount_3        		 = json.loads(request.POST['e_amount_3'])
	e_service_4        		 = json.loads(request.POST['e_service_4'])
	e_invoice_no_4        	 = json.loads(request.POST['e_invoice_no_4'])
	e_amount_4        		 = json.loads(request.POST['e_amount_4'])
	e_service_5        		 = json.loads(request.POST['e_service_5'])
	e_invoice_no_5        	 = json.loads(request.POST['e_invoice_no_5'])
	e_amount_5        		 = json.loads(request.POST['e_amount_5'])
	e_service_6        		 = json.loads(request.POST['e_service_6'])
	e_invoice_no_6        	 = json.loads(request.POST['e_invoice_no_6'])
	e_amount_6        		 = json.loads(request.POST['e_amount_6'])
	e_service_7        		 = json.loads(request.POST['e_service_7'])
	e_invoice_no_7        	 = json.loads(request.POST['e_invoice_no_7'])
	e_amount_7        		 = json.loads(request.POST['e_amount_7'])
	e_service_8        		 = json.loads(request.POST['e_service_8'])
	e_invoice_no_8        	 = json.loads(request.POST['e_invoice_no_8'])
	e_amount_8        		 = json.loads(request.POST['e_amount_8'])
	e_service_9        		 = json.loads(request.POST['e_service_9'])
	e_invoice_no_9        	 = json.loads(request.POST['e_invoice_no_9'])
	e_amount_9        		 = json.loads(request.POST['e_amount_9'])
	e_service_10        	 = json.loads(request.POST['e_service_10'])
	e_invoice_no_10        	 = json.loads(request.POST['e_invoice_no_10'])
	e_amount_10        		 = json.loads(request.POST['e_amount_10'])
	e_service_11        	 = json.loads(request.POST['e_service_11'])
	e_invoice_no_11        	 = json.loads(request.POST['e_invoice_no_11'])
	e_amount_11        		 = json.loads(request.POST['e_amount_11'])
	e_service_12        	 = json.loads(request.POST['e_service_12'])
	e_invoice_no_12        	 = json.loads(request.POST['e_invoice_no_12'])
	e_amount_12        		 = json.loads(request.POST['e_amount_12'])
	e_service_13        	 = json.loads(request.POST['e_service_13'])
	e_invoice_no_13        	 = json.loads(request.POST['e_invoice_no_13'])
	e_amount_13        		 = json.loads(request.POST['e_amount_13'])
	e_service_14        	 = json.loads(request.POST['e_service_14'])
	e_invoice_no_14        	 = json.loads(request.POST['e_invoice_no_14'])
	e_amount_14        		 = json.loads(request.POST['e_amount_14'])
	e_service_15        	 = json.loads(request.POST['e_service_15'])
	e_invoice_no_15        	 = json.loads(request.POST['e_invoice_no_15'])
	e_amount_15        		 = json.loads(request.POST['e_amount_15'])
	e_tds_deducted       	 = json.loads(request.POST['e_tds_deducted'])
	
	chk1 = models.remittance_data.objects.filter(invoice_no1=e_invoice_no_1,services1=e_service_1,invoice_no2=e_invoice_no_2,services2=e_service_2,invoice_no3=e_invoice_no_3,services3=e_service_3,invoice_no4=e_invoice_no_4,services4=e_service_4,invoice_no5=e_invoice_no_5,services5=e_service_5,invoice_no6=e_invoice_no_6,services6=e_service_6,invoice_no7=e_invoice_no_7,services7=e_service_7,invoice_no8=e_invoice_no_8,services8=e_service_8,invoice_no9=e_invoice_no_9,services9=e_service_9,invoice_no10=e_invoice_no_10,services10=e_service_10,invoice_no11=e_invoice_no_11,services11=e_service_11,invoice_no12=e_invoice_no_12,services12=e_service_12,invoice_no13=e_invoice_no_13,services13=e_service_13,invoice_no14=e_invoice_no_14,services14=e_service_14,invoice_no15=e_invoice_no_15,services15=e_service_15).count()	
	# if chk1!=0:
	# 	return HttpResponse(json.dumps('already'))  	
   #	print '------------',e_get_status
   	db 						= models.remittance_data.objects.filter(id=e_get_id).first()
   	db.remittance_date 		= e_remittance_date
	db.inward_amount 		= e_inward_amount
	db.certificate_no 		= e_certificate_no
	db.currency 			= e_currency
	db.rate 				= e_rate
	db.bank_charges 		= e_get_bank_deduction
	db.status 				= e_get_status
	db.inr_amount 			= e_amount_inr
	db.net_amount_ref 		= e_get_net_inward_amt_ref
	db.total_amount 		= e_get_total_amt
	db.client_name 			= e_get_client_name

	try:
		db.tds_deducted 	= e_tds_deducted
	except:
		db.tds_deducted 	= 0
	
	db.services1    		= e_service_1    
	try:
		db.invoice_no1 		= e_invoice_no_1 
	except:
		db.invoice_no1      = ''
	
	try:
		db.amount1     		= e_amount_1
	except:
		db.amount1     		= 0

	
	db.services2    		= e_service_2    
	try:
		db.invoice_no2 		= e_invoice_no_2 
	except:
		db.invoice_no2 		= '' 
	
	try:
		db.amount2     		= e_amount_2
	except:
		db.amount2     		= 0

	db.services3    		= e_service_3    
	try:
		db.invoice_no3 		= e_invoice_no_3 
	except:
		db.invoice_no3 		= ''

	try:
		db.amount3     		= e_amount_3
	except:
	   db.amount3     		= 0
	
	db.services4    		= e_service_4
	try:
		db.invoice_no4 		= e_invoice_no_4
	except:
		db.invoice_no4 		= ''
	
	try:
		db.amount4     		= e_amount_4
	except:
	   db.amount4     		= 0
	
	db.services5    		= e_service_5    
	try:
		db.invoice_no5 		= e_invoice_no_5 
	except:
		db.invoice_no5 		= ''

	try:
		db.amount5     		= e_amount_5
	except:
	    db.amount5     		= 0
	
	db.services6    		= e_service_6    
	try:
		db.invoice_no6 		= e_invoice_no_6 
	except:
		db.invoice_no6 		= ''
	
	try:
		db.amount6     		= e_amount_6
	except:
	    db.amount6     		= 0 
	
	db.services7    		= e_service_7    
	try:
		db.invoice_no7 		= e_invoice_no_7 
	except:
		db.invoice_no7 		= ''

	try:
		db.amount7     		= e_amount_7
	except:
		db.amount7     		= 0
	
	db.services8    		= e_service_8    
	try:
		db.invoice_no8 		= e_invoice_no_8 
	except:
		db.invoice_no8 		= ''

	try:
		db.amount8     		= e_amount_8
	except:
		db.amount8     		= 0
	
	db.services9    		= e_service_9    
	try:
		db.invoice_no9 		= e_invoice_no_9 
	except:
		db.invoice_no9 		= ''

	try:
		db.amount9     		= e_amount_9
	except:
		db.amount9     		= 0

	
	db.services10   		= e_service_10   
	try:
		db.invoice_no10		= e_invoice_no_10
	except:
		db.invoice_no10		= ''
	
	try:
		db.amount10    		= e_amount_10
	except:
		db.amount10    		= None

	db.services11   		= e_service_11   
	try:
		db.invoice_no11		= e_invoice_no_11
	except:
		db.invoice_no11		= ''
	
	try:
		db.amount11    		= e_amount_11
	except:
		db.amount11    		= 0

	db.services12   		= e_service_12   
	try:
		db.invoice_no12		= e_invoice_no_12
	except:
		db.invoice_no12		= ''

	try:
		db.amount12    		= e_amount_12
	except:
		db.amount12    		= 0

	db.services13   		= e_service_13   
	try:
		db.invoice_no13		= e_invoice_no_13
	except:
		db.invoice_no13		= ''

	try:
		db.amount13    		= e_amount_13
	except:
		db.amount13    		= 0

	db.services14   		= e_service_14   
	try:
		db.invoice_no14		= e_invoice_no_14
	except:
		db.invoice_no14		= ''
	
	try:
		db.amount14    		= e_amount_14
	except:
		db.amount14    		= 0

	db.services15   		= e_service_15   
	try:
		db.invoice_no15		= e_invoice_no_15
	except:
		db.invoice_no15		= ''

	try:
		db.amount15    		= e_amount_15
	except:
		db.amount15    		= 0

	db.reference_no 		= e_reference_no
	db.entry_date 			= datetime.now().date() 
	db.save() 
	return HttpResponse(json.dumps('done'))


def calculation_remittance(request):
	e_n 		= json.loads(request.GET['n'])
	e_service 	= json.loads(request.GET['e_service'])
	e_invoice 	= json.loads(request.GET['e_invoice'])
	e_currency  = json.loads(request.GET['e_currency'])
	
	if e_currency=='INR':
		get_calc 	= models.invoice.objects.filter(proj_name=e_service,invoice_no=e_invoice,inr=e_currency).first()
		get_amount 	= models.invoice.objects.filter(proj_name=e_service,invoice_no=e_invoice,inr=e_currency).aggregate(Sum('price'))
	else:
		get_calc 	= models.invoice.objects.filter(proj_name=e_service,invoice_no=e_invoice).first()
		get_amount 	= models.invoice.objects.filter(proj_name=e_service,invoice_no=e_invoice).aggregate(Sum('price'))

	try:
		month_select = datetime.strptime(str(get_calc.month),"%Y-%m-%d").strftime('%m')
		year_select  = datetime.strptime(str(get_calc.month),"%Y-%m-%d").strftime('%Y')
		get_month    = monthrange(int(year_select),int(month_select))[1]
	except:
		get_month    = 0

	query1 		= "SELECT sum(price*qty)/"+str(get_month)+" FROM invoice.it_invoice where invoice_no='"+str(e_invoice)+"';"	
	cursor 		= connection.cursor()
	cursor.execute(query1)
	totamts 	= cursor.fetchall()
	amountss 	= round(totamts[0][0],0)
	try:
		if get_calc.client_id==112:
			amount = get_calc.usd_amount
			print '--------wala hoga'
		elif get_calc.client_id==20017:
			amount = get_amount['price__sum']
			print '--------22 hoga'
		else:
			if get_calc.qty>1:
				amt_qty = get_amount['price__sum']*float(get_calc.qty)
				amount  = amt_qty
				print '--------usd wala hoga'
			else:
				if e_currency=='INR':
					amount		= get_amount['price__sum']*float(get_calc.rate)
					print '------inr wala hoga'
				else:
					in_amt = get_amount['price__sum']
					amount		= (get_amount['price__sum']*float(get_calc.rate))
					print '-------yahi tha kya'

					
	except:
		amount   = 0
		amountss = 0

	print '---------',get_calc.client_id

	
	
	context={
		'sno' 		: e_n,
		'amount' 	: amount,
	}

	return HttpResponse(json.dumps(context))


def reconciliation_entry(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		login_user 	   = request.user
		currency_list  = []
		query 		= "SELECT id,client_name,proj_name,status FROM invoice.it_client where status='1' group by client_name,proj_name order by client_name,proj_name"
		#print '----',query
		cursor 		= connection.cursor()
		cursor.execute(query)
		client_det 	= cursor.fetchall()
		cl_array    = []
		for t in client_det:					
			cl_array.append({
				'id' 		  : t[0],
				'client_name' : t[1],
				'client_list' : str(t[1])+' ('+str(t[2])+')'
			})

		currency_details = models.currency_data.objects.all().order_by('currency')
		for x in currency_details:
			currency_list.append({
				'id' 			: x.id,
				'currency_name' : x.currency
			})

		context={
		'login_user' 	: login_user,
		'cl_array'   	: cl_array,
		'currency_list' : currency_list,
		}

		return render_to_response("invoice_display/reconciliation_entry.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def save_remittance(request):
	client_name 		= json.loads(request.GET['client_array'])
	try:
		net_inward_ref 	= json.loads(request.GET['net_inward_ref'])
	except:
		net_inward_ref  = 0.0

	remittance_date		= json.loads(request.GET['remittance_dat'])
	
	try:
		status 			= json.loads(request.GET['status'])
	except:
		status 			= 0.0

	invoice_no_5 		= json.loads(request.GET['invoice_no_5'])
	try:
		invoice_no_9 		= json.loads(request.GET['invoice_no_9'])
	except:
		invoice_no_9 = None
	certificate 		= json.loads(request.GET['certificate'])
	bank_charges 		= json.loads(request.GET['bank_charges'])
	amount_5 			= json.loads(request.GET['amount_5'])
	amount_9 			= json.loads(request.GET['amount_9'])
	inward_amount 		= json.loads(request.GET['inward_amount'])
	total_amount 		= json.loads(request.GET['total_amount'])
	service_6 			= json.loads(request.GET['service_6'])
	service_10 			= json.loads(request.GET['service_10'])
	currency 			= json.loads(request.GET['currency'])
	rate 				= json.loads(request.GET['rate'])
	invoice_no_6 		= json.loads(request.GET['invoice_no_6'])
	invoice_no_10 		= json.loads(request.GET['invoice_no_10'])
	amount_inr 			= json.loads(request.GET['amount_inr'])
	amount_6 			= json.loads(request.GET['amount_6'])
	amount_10 			= json.loads(request.GET['amount_10'])
	service_1 			= json.loads(request.GET['service_1'])
	service_3 			= json.loads(request.GET['service_3'])
	service_7 			= json.loads(request.GET['service_7'])
	invoice_no_1 		= json.loads(request.GET['invoice_no_1'])
	invoice_no_3 		= json.loads(request.GET['invoice_no_3'])
	invoice_no_7 		= json.loads(request.GET['invoice_no_7'])
	amount_1 			= json.loads(request.GET['amount_1'])
	amount_3 			= json.loads(request.GET['amount_3'])
	amount_7 			= json.loads(request.GET['amount_7'])
	service_2 			= json.loads(request.GET['service_2'])
	service_4 			= json.loads(request.GET['service_4'])
	service_8 			= json.loads(request.GET['service_8'])
	invoice_no_2 		= json.loads(request.GET['invoice_no_2'])
	invoice_no_4 		= json.loads(request.GET['invoice_no_4'])
	invoice_no_8 		= json.loads(request.GET['invoice_no_8'])
	amount_2 			= json.loads(request.GET['amount_2'])
	amount_4 			= json.loads(request.GET['amount_4'])
	amount_8 			= json.loads(request.GET['amount_8'])
	service_5 			= json.loads(request.GET['service_5'])
	service_9 			= json.loads(request.GET['service_9'])
	reference_no 		= json.loads(request.GET['reference_no'])
	e_remarks 			= json.loads(request.GET['e_remarks'])


	check 	= models.remittance_data.objects.filter(certificate_no=certificate).count()
	#print '-----',check
	if check!=0:
		db =  models.remittance_data.objects.filter(certificate_no=certificate).first()
	else:
		db =  models.remittance_data()

   	db.remittance_date 		= remittance_date
	db.inward_amount 		= inward_amount
	db.certificate_no 		= certificate
	db.currency 			= currency
	db.rate 				= rate
	db.bank_charges 		= bank_charges
	db.status 				= status
	db.inr_amount 			= amount_inr
	try:
		db.net_amount_ref 	= net_inward_ref
	except:
		db.net_amount_ref 	= inward_amount
	db.total_amount 		= total_amount
	db.client_name 			= client_name
	
	db.services1    		= service_1    
	db.invoice_no1 			= invoice_no_1 
	db.amount1     			= amount_1     
	
	db.services2    		= service_2    
	db.invoice_no2 			= invoice_no_2 
	db.amount2     			= amount_2     
	
	db.services3    		= service_3    
	db.invoice_no3 			= invoice_no_3 
	db.amount3     			= amount_3     
	
	db.services4    		= service_4    
	db.invoice_no4 			= invoice_no_4 
	db.amount4     			= amount_4     
	
	db.services5    		= service_5    
	db.invoice_no5 			= invoice_no_5 
	db.amount5     			= amount_5     
	
	db.services6    		= service_6    
	db.invoice_no6 			= invoice_no_6 
	db.amount6     			= amount_6     
	
	db.services7    		= service_7    
	db.invoice_no7 			= invoice_no_7 
	db.amount7     			= amount_7     
	
	db.services8    		= service_8    
	db.invoice_no8 			= invoice_no_8 
	db.amount8     			= amount_8     
	
	db.services9    		= service_9    
	db.invoice_no9 			= invoice_no_9 
	db.amount9     			= amount_9     
	
	db.services10   		= service_10   
	db.invoice_no10			= invoice_no_10
	db.amount10    			= amount_10 
	db.entry_date 			= datetime.now().date() 
	db.reference_no 		= reference_no
	db.remarks 				= e_remarks
	db.save()		
	return HttpResponse(json.dumps('done'))
		
def get_remittance(request):
	arr 				= []
	certificate_value 	= json.loads(request.GET['certificate_value'])
	get_certicate 		= models.master_inw_data.objects.filter(inward_no=certificate_value).first()
	#print '--------',get_certicate
	try:
		qrty 				= "SELECT reference_no,amount_inr,referencez,txn_date FROM invoice.it_bank_statement where descs like '%"+str(certificate_value)+"%'"
		cursor 				= connection.cursor()
		cursor.execute(qrty)
		get_reference_no  	= cursor.fetchall()
		g_reference 		= get_reference_no[0][0]
		g_amt_inr  			= get_reference_no[0][1]
		_rates 				= get_reference_no[0][2]
		g_split_rate 		= _rates.split('@')
		g_rates 			= g_split_rate[1]
		g_currecny 			= str(g_split_rate[0])[:3]	
		g_txn_date 			= get_reference_no[0][3]
		#print '----------',	qrty
		
	except:
		g_reference 		= 0
		g_amt_inr  			= 0
		g_rates 			= 0
		g_currecny 			= 0
		g_txn_date 			= ''

	
	try:
		arr.append({
			'rate' 			: get_certicate.rate,
			'currency' 		: get_certicate.currency,		
			'remit_date' 	: str(get_certicate.txn_date),
			'client_name'	: get_certicate.client_id.client_name,
			'inward_amount' : get_certicate.amount,
			'reference' 	: g_reference,
			'amt_inr'		: g_amt_inr,
			'remarks'		: get_certicate.remarks,
		})
	except:
		arr.append({
			'rate' 			: g_rates,
			'currency' 		: g_currecny,		
			'remit_date' 	: str(g_txn_date),
			'client_name'	: '',
			'inward_amount' : '0',
			'reference' 	: g_reference,
			'amt_inr'		: g_amt_inr,
		})

	return HttpResponse(json.dumps(arr))


def delete_remiitance(request):
	delete_id 	= request.GET['del_id']
	models.remittance_data.objects.filter(id=delete_id).delete()

	return HttpResponse(json.dumps('done'))



def filter_remiitance_data(request):
	stt_date  = json.loads(request.GET['st_date'])
	ett_date  = json.loads(request.GET['et_date'])
	cl_name   = json.loads(request.GET['cl_name'])
	pr_name   = json.loads(request.GET['proj_name'])
	_pro_name = pr_name[-5:]#pr_name.split(' ')
	projnme1  = _pro_name.replace('(','').replace(')','')
	
	proj_name = projnme1
	

	
	login_user 		   	= request.user
	invoice_array 	   	= []
	curr_date 		   	= datetime.now().date()
	gt_prev_year	   	= curr_date.strftime('%Y')
	curr_year 		   	= int(gt_prev_year)
	prev_year 		   	= int(gt_prev_year)
	previous_financial  = str(prev_year)+'-04-01'
	#current_financial   = str(curr_year)+'-03-31'
	current_financial   = '2025-03-31'#str(curr_year)+'-12-31'

	st_date 			= previous_financial
	et_date 			= current_financial
	print '------',st_date,'-----',et_date
	
	if stt_date!='' and ett_date!='' and cl_name=='0':
		#print '-----------1'
		remittance_details 	= models.remittance_data.objects.filter(remittance_date__gte=stt_date,remittance_date__lte=ett_date).order_by('-id')
	elif stt_date!='' and ett_date!='' and cl_name!='0':
		#print '----------2'
		remittance_details 	= models.remittance_data.objects.filter(remittance_date__gte=stt_date,remittance_date__lte=ett_date,client_name=cl_name,services1=proj_name).order_by('-id')
	elif stt_date=='' and ett_date=='' and cl_name!='0':
		#print '-----------3'
		remittance_details 	= models.remittance_data.objects.filter(client_name=cl_name,services1=proj_name).order_by('-id')	
	else:
		#print '----------4'
		remittance_details 	= models.remittance_data.objects.filter(remittance_date__gte=st_date,remittance_date__lte=et_date).order_by('-id')
	
	cl_array 			= []
	query 				= "SELECT id,client_name,proj_name,status FROM invoice.it_client where status='1' group by client_name,proj_name order by client_name,proj_name"
	cursor 				= connection.cursor()
	cursor.execute(query)
	gt_client_det   	= cursor.fetchall()
	for t in gt_client_det:
		cl_array.append({
			'id' 		  : t[0],
			'client_name' : t[1],
			'client_list' : str(t[1])+' ('+str(t[2])+')'
		})



	

	for x in remittance_details:
		qrty 				= "SELECT reference_no,amount_inr,referencez,txn_date FROM invoice.it_bank_statement where descs like '%"+str(x.certificate_no)+"%'"
		cursor 				= connection.cursor()
		cursor.execute(qrty)
		get_reference_no  	= cursor.fetchall()
		try:
			gt_reference_no = get_reference_no[0][0]
		except:
			gt_reference_no = 0

		#print '-----------',x.tds_deducted
		if x.tds_deducted!=None and x.tds_deducted!=0:
			try:
				amt_deduct = ((x.amount1+x.amount2+x.amount3+x.amount4+x.amount5+x.amount6+x.amount7+x.amount8+x.amount9+x.amount10+x.amount11+x.amount12+x.amount13+x.amount14+x.amount15)-x.tds_deducted)
			except:
				amt_deduct = 0
		else:			
			try:
				amt_deduct = ((x.amount1+x.amount2+x.amount3+x.amount4+x.amount5+x.amount6+x.amount7+x.amount8+x.amount9+x.amount10+x.amount11+x.amount12+x.amount13+x.amount14+x.amount15)-x.bank_charges)
			except:
				amt_deduct = 0

		
		
		if x.net_amount_ref==amt_deduct:
			status 			= "Reconciled"
			bg_color 		= "white"
		else:
			status 			= "Not Reconciled"
			bg_color 		= "#f58787"


		if x.bank_charges!=0:
			chargs_color   	= '#f7aba6'
			chrgs 			= x.bank_charges
		else:
			chargs_color 	= ''
			chrgs 			= ''

		try:
			if x.approved=='yes':
				bg_collors = '#c0f7b7'
			else:
				bg_collors = ''
		except:
			pass

		if x.services1!=None:
			service1 = x.services1
		else:
			service1 = ''

		if x.services2!=None:
			service2 = x.services2
		else:
			service2 = ''

		if x.services3!=None:
			service3 = x.services3
		else:
			service3 = ''

		if x.services4!=None:
			service4 = x.services4
		else:
			service4 = ''

		if x.services5!=None:
			service5 = x.services5
		else:
			service5 = ''

		if x.services6!=None:
			service6 = x.services6
		else:
			service6 = ''

		if x.services7!=None:
			service7 = x.services7
		else:
			service7 = ''

		if x.services8!=None:
			service8 = x.services8
		else:
			service8 = ''

		if x.services9!=None:
			service9 = x.services9
		else:
			service9 = ''

		if x.services10!=None:
			service10 = x.services10
		else:
			service10 = ''

		if x.services11!=None:
			service11 = x.services11
		else:
			service11 = ''

		if x.services12!=None:
			service12 = x.services12
		else:
			service12 = ''

		if x.services13!=None:
			service13 = x.services13
		else:
			service13 = ''

		if x.services14!=None:
			service14 = x.services14
		else:
			service14 = ''

		if x.services15!=None:
			service15 = x.services15
		else:
			service15 = ''

		# if x.invoice_no1==None:
		# 	curr = 'white'
		# else:
		# 	curr = 'white'
		if x.tds_deducted!=None and x.tds_deducted!=0:
			tds_ded 		= x.tds_deducted
			chargs_colorx  	= '#e4c2f0'
		else:
			tds_ded 		= ''
			chargs_colorx   = 'white'

		try:
			cl_namess = x.client_name.strip()
		except:
			cl_namess = ''


		
		

		invoice_array.append({				
			'get_id'			 : x.id,
			'remittance_date'	 : str(x.remittance_date), 					
			'certificate_no' 	 : x.certificate_no,
			'inward_amount' 	 : x.inward_amount,
			'currency' 			 : x.currency,
			'rate' 				 : x.rate,
			'amount_inr' 		 : x.inr_amount,
			'status' 			 : status,
			'bank_deduction' 	 : chrgs,
			'net_inward_amt_ref' : x.net_amount_ref,
			'total_amount'		 : amt_deduct,
			'tds_deducted'		 : tds_ded,

			'service1'			 : x.services1,
			'invoice_no1'		 : x.invoice_no1,
			'amount1'			 : x.amount1,
			'service2'			 : service2, #x.services2,
			'invoice_no2'		 : x.invoice_no2,
			'amount2'			 : x.amount2,
			'service3'			 : service3,#x.services3,
			'invoice_no3'		 : x.invoice_no3,
			'amount3'			 : x.amount3,
			'service4'			 : service4,#x.services4,
			'invoice_no4'		 : x.invoice_no4,
			'amount4'			 : x.amount4,
			'service5'			 : service5,#x.services5,
			'invoice_no5'		 : x.invoice_no5,
			'amount5'			 : x.amount5,
			'service6'			 : service6,#x.services6,
			'invoice_no6'		 : x.invoice_no6,
			'amount6'			 : x.amount6,
			'service7'			 : service7,#x.services7,
			'invoice_no7'		 : x.invoice_no7,
			'amount7'			 : x.amount7,
			'service8'			 : service8,#x.services8,
			'invoice_no8'		 : x.invoice_no8,
			'amount8'			 : x.amount8,
			'service9'			 : service9,#x.services9,
			'invoice_no9'		 : x.invoice_no9,
			'amount9'			 : x.amount9,				
			'service10'			 : service10,#x.services10,
			'invoice_no10'		 : x.invoice_no10,
			'amount10'			 : x.amount10,
			'service11'			 : service11,#x.services11,
			'invoice_no11'		 : x.invoice_no11,
			'amount11'			 : x.amount11,	
			'service12'			 : service12,#x.services12,
			'invoice_no12'		 : x.invoice_no12,
			'amount12'			 : x.amount12,	
			'service13'			 : service13,#x.services13,
			'invoice_no13'		 : x.invoice_no13,
			'amount13'			 : x.amount13,	
			'service14'			 : service14,#x.services14,
			'invoice_no14'		 : x.invoice_no14,
			'amount14'			 : x.amount14,	
			'service15'			 : service15,#x.services15,
			'invoice_no15'		 : x.invoice_no15,
			'amount15'			 : x.amount15,	
			

			'client_name'   	 : cl_namess,	
			'bg_color' 			 : bg_color,
			'remit_date'		 : str(x.remittance_date), 
			'gt_reference_no'    : gt_reference_no,		
			'remarks'			 : x.remarks,
			'chargs_color'		 : chargs_color,
			'bg_collors' 		 : bg_collors,		
			'format_date'	 	 : (x.remittance_date).strftime('%d-%b-%Y'), 
			'chargs_colorx' 	 : chargs_colorx,
			#'curr_color'		 : curr, 					
		})

	

	context={
		'invoice_array' 	 : invoice_array,		
		'cl_array'			 : cl_array,
		'previous_financial' : str(previous_financial),
		'current_financial'  : str(current_financial),
	}

	

	return HttpResponse(json.dumps(context))

@csrf_exempt
def export_remiitance_data(request):	
	cl_name 	  		= json.loads(request.POST['client_name'])
	stt_date 	  		= json.loads(request.POST['start_date'])
	ett_date 	  		= json.loads(request.POST['end_date'])

	curr_date 			= datetime.now().date()
	gt_prev_year		= curr_date.strftime('%Y')
	curr_year 			= int(gt_prev_year)
	prev_year 			= int(gt_prev_year)-1
	previous_financial  = str(prev_year)+'-04-01'
	current_financial   = str(curr_year)+'-03-31'

	st_date 			= previous_financial
	et_date 			= current_financial

	select_dir    		= os.path.dirname(__file__)
	srcfile       		= select_dir+'/static/reconciliation.xlsx'
	dstroot       		= select_dir+'/reconciliation.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      		= os.path.dirname(__file__)		
	location      		= (my_dir+'/reconciliation.xlsx')
	wb 	   	      		= load_workbook(my_dir+'/reconciliation.xlsx')
	ws_usd        		= wb.get_sheet_by_name("Reconciliation")
	thin_border   		= Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      		= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	ii   		  		= 5
	jj 	 		  		= 0

	pr_name   = json.loads(request.POST['proj_name'])
	_pro_name = pr_name.split(' ')
	projnme1  = _pro_name[1].replace('(','').replace(')','')	
	proj_name = projnme1


	if stt_date!='' and ett_date!='' and cl_name=='0':		
		remittance_details 	= models.remittance_data.objects.filter(remittance_date__gte=stt_date,remittance_date__lte=ett_date).order_by('-id')
	elif stt_date!='' and ett_date!='' and cl_name!='0':		
		remittance_details 	= models.remittance_data.objects.filter(remittance_date__gte=stt_date,remittance_date__lte=ett_date,client_name=cl_name,services1=proj_name).order_by('-id')
	elif stt_date=='' and ett_date=='' and cl_name!='0':
		remittance_details 	= models.remittance_data.objects.filter(client_name=cl_name,services1=proj_name).order_by('-id')		
	else:	
		remittance_details 	= models.remittance_data.objects.filter(remittance_date__gte=st_date,remittance_date__lte=et_date).order_by('-id')
		
	thin_border    = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style       = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	_row_style     = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='center'))

	for c in remittance_details:
		rem_date 						= c.remittance_date.strftime('%d-%b-%Y')		
		ws_usd.cell('A'+str(ii)).value 	= jj+1
		ws_usd.cell('B'+str(ii)).value 	= c.client_name
		ws_usd.cell('C'+str(ii)).value 	= rem_date
		ws_usd.cell('D'+str(ii)).value 	= c.certificate_no
		ws_usd.cell('E'+str(ii)).value 	= c.inward_amount
		ws_usd.cell('F'+str(ii)).value 	= c.currency

		if c.rate!=0:
			ws_usd.cell('G'+str(ii)).value 	= c.rate
		else:
			ws_usd.cell('G'+str(ii)).value 	= ""
		
		if c.inr_amount!=0:
			ws_usd.cell('H'+str(ii)).value 	= c.inr_amount
		else:
			ws_usd.cell('H'+str(ii)).value 	= ""

		ws_usd.cell('I'+str(ii)).value 	= c.net_amount_ref

		if c.total_amount!=0:
			ws_usd.cell('J'+str(ii)).value 	= c.total_amount
		else:
			ws_usd.cell('J'+str(ii)).value 	= ""

		

		if c.bank_charges!=0:
			ws_usd.cell('L'+str(ii)).value 	= c.bank_charges
		else:
			ws_usd.cell('L'+str(ii)).value 	= ""
			

		# try:
		# 	if c.tds_deducted!=0:
		# 		tot_sum = ((c.amount1+c.amount2+c.amount3+c.amount4+c.amount5+c.amount6+c.amount7+c.amount8+c.amount9+c.amount10+c.amount11+c.amount12+c.amount13+c.amount14+c.amount15)-c.tds_deducted)
		# 	else:
		# 		tot_sum = 0
		#Mohd_Arbaaz	
		#Boss@123
		# except:
		# 	tot_sum = 0
		
		
		
		#ws_usd.cell('K'+str(ii)).value 	= 'Reconciled' #c.status
		ws_usd.cell('M'+str(ii)).value 	= c.tds_deducted

		ws_usd.cell('N'+str(ii)).value 	= c.services1
		ws_usd.cell('O'+str(ii)).value 	= c.invoice_no1
		ws_usd.cell('P'+str(ii)).value 	= c.amount1

		ws_usd.cell('Q'+str(ii)).value 	= c.services2
		ws_usd.cell('R'+str(ii)).value 	= c.invoice_no2
		ws_usd.cell('S'+str(ii)).value 	= c.amount2

		ws_usd.cell('T'+str(ii)).value 	= c.services3
		ws_usd.cell('U'+str(ii)).value 	= c.invoice_no3
		ws_usd.cell('V'+str(ii)).value 	= c.amount3

		ws_usd.cell('W'+str(ii)).value 	= c.services4
		ws_usd.cell('X'+str(ii)).value 	= c.invoice_no4
		ws_usd.cell('Y'+str(ii)).value 	= c.amount4

		ws_usd.cell('Z'+str(ii)).value 		= c.services5
		ws_usd.cell('AA'+str(ii)).value 	= c.invoice_no5
		ws_usd.cell('AB'+str(ii)).value 	= c.amount5

		ws_usd.cell('AC'+str(ii)).value 	= c.services6
		ws_usd.cell('AD'+str(ii)).value 	= c.invoice_no6
		ws_usd.cell('AE'+str(ii)).value 	= c.amount6

		ws_usd.cell('AF'+str(ii)).value 	= c.services7
		ws_usd.cell('AG'+str(ii)).value 	= c.invoice_no7
		ws_usd.cell('AH'+str(ii)).value 	= c.amount7

		ws_usd.cell('AI'+str(ii)).value 	= c.services8
		ws_usd.cell('AJ'+str(ii)).value 	= c.invoice_no8
		ws_usd.cell('AK'+str(ii)).value 	= c.amount8

		ws_usd.cell('AL'+str(ii)).value 	= c.services9
		ws_usd.cell('AM'+str(ii)).value 	= c.invoice_no9
		ws_usd.cell('AN'+str(ii)).value 	= c.amount9


		ws_usd.cell('AO'+str(ii)).value 	= c.services10
		ws_usd.cell('AP'+str(ii)).value 	= c.invoice_no10
		ws_usd.cell('AQ'+str(ii)).value 	= c.amount10


		ws_usd.cell('AR'+str(ii)).value 	= c.services11
		ws_usd.cell('AS'+str(ii)).value 	= c.invoice_no11
		ws_usd.cell('AT'+str(ii)).value 	= c.amount11


		ws_usd.cell('AU'+str(ii)).value 	= c.services12
		ws_usd.cell('AV'+str(ii)).value 	= c.invoice_no12
		ws_usd.cell('AW'+str(ii)).value 	= c.amount12


		ws_usd.cell('AX'+str(ii)).value 	= c.services13
		ws_usd.cell('AY'+str(ii)).value 	= c.invoice_no13
		ws_usd.cell('AZ'+str(ii)).value 	= c.amount13


		ws_usd.cell('BA'+str(ii)).value 	= c.services14
		ws_usd.cell('BB'+str(ii)).value 	= c.invoice_no14
		ws_usd.cell('BC'+str(ii)).value 	= c.amount14

		ws_usd.cell('BD'+str(ii)).value 	= c.services15
		ws_usd.cell('BE'+str(ii)).value 	= c.invoice_no15
		ws_usd.cell('BF'+str(ii)).value 	= c.amount15

		ws_usd.cell('A'+str(ii)).style = _row_style
		ws_usd.cell('B'+str(ii)).style = _row_style	 
		ws_usd.cell('C'+str(ii)).style = _row_style
		ws_usd.cell('D'+str(ii)).style = _row_style
		ws_usd.cell('E'+str(ii)).style = _row_style
		ws_usd.cell('F'+str(ii)).style = _row_style
		ws_usd.cell('G'+str(ii)).style = _row_style
		ws_usd.cell('H'+str(ii)).style = _row_style
		ws_usd.cell('I'+str(ii)).style = _row_style
		ws_usd.cell('J'+str(ii)).style = _row_style
		ws_usd.cell('K'+str(ii)).style = _row_style
		ws_usd.cell('L'+str(ii)).style = _row_style
		ws_usd.cell('M'+str(ii)).style = _row_style		

		ws_usd.cell('N'+str(ii)).style = _row_style
		ws_usd.cell('O'+str(ii)).style = _row_style
		ws_usd.cell('P'+str(ii)).style = _row_style

		ws_usd.cell('Q'+str(ii)).style = _row_style
		ws_usd.cell('R'+str(ii)).style = _row_style
		ws_usd.cell('S'+str(ii)).style = _row_style

		ws_usd.cell('T'+str(ii)).style = _row_style
		ws_usd.cell('U'+str(ii)).style = _row_style
		ws_usd.cell('V'+str(ii)).style = _row_style

		ws_usd.cell('W'+str(ii)).style = _row_style
		ws_usd.cell('X'+str(ii)).style = _row_style
		ws_usd.cell('Y'+str(ii)).style = _row_style

		ws_usd.cell('Z'+str(ii)).style = _row_style
		ws_usd.cell('AA'+str(ii)).style = _row_style
		ws_usd.cell('AB'+str(ii)).style = _row_style

		ws_usd.cell('AC'+str(ii)).style = _row_style
		ws_usd.cell('AD'+str(ii)).style = _row_style
		ws_usd.cell('AE'+str(ii)).style = _row_style

		ws_usd.cell('AF'+str(ii)).style = _row_style
		ws_usd.cell('AG'+str(ii)).style = _row_style
		ws_usd.cell('AH'+str(ii)).style = _row_style

		ws_usd.cell('AI'+str(ii)).style = _row_style
		ws_usd.cell('AJ'+str(ii)).style = _row_style
		ws_usd.cell('AK'+str(ii)).style = _row_style

		ws_usd.cell('AL'+str(ii)).style = _row_style
		ws_usd.cell('AM'+str(ii)).style = _row_style
		ws_usd.cell('AN'+str(ii)).style = _row_style

		ws_usd.cell('AO'+str(ii)).style = _row_style
		ws_usd.cell('AP'+str(ii)).style = _row_style
		ws_usd.cell('AQ'+str(ii)).style = _row_style

		ws_usd.cell('AR'+str(ii)).style = _row_style
		ws_usd.cell('AS'+str(ii)).style = _row_style
		ws_usd.cell('AT'+str(ii)).style = _row_style

		ws_usd.cell('AU'+str(ii)).style = _row_style
		ws_usd.cell('AV'+str(ii)).style = _row_style
		ws_usd.cell('AW'+str(ii)).style = _row_style

		ws_usd.cell('AX'+str(ii)).style = _row_style
		ws_usd.cell('AY'+str(ii)).style = _row_style
		ws_usd.cell('AZ'+str(ii)).style = _row_style

		ws_usd.cell('BA'+str(ii)).style = _row_style
		ws_usd.cell('BB'+str(ii)).style = _row_style
		ws_usd.cell('BC'+str(ii)).style = _row_style

		ws_usd.cell('BD'+str(ii)).style = _row_style
		ws_usd.cell('BE'+str(ii)).style = _row_style
		ws_usd.cell('BF'+str(ii)).style = _row_style

		ws_usd.cell('E'+str(ii)).number_format = '0.0'
		ws_usd.cell('G'+str(ii)).number_format = '0.0'
		ws_usd.cell('H'+str(ii)).number_format = '0.0'
		ws_usd.cell('I'+str(ii)).number_format = '0.0'
		ws_usd.cell('J'+str(ii)).number_format = '0.0'
		ws_usd.cell('L'+str(ii)).number_format = '0.0'
		ws_usd.cell('M'+str(ii)).number_format = '0.0'

		ws_usd.cell('P'+str(ii)).number_format = '0.0'
		ws_usd.cell('S'+str(ii)).number_format = '0.0'
		ws_usd.cell('V'+str(ii)).number_format = '0.0'

		ws_usd.cell('Y'+str(ii)).number_format = '0.0'
		ws_usd.cell('AB'+str(ii)).number_format = '0.0'
		ws_usd.cell('AE'+str(ii)).number_format = '0.0'
		ws_usd.cell('AH'+str(ii)).number_format = '0.0'
		ws_usd.cell('AK'+str(ii)).number_format = '0.0'
		ws_usd.cell('AN'+str(ii)).number_format = '0.0'
		ws_usd.cell('AQ'+str(ii)).number_format = '0.0'
		ws_usd.cell('AT'+str(ii)).number_format = '0.0'
		ws_usd.cell('AW'+str(ii)).number_format = '0.0'
		ws_usd.cell('AZ'+str(ii)).number_format = '0.0'
		ws_usd.cell('BC'+str(ii)).number_format = '0.0'
		ws_usd.cell('BF'+str(ii)).number_format = '0.0'


		ii+=1
		jj+=1

	wb.save(my_dir+'/reconciliation.xlsx')

	return HttpResponse(json.dumps('done'))


def export_reconciliation_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/reconciliation.xlsx')
	filename 	   					= my_dir+'/reconciliation.xlsx'
	download_name  					= "reconciliation.xlsx"	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response