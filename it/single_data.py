from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################

def single_data(request):
	if request.user.is_authenticated():		
		login_user 		= request.user
		
		
		context={
		'login_user' :  login_user
		}			
		return render_to_response("invoice_display/single_data.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')

@csrf_exempt
def update_single_invoice(request):
	if request.user.is_authenticated():
		table_details = json.loads(request.POST['table_details'])
		for f in table_details:
			update_db 				= models.invoice.objects.filter(id=f['id'],invoice_no=f['invoice_no']).order_by('ship_name').first()
			update_db.id 			= f['id']
			update_db.qty 			= f['day']
			update_db.price 		= f['price']
			update_db.invoice_no 	= f['invoice_no']
			update_db.ship_name 	= f['ship_name']
			# update_db.heading 		= f['heading']
			update_db.save()

		return HttpResponse(json.dumps('done'))
	else:
		return HttpResponseRedirect('/it/user_login')
	

@csrf_exempt
def single(request):
	if request.user.is_authenticated():
		type_invoice_no 	= json.loads(request.GET['type_invoice_no'])[0]
		#print '-----------',type_invoice_no
		
		try:
			get_id = models.invoice.objects.filter(client_id=112,invoice_no=type_invoice_no).order_by('-id').first()
			get_invoice_date = str(get_id.invoice_date)
		except:
			get_id 			 = ""
			get_invoice_date = ''

		try:
			period_date 	= get_id.month
		except:
			period_date 	= ''

		get_invoice_date 	= get_invoice_date		
		gt_splt 			= get_invoice_date.split('-')
		gt_month 			= (gt_splt[2][:2])

		cal_invoice_date 	= period_date		
		cl_splt 			= period_date.split('-')
		cl_month 			= (cl_splt[2][:2])
		cl_year 			= 24#(cl_splt[0][:1])
		

		invoice_details 	= models.invoice.objects.filter(client_id=112,invoice_date=get_invoice_date,invoice_no=type_invoice_no).order_by('ship_name')
		inv_det 			= []
		#get_month 			= int(gt_month)#31
		calc_price 			= 0.0
		s 					= 0
		get_price 			= get_id.price #invoice_details[0].client.price
		get_month   		= 31#monthrange(int(cl_year),int(cl_month))[1]
		print '------------>>>',cl_month,'---',cl_year

		for c in invoice_details:
			calc_price  = ((c.qty*get_price)/31) #(get_month))
			#print '=====',get_price
			s+=calc_price
			inv_det.append({
				'id'			: c.id,
				'ship_name' 	: c.ship_name,
				'invoice_no' 	: c.invoice_no,
				'day'			: c.qty,
				'price'			: c.price,
				'calc_price'	: calc_price
				})

			
			dt_date = (get_invoice_date)[:10]
			split_d = (dt_date).split('-')
			get_pro_date = str(split_d[2])+'-'+str(split_d[1])+'-'+str(split_d[0])
			
		
		
		context={
			'total_amt'	: s,
			'inv_det'	: inv_det,
			'get_date' 	: get_pro_date,
		}

		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')