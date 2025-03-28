from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

#####################################################
import socket
import platform
import os

def fin_bluewater(request):
	if request.user.is_authenticated():
		bluewater_log 	= models.BlueWater.objects.all()
		tag_bwtw 		= []

		for c in bluewater_log:
			tag_bwtw.append({
				'c_id' 		  	: c.id,
				'tag_tin_name' 	: c.tin_number,
				'client_name'  	: c.client_name,				
			})

		f_invoice_no 		= models.financial_invoice_no.objects.all().order_by('-id').first()
		get_fin_invoice_no  = f_invoice_no.fin_invoice_no
		login_user 			= request.user

		context={
			'login_user' 		: login_user,	
			'tag_bwtw' 			: tag_bwtw,
			'tag_invoice_no' 	: get_fin_invoice_no,
		}	

		return render_to_response("invoice_display/fin_bluewater.html",context)
	else:
		return HttpResponseRedirect('/it/user_login')	



def update_fin_bluewater(request):	
	b_invoice_no  		= request.GET['b_invoice_no']	
	up 					= models.financial_invoice_no.objects.filter(id=2).first()
	up.fin_invoice_no 	= b_invoice_no
	up.save()

	return HttpResponse(json.dumps('done'))


@csrf_exempt
def gstin_details(request):
	bluewater_log 	= models.BlueWater.objects.all().order_by('client_name')
	tag_bwtw 		= []

	for c in bluewater_log:
		tag_bwtw.append({
			'c_id' 		  	: c.id,
			'tag_tin_name' 	: c.tin_number,
			'client_name'  	: c.client_name,
			'proj_name'		: c.proj_name,				
		})


	context={
		'tag_bwtw'  : tag_bwtw,
	}	
	
	return HttpResponse(json.dumps(context))

@csrf_exempt
def gst_details_entry(request):	
	tabs_name=json.loads(request.POST['tabs_name'])	
	for t in tabs_name:	
		
		check = models.BlueWater.objects.filter(id=t['c_id']).count()
		if check==1:
			net = models.BlueWater.objects.filter(id=t['c_id']).first()
		else:		
			net = models.BlueWater()		

		if t['c_id']!=None or t['client_name']!=None or t['tag_tin_name']!=None or t['proj_name']:
		 	net.id 			= t['c_id']
		 	net.client_name = t['client_name']
		 	net.tin_number 	= t['tag_tin_name']
		 	net.proj_name   = t['proj_name']		
		  	net.save()
			
	return HttpResponse(json.dumps("done"))