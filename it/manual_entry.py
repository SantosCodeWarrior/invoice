from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

#####################################################
import socket
import platform
import os

def manual_entry(request):
	if request.user.is_authenticated():
		
		
		print socket.gethostname()
	

		return render_to_response("invoice_display/manual_edit.html")
	else:
		return HttpResponseRedirect('/it/user_login')	


def invoice_details_chm(request):
	invoice_no = json.loads(request.GET['invoice_no'])[0]
	invoice_id = json.loads(request.GET['invoice_id'])[0]

	get_invoice_deta = models.invoice.objects.filter(invoice_no=invoice_no).first()
	disching_date 	 = get_invoice_deta.disch_date.date()

	context={
		'qty' 				: get_invoice_deta.qty,
		'price'				: get_invoice_deta.price,
		'vm_name'			: get_invoice_deta.vm_name,
		'voyage_no'   		: get_invoice_deta.voyage_no,
		'disch_date'  		: str(disching_date),
		'disch_port'  		: get_invoice_deta.disch_port,
		'vessel_name' 		: get_invoice_deta.ship_name,
		'total_amount' 		: get_invoice_deta.total_amount,
		'nomination_date'  	: str(get_invoice_deta.nomination_date),
	}
	return HttpResponse(json.dumps(context))


def update_invoice_details_chm(request):
	invoice_no 	 	= json.loads(request.GET['_invoice_no'])
	vessel_name 	= json.loads(request.GET['_vessel_name'])
	voyage_no 		= json.loads(request.GET['_voyage_no'])
	disch_port 	 	= json.loads(request.GET['_disch_port'])
	disch_date 	 	= json.loads(request.GET['_disch_date'])
	nomination_date = json.loads(request.GET['_nomination_date'])
	vm_name 		= json.loads(request.GET['_vm_name'])
	invoice_id 		= json.loads(request.GET['invoice_id'])

	_qty 			= json.loads(request.GET['_qty'])
	_tot_amount 	= json.loads(request.GET['_tot_amount'])
	_price 			= json.loads(request.GET['_price'])


	Update_invoice  = models.invoice.objects.filter(id=invoice_id,invoice_no=invoice_no).first()
	
	Update_invoice.disch_port 		= disch_port
	Update_invoice.disch_date 		= disch_date
	Update_invoice.nomination_date 	= nomination_date
	Update_invoice.ship_name 		= vessel_name
	Update_invoice.voyage_no 		= voyage_no
	Update_invoice.id 				= invoice_id
	Update_invoice.invoice_no    	= invoice_no
	Update_invoice.vm_name 			= vm_name
	Update_invoice.total_amount 	= _tot_amount
	Update_invoice.qty 				= _qty
	Update_invoice.price 			= _price
	Update_invoice.save()  	


	return HttpResponse(json.dumps('done'))