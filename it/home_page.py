from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################
def home(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		curr_date_invoice 	= datetime.now().date()
		login_user 			= request.user
		#curr_date_invoice   = '2023-12-31'
		tot_boss_client 	= models.boss_client.objects.filter(status=1).count()
		boss_non_active 	= models.boss_client.objects.all().count()
		tot_chm_client 		= models.Client.objects.filter(proj_name='CHM',status=1).count()
		chm_non_active 		= models.Client.objects.filter(proj_name='CHM').count()
		
		check_curr 	  		= models.invoice.objects.filter(invoice_date=curr_date_invoice).count()		
		if check_curr>1:
			show_table = ''
		else:
			show_table = 'none'

		curr_invoice 	  	= models.invoice.objects.filter(invoice_date=curr_date_invoice).order_by('invoice_no')
		curr_array 		  	= []
		inv_arr 			= []
		for x in curr_invoice:
			if x.invoice_no not in inv_arr:
				inv_arr.append(x.invoice_no)

		for b in inv_arr:
			invoice_log    = models.invoice.objects.filter(invoice_no=b).first()
			curr_array.append({
				'proj_name'    : invoice_log.proj_name,
				'client_name'  : invoice_log.client.client_name,
				'invoice_no'   : invoice_log.invoice_no,
				'invoice_date' : str(invoice_log.invoice_date.date()),
				})
		

		context={
		
		'chm'		 : str(tot_chm_client)+' of '+str(chm_non_active),
		'boss'		 : str(tot_boss_client)+' of '+str(boss_non_active),
		'login_user' : login_user,
		'curr_array' : curr_array,
		'show_table' : show_table,
		}
		return render_to_response("invoice_display/home.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')


