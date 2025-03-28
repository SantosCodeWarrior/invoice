from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
#####################################################

def conv_rate(request):
	if request.user.is_authenticated():
		login_user = request.user
		
		context={
			'login_user' 		: login_user,
		}
		
		return render_to_response("invoice_display/conv_rate.html",context)
	else:
		return HttpResponseRedirect('/it/user_login')

def insert_for_conv_rate(request):
	if request.user.is_authenticated():
		table_details = json.loads(request.GET['table_details'])
		for c in table_details:
			if c[1]!=None or c[2]!=None or c[3]!=None or c[4]!=None or c[5]!=None or c[6]!=None or c[0]!=None:
				CHK = models.conversion_rate.objects.filter(id=c[0]).count()
				if CHK!=0:
					db  = models.conversion_rate.objects.filter(id=c[0]).first()
				else:
					db  = models.conversion_rate()

				frate_date    	= datetime.strptime(c[1], "%d-%b-%Y").strftime('%Y-%m-%d')
				db.id 			= c[0]
				db.rate_date 	= frate_date
				db.rate_price 	= c[2]
				db.rate_open 	= c[3]
				db.rate_high 	= c[4]
				db.entry_date 	= datetime.now().date()
				db.rate_low 	= c[5]
				db.rate_avg 	= c[6]
				db.remarks 	    = None
				db.save()
		return HttpResponse(json.dumps('done'))
	else:
		return HttpResponseRedirect('/it/user_login')


def conv_rate_list(request):
	if request.user.is_authenticated():
		rate_details = models.conversion_rate.objects.all().order_by('-rate_date')
		rate_array   = []
		for c in rate_details:
			rate_array.append({
				'id' 			: c.id,
				'rate_date' 	: (c.rate_date).strftime('%d-%b-%Y'),
				'rate_price' 	: c.rate_price,
				'rate_open' 	: c.rate_open,
				'rate_high' 	: c.rate_high,
				'rate_low' 		: c.rate_low,
				'rate_avg' 		: c.rate_avg,
			})
			
		context={
			'rate_array' : rate_array,
		}
		
		return HttpResponse(json.dumps(context))
	else:
		return HttpResponseRedirect('/it/user_login')
