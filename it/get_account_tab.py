from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection
##############################################
##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt

from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

#####################################################
def get_account_tab(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		login_user 	= request.user

		

		context={
			'login_user' : login_user,
		}
		
		return render_to_response("invoice_display/get_account_tab.html",context)
	else:
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
		return HttpResponseRedirect('/it/user_login')


@csrf_exempt
def account_fetch_details(request):
	print '-0fdfdf'
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		api_l 		   = "https://aboss.bwesglobal.com/api/get_static_details/"		
		api_l_method   = "POST"			
		parameters_l   = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b"}
		responses      = requests.get(api_l, params=parameters_l,verify=False)		
		cl_array       = json.loads(responses.content)
		cl_dets 	   = cl_array['ship_details']
		account_array  = []
		current_dates  = datetime.now().date()
		for key in cl_dets:
			patn 	   = re.sub(r"[\([{})\]]", "", str(cl_dets[key]["clients_id"]))			
			client_id  = patn	
			
			if patn=="64":
				CHK = models.api_client_data.objects.filter(client_id=client_id,ship_name=cl_dets[key]["ship_id"],curr_date=current_dates).count()
				if CHK>1:
					db = models.api_client_data.objects.filter(client_id=client_id,ship_name=cl_dets[key]["ship_id"],curr_date=current_dates).first()
				else:
					db = models.api_client_data()

				
				db.account_tab 	= cl_dets[key]["account_tab"]
				db.ship_name   	= cl_dets[key]["ship_name"]
				#db.ship_id 		= cl_dets[key]["id"]
				db.client_id   	= client_id
				db.address 		= cl_dets[key]["address"]
				db.curr_date 	= str(current_dates)
				db.save()


		account_array.append({
			'account_tab' 	: cl_dets[key]["account_tab"],
			'ship_name'		: cl_dets[key]["ship_name"],
			'ship_id'		: 'cl_dets[key]["id"]',
			'address'		: cl_dets[key]["address"],
			'curr_date'     : str(current_dates),
			'client_id' 	: patn,

		})	
		print '-----',	account_array	

		return HttpResponse(json.dumps(account_array))
	else:
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
		return HttpResponseRedirect('/it/user_login')

		