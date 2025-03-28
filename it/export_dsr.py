from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection

from it import apj_template as apj
from it import clearlake_template as ct
from it import shell_template as sh
from it import vessel_template as vt
from django.db.models import Q
import time

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
#####################################################
import urllib
from bs4 import BeautifulSoup
import json

def export_dsr_details(request):
	clicks 				= request.GET['clicks']
	#name  	   			= json.loads(request.GET['client_lt'])	
	#start_date 			= json.loads(request.GET['start_dtt'])
	#end_date   			= json.loads(request.GET['end_dtt'])
	#ship_type  		= json.loads(request.GET['vessel_type'])
	#start_date_format 	= datetime.strptime(start_date,"%m/%d/%Y").strftime('%Y-%m-%d')
	#end_date_format   	= datetime.strptime(end_date,"%m/%d/%Y").strftime('%Y-%m-%d')
	# api_url 	    	= "https://chm.bwesglobal.com/hb/get_chm_data/"
	# api_method      	= "GET"
	# parameters      	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': name,'ship_type':ship_type,'start_date':str(start_date_format),'end_date': str(end_date_format)}
	# response        	= requests.get(api_url, params=parameters,verify=False)
	# ch_array        	= json.loads(response.content)
	# ch_array        	= ch_array['finance_array']
	# api_url 	    	= "https://chm.bwesglobal.com/hb/get_chm_data/"
	# api_method      	= "GET"
	# parameters      	= {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': name,'ship_type':ship_type,'start_date':str(start_date_format),'end_date': str(end_date_format)}
	# response        	= requests.get(api_url, params=parameters,verify=False)
	# ch_array        	= json.loads(response.content)
	# ch_array        	= ch_array['finance_array']
	chm_array       	= []
	s_no 		    	= 1
	ship_name       	= 0
	voyage_no       	= 0
	disch_port     	 	= 0
	disch_date      	= 0
	vm_name 	    	= 0
	voyage_id       	= 0
	s_no            	= 0
	address 	    	= 0
	color           	= 0
	href            	= 0
	active          	= 0
	clientid        	= 0
	alert_ship_name 	= ''
	ii 					= 2
	jj 					= 0
	
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/CHM_DSR.xlsx'
	dstroot       = select_dir+'/CHM_DSR.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/CHM_DSR.xlsx')
	wb 	   	      = load_workbook(my_dir+'/CHM_DSR.xlsx')
	ws_usd        = wb.get_sheet_by_name("CHM")
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	ii   		  = 2
	jj 	 		  = 1
	
	single_client = []
	get_client    = models.Client.objects.filter(proj_name='CHM',status=1)
	for w in get_client:
		if w.client_name not in single_client:
			single_client.append(w.client_name)



	#print '-----',sorted(single_client)
	today 		   = datetime.now().date()
	format_year    = (today).strftime('%Y')
	format_month   = (today).strftime('%b')
	int_mon		   = (today).strftime('%m')
	format_date    = (today).strftime('%d')
	int_mont 	   = int(int_mon)
	mrge_date 	   = str(format_date)+'-'+str(format_month)+'-'+str(format_year)#+'_'+str(clicks)
	#print '-----------',mrge_date
	
	if int_mont<=9:
		int_month = "0"+str(int_mont)
	else:
		int_month = str(int_mont)

	get_month 		= monthrange(int(format_year),int(int_month))[1]	
	stat_dt_format 	= '2025-02-01' #str(format_year)+'-'+str(int_month)+'-01'
	end_dt_format 	= '2025-03-31' #str(format_year)+'-'+str(int_month)+'-'+str(get_month)
	#print '-----',stat_dt_format,'----',end_dt_format
	#client_details 	= ['Cargill','Litasco','Litasco_Dubai','Navig8','Phillips 66','Shell','Shell NWE','Stena Bulk','Stena Bulk Veg Oil','Teekay','Ultranav','Union Maritime']
	ship_type 		= '0'
	#format_fin_date = ''
	# try:
	# 	models.merge_data.objects.filter(client='Shell NWE').delete()
	# except:
	# 	pass

	for c in sorted(single_client):		
		api_url 	    = "https://chm.bwesglobal.com/hb/get_chm_data/"	
		api_method      = "GET"
		parameters      = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': c,'ship_type':ship_type,'start_date':str(stat_dt_format),'end_date': str(end_dt_format)}
		response        = requests.get(api_url, params=parameters,verify=False)
		ch_array        = json.loads(response.content)
		ch_array        = ch_array['finance_array']
		single_array 	= []
		
		for t in ch_array:
		 	if t['finance_date']>=stat_dt_format and t['finance_date']<=end_dt_format:
				try:
					clientx   	 = models.Client.objects.filter(client_name=c,proj_name='CHM').first()
					pool_lisx 	 = models.pool_master.objects.filter(client_id=clientx.id,pool=ship_type).first()			
					ship_details = models.Ship.objects.filter(ship_name=t['ship_name']).first()
					
					if c=='Shell NWE':
						dbs 			= models.merge_data()
						shipname 		= t['ship_name']
						voyageno 		= t['voyage_no']
						dbs.ship_name 	= shipname
						dbs.voyage_no 	= voyageno
						dbs.client 		= c
						dbs.save()
			
					try:
						missing_ship  = ship_details.ship_name
						missing_color = 'black'
					except:
						missing_ship  = t['ship_name']
						missing_color = 'black' #

					if t['account_tab']!=None:
						account_slab  = t['account_tab']
					else:
						account_slab  = ''

					if ship_type!='0':
						if t['pool_namex']!=pool_lisx.pool:
							remove_space_vessel = t['ship_name'].replace(' ', '_')
							ship_name  = t['ship_name']
							voyage_no  = t['voyage_no']
							disch_port = t['disch_port']
							try:
								disch_date = datetime.strptime(t['disch_date'], "%Y-%m-%d").strftime('%d-%b-%Y') #t['disch_date']
							except:
								disch_date = ''
								
							vm_name    = t['vm_name']
							voyage_id  = t['voyage_id']					
							s_no 	   = s_no
							address    = t['account_address'] 
							color 	   = ''
							href 	   = '#'
							active 	   = 'style="pointer-events: none;cursor: default;"'
							clientid   = clientx.id
							accoun_tab = account_slab
														
							try:
								format_fin_date = datetime.strptime(str(t['disch_date']), "%Y-%m-%d").strftime('%d-%b-%Y')
							except:
								format_fin_date = ''

							try:
								check_cancel     = models.invoice.objects.filter(ship_name=t['ship_name'],voyage_no=t['voyage_no'],payment_status='Cancel',cancel_invoice='1',vessel_type=ship_type).first()
								cancel_vessel    = check_cancel.ship_name
								cancel_voyage_no = check_cancel.voyage_no
								cancelo 		 = check_cancel.cancel_invoice
							except:
								cancel_vessel    = ''
								cancel_voyage_no = ''
								cancelo  		 = '0'

							chm_array.append({
								'ship_name'       : t['ship_name'],
								'voyage_no'       : t['voyage_no'],
								'disch_port'      : t['disch_port'],						
								'vm_name' 	      : t['vm_name'],
								'client_id'       : clientx.id,
								'voyage_id'       : t['voyage_id'],
								's_no' 		      : s_no,
								'alert_color'     : color,
								'alert_ship_name' : alert_ship_name,
								'alert_href' 	  : href,
								'alert_active'    : active,
								'disch_date'	  : str(format_fin_date),
								'missing_ship'    : missing_ship,
								'missing_color'   : missing_color,
								'mix_vessel'	  : remove_space_vessel,
								'cancel_vessel'   : cancel_vessel,
								'cancel_voyage_no': cancel_voyage_no,
								'cancelo'  		  : cancelo,
								'account_slab'    : accoun_tab,
								'nomination_date' : t['finance_date'],
								})
							s_no+=1			

					thin_border    				   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
					my_style       				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
					_row_style     				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
					row_alignment 				   = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='center'))

					try:
						format_fin_date = datetime.strptime(str(t['disch_date']), "%Y-%m-%d").strftime('%d-%b-%Y')
					except:
						format_fin_date = ''

					try:
						format_nomination_date = datetime.strptime(str(t['finance_date']), "%Y-%m-%d").strftime('%d-%b-%Y')
					except:
						format_nomination_date = ''
					
					ship_names = t['ship_name']
					voyage_nos = t['voyage_no']

					ws_usd.cell('A'+str(ii)).value = jj
					ws_usd.cell('B'+str(ii)).value = ship_names
					ws_usd.cell('C'+str(ii)).value = voyage_nos 
					ws_usd.cell('D'+str(ii)).value = t['vm_name']
					ws_usd.cell('E'+str(ii)).value = t['disch_port']		
					ws_usd.cell('F'+str(ii)).value = str(format_fin_date)
					
					if c=='Litasco_Dubai':
						ws_usd.cell('G'+str(ii)).value = t['account_tab']
						ws_usd.cell('J'+str(ii)).value = t['ref_name']
					else:
						ws_usd.cell('G'+str(ii)).value = t['account_tab']
						ws_usd.cell('J'+str(ii)).value = ""

					try:
						ws_usd.cell('K'+str(ii)).value = t['edit_reasons']
					except:
						ws_usd.cell('K'+str(ii)).value = ""

					ws_usd.cell('H'+str(ii)).value = c
					ws_usd.cell('I'+str(ii)).value = format_nomination_date
					ws_usd.cell('A'+str(ii)).style = row_alignment	 
					ws_usd.cell('B'+str(ii)).style = _row_style	 
					ws_usd.cell('C'+str(ii)).style = _row_style	 
					ws_usd.cell('D'+str(ii)).style = _row_style	 
					ws_usd.cell('E'+str(ii)).style = _row_style	 
					ws_usd.cell('F'+str(ii)).style = row_alignment		
					ws_usd.cell('G'+str(ii)).style = _row_style	
					ws_usd.cell('H'+str(ii)).style = _row_style	 
					ws_usd.cell('I'+str(ii)).style = row_alignment	
					ws_usd.cell('J'+str(ii)).style = _row_style	 
					ws_usd.cell('K'+str(ii)).style = _row_style	 	
					ii+=1
					jj+=1
					datef = mrge_date	
					wb.save(my_dir+'/DSR/CHM_DSR_'+str(mrge_date)+'.xlsx')					
				except:
					pass				
			else:
				try:
					if t['account_tab']!=None:
						account_slab = t['account_tab']
					else:
						account_slab = ''

					clien_x    = models.Client.objects.filter(client_name=c,proj_name='CHM').first()
					pool_lisx  = models.pool_master.objects.filter(client_id=clien_x.id).first()
					if t['client_name']==c:
						ship_details = models.Ship.objects.filter(ship_name=t['ship_name']).first()
						try:
							missing_ship  = ship_details.ship_name
							missing_color = 'black'
						except:
							missing_ship  = t['ship_name']
							missing_color = 'black'

						remove_space_vessel = t['ship_name'].replace(' ', '_')
						ship_name           = t['ship_name']
						voyage_no           = t['voyage_no']
						disch_port          = t['disch_port']
						try:
							disch_date = datetime.strptime(t['disch_date'], "%Y-%m-%d").strftime('%d-%b-%Y') #t['disch_date']
						except:
							disch_date = ''
						
						vm_name             = t['vm_name'],
						voyage_id           = t['voyage_id']
						s_no 	            = s_no
						address             = t['account_address']
						color 	            = ''
						href 	            = '#'
						active 	            = 'style="pointer-events: none;cursor: default;"'
						clientid            = clientx.id				

						try:
							format_fin_date = datetime.strptime(t['disch_date'],"%Y-%m-%d").strftime('%d-%b-%Y')
						except:
							format_fin_date = ''

						try:
							check_cancel     = models.invoice.objects.filter(ship_name=t['ship_name'],voyage_no=t['voyage_no'],payment_status='Cancel',cancel_invoice='1').first()
							cancel_vessel    = check_cancel.ship_name
							cancel_voyage_no = check_cancel.voyage_no
							cancelo 		 = check_cancel.cancel_invoice
						except:
							cancel_vessel    = ''
							cancel_voyage_no = ''
							cancelo  		 = '0'					
					
						chm_array.append({
							'ship_name'       : t['ship_name'],
							'voyage_no'       : t['voyage_no'],
							'disch_port'      : t['disch_port'],									
							'vm_name' 	      : t['vm_name'],
							'client_id'       : clientx.id,
							'voyage_id'       : t['voyage_id'],
							's_no' 		      : s_no,
							'alert_color'     : color,
							'alert_ship_name' : alert_ship_name,
							'alert_href' 	  : href,
							'alert_active'    : active,
							'disch_date'	  : str(format_fin_date),
							'cancel_vessel'   : cancel_vessel,
							'cancel_voyage_no': cancel_voyage_no,
							'cancelo'  		  : cancelo,
							'missing_ship'    : missing_ship,
							'missing_color'   : missing_color,
							'mix_vessel'	  : remove_space_vessel,
							'account_slab'    : account_slab,
							'nomination_date' : t['finance_date']
							})

						s_no+=1
				except:
					pass
		
	return HttpResponse(json.dumps('done'))
	
	

def export_dsr_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	today 							= datetime.now().date()
	format_year 					= (today).strftime('%Y')
	format_month 					= (today).strftime('%b')
	int_mon							= (today).strftime('%m')
	format_date 					= (today).strftime('%d')
	int_mont 						= int(int_mon)
	mrge_date 						= str(format_date)+'-'+str(format_month)+'-'+str(format_year)	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/DSR/CHM_DSR_'+str(mrge_date)+'.xlsx')
	filename 	   					= my_dir+'/DSR/CHM_DSR_'+str(mrge_date)+'.xlsx'
	download_name  					= 'CHM_DSR_'+str(mrge_date)+'.xlsx'	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response