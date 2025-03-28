from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
#####################################################


def rec_inr(request):
	if request.user.is_authenticated():
		login_user = request.user	
		

		context={		
			'login_user'   : login_user,		
		}
		return render_to_response('invoice_display/rec_inr.html',context)
	else:
		return HttpResponseRedirect('/it/user_login')

def calc_remittance_datas_inr(request):
	e_n 		= json.loads(request.GET['n'])	
	e_invoice 	= json.loads(request.GET['invoice_no'])	
	amountss    = 0
	query1 		= "SELECT sum(ins.price*.ins.qty*ins.rate),cl.client_name FROM invoice.it_invoice ins,invoice.it_client cl where (cl.id=ins.client_id) and ins.invoice_no='"+str(e_invoice)+"';"	
	cursor 		= connection.cursor()
	cursor.execute(query1)
	totamts 	= cursor.fetchall()
	qry_deta    = []
	try:
		amt = round(totamts[0][0],1)
	except:
		amt = 0.0

	qry_deta.append({
		'amountss' 		: amt,
		'client_name1' 	: totamts[0][1],
	})
	


	context={
	'qry_deta' : qry_deta
	}
	
	
	return HttpResponse(json.dumps(context))


def submitting_inr_data(request):
	remitt_date  		= json.loads(request.GET['remitt_date'])
	reference_no  		= json.loads(request.GET['reference_no'])
	amount_received  	= json.loads(request.GET['amount_received'])
	status  			= json.loads(request.GET['status'])
	total_remitt  		= json.loads(request.GET['total_remitt'])
	charges_any  		= json.loads(request.GET['charges_any'])
	remitt_amount  		= json.loads(request.GET['remitt_amount'])
	
	invoice_no1  		= json.loads(request.GET['invoice_no1'])
	client_name1  		= json.loads(request.GET['client_name1'])
	principle1  		= json.loads(request.GET['principle1'])
	igst1  				= json.loads(request.GET['igst1'])
	igst_value1  		= json.loads(request.GET['igst_value1'])
	tds1  				= json.loads(request.GET['tds1'])
	tds_amount1  		= json.loads(request.GET['tds_amount1'])
	tds_igst1  			= json.loads(request.GET['tds_igst1'])
	tds_igst_amt1  		= json.loads(request.GET['tds_igst_amt1'])
	principle_tds_ded1  = json.loads(request.GET['principle_tds_ded1'])
	remitt1  			= json.loads(request.GET['remitt1'])
	
	invoice_no2  		= json.loads(request.GET['invoice_no2'])
	client_name2  		= json.loads(request.GET['client_name2'])
	principle2  		= json.loads(request.GET['principle2'])
	igst2  				= json.loads(request.GET['igst2'])
	igst_value2  		= json.loads(request.GET['igst_value2'])
	tds2  				= json.loads(request.GET['tds2'])
	tds_amount2  		= json.loads(request.GET['tds_amount2'])
	tds_igst2  			= json.loads(request.GET['tds_igst2'])
	tds_igst_amt2  		= json.loads(request.GET['tds_igst_amt2'])
	principle_tds_ded2  = json.loads(request.GET['principle_tds_ded2'])
	remitt2  			= json.loads(request.GET['remitt2'])
	
	invoice_no3  		= json.loads(request.GET['invoice_no3'])
	client_name3  		= json.loads(request.GET['client_name3'])
	principle3  		= json.loads(request.GET['principle3'])
	igst3  				= json.loads(request.GET['igst3'])
	igst_value3  		= json.loads(request.GET['igst_value3'])
	tds3  				= json.loads(request.GET['tds3'])
	tds_amount3  		= json.loads(request.GET['tds_amount3'])
	tds_igst3  			= json.loads(request.GET['tds_igst3'])
	tds_igst_amt3  		= json.loads(request.GET['tds_igst_amt3'])
	principle_tds_ded3  = json.loads(request.GET['principle_tds_ded3'])
	remitt3  			= json.loads(request.GET['remitt3'])
	
	invoice_no4  		= json.loads(request.GET['invoice_no4'])
	client_name4  		= json.loads(request.GET['client_name4'])
	principle4  		= json.loads(request.GET['principle4'])
	igst4  				= json.loads(request.GET['igst4'])
	igst_value4  		= json.loads(request.GET['igst_value4'])
	tds4  				= json.loads(request.GET['tds4'])
	tds_amount4 	 	= json.loads(request.GET['tds_amount4'])
	tds_igst4  			= json.loads(request.GET['tds_igst4'])
	tds_igst_amt4  		= json.loads(request.GET['tds_igst_amt4'])
	principle_tds_ded4  = json.loads(request.GET['principle_tds_ded4'])
	remitt4  			= json.loads(request.GET['remitt4'])
	
	invoice_no5  		= json.loads(request.GET['invoice_no5'])
	client_name5  		= json.loads(request.GET['client_name5'])
	principle5  		= json.loads(request.GET['principle5'])
	igst5  				= json.loads(request.GET['igst5'])
	igst_value5  		= json.loads(request.GET['igst_value5'])
	tds5  				= json.loads(request.GET['tds5'])
	tds_amount5  		= json.loads(request.GET['tds_amount5'])
	tds_igst5  			= json.loads(request.GET['tds_igst5'])
	tds_igst_amt5  		= json.loads(request.GET['tds_igst_amt5'])
	principle_tds_ded5  = json.loads(request.GET['principle_tds_ded5'])
	remitt5  			= json.loads(request.GET['remitt5'])
	
	invoice_no6  		= json.loads(request.GET['invoice_no6'])
	client_name6  		= json.loads(request.GET['client_name6'])
	principle6  		= json.loads(request.GET['principle6'])
	igst6  				= json.loads(request.GET['igst6'])
	igst_value6  		= json.loads(request.GET['igst_value6'])
	tds6  				= json.loads(request.GET['tds6'])
	tds_amount6  		= json.loads(request.GET['tds_amount6'])
	tds_igst6  			= json.loads(request.GET['tds_igst6'])
	tds_igst_amt6  		= json.loads(request.GET['tds_igst_amt6'])
	principle_tds_ded6  = json.loads(request.GET['principle_tds_ded6'])
	remitt6  			= json.loads(request.GET['remitt6'])
	
	invoice_no7  		= json.loads(request.GET['invoice_no7'])
	client_name7  		= json.loads(request.GET['client_name7'])
	principle7  		= json.loads(request.GET['principle7'])
	igst7  				= json.loads(request.GET['igst7'])
	igst_value7  		= json.loads(request.GET['igst_value7'])
	tds7  				= json.loads(request.GET['tds7'])
	tds_amount7  		= json.loads(request.GET['tds_amount7'])
	tds_igst7  			= json.loads(request.GET['tds_igst7'])
	tds_igst_amt7  		= json.loads(request.GET['tds_igst_amt7'])
	principle_tds_ded7  = json.loads(request.GET['principle_tds_ded7'])
	remitt7  			= json.loads(request.GET['remitt7'])
	
	invoice_no8  		= json.loads(request.GET['invoice_no8'])
	client_name8  		= json.loads(request.GET['client_name8'])
	principle8  		= json.loads(request.GET['principle8'])
	igst8  				= json.loads(request.GET['igst8'])
	igst_value8  		= json.loads(request.GET['igst_value8'])
	tds8  				= json.loads(request.GET['tds8'])
	tds_amount8  		= json.loads(request.GET['tds_amount8'])
	tds_igst8  			= json.loads(request.GET['tds_igst8'])
	tds_igst_amt8  		= json.loads(request.GET['tds_igst_amt8'])
	principle_tds_ded8  = json.loads(request.GET['principle_tds_ded8'])
	remitt8  			= json.loads(request.GET['remitt8'])
	
	invoice_no9  		= json.loads(request.GET['invoice_no9'])
	client_name9  		= json.loads(request.GET['client_name9'])
	principle9  		= json.loads(request.GET['principle9'])
	igst9  				= json.loads(request.GET['igst9'])
	igst_value9  		= json.loads(request.GET['igst_value9'])
	tds9  				= json.loads(request.GET['tds9'])
	tds_amount9  		= json.loads(request.GET['tds_amount9'])
	tds_igst9  			= json.loads(request.GET['tds_igst9'])
	tds_igst_amt9  		= json.loads(request.GET['tds_igst_amt9'])
	principle_tds_ded9  = json.loads(request.GET['principle_tds_ded9'])
	remitt9  			= json.loads(request.GET['remitt9'])
	
	invoice_no10  		= json.loads(request.GET['invoice_no10'])
	client_name10  		= json.loads(request.GET['client_name10'])
	principle10  		= json.loads(request.GET['principle10'])
	igst10  			= json.loads(request.GET['igst10'])
	igst_value10  		= json.loads(request.GET['igst_value10'])
	tds10  				= json.loads(request.GET['tds10'])
	tds_amount10  		= json.loads(request.GET['tds_amount10'])
	tds_igst10  		= json.loads(request.GET['tds_igst10'])
	tds_igst_amt10  	= json.loads(request.GET['tds_igst_amt10'])
	principle_tds_ded10 = json.loads(request.GET['principle_tds_ded10'])
	remitt10  			= json.loads(request.GET['remitt10'])

	remitt_amount2 		= json.loads(request.GET['remitt_amount2'])
	remitt_amount3 		= json.loads(request.GET['remitt_amount3'])
	remitt_amount4 		= json.loads(request.GET['remitt_amount4'])
	remitt_amount5 		= json.loads(request.GET['remitt_amount5'])
	remitt_amount6 		= json.loads(request.GET['remitt_amount6'])
	remitt_amount7 		= json.loads(request.GET['remitt_amount7'])
	remitt_amount8 		= json.loads(request.GET['remitt_amount8'])
	remitt_amount9 		= json.loads(request.GET['remitt_amount9'])
	remitt_amount10 	= json.loads(request.GET['remitt_amount10'])
	

	CHK = models.remittance_data_inr.objects.filter(reference_no=reference_no).count()
	if CHK!=0:
		db = models.remittance_data_inr.objects.filter(reference_no=reference_no).first()
	else:
		db = models.remittance_data_inr()

	

	db.remittance_date		= remitt_date
	db.reference_no 		= reference_no
	db.amount_received 		= amount_received
	db.status 				= status
	db.total_remitt 		= total_remitt
	db.charges_any 			= charges_any
	db.remitt_amount 		= remitt_amount
	
	if invoice_no1!='':
		db.invoice_no1 			= invoice_no1
	else:
		db.invoice_no1 			= ""
	
	if client_name1!='':
		db.client_name1 		= client_name1
	else:
		db.client_name1 		= ""
	
	if principle1!='':
		db.principle1 			= principle1
	else:
		db.principle1 			= 0.0	
	
	if igst1!='':
		db.igst1 				= igst1
	else:
		db.igst1 				= ""
	
	if igst_value1!='':
		db.igst_value1 			= igst_value1
	else:
		db.igst_value1 			= 0.0	
	
	if tds1!='':
		db.tds1 				= tds1
	else:
		db.tds1 				= 0.0	
	
	if  tds_amount1!='':
		db.tds_amount1 			= tds_amount1
	else:
		db.tds_amount1 			= 0.0	
	
	if tds_igst1!='':
		db.tds_igst1 			= tds_igst1
	else:
		db.tds_igst1 			= 0.0	
	
	if tds_igst_amt1!='':
		db.tds_igst_amt1 		= tds_igst_amt1
	else:
		db.tds_igst_amt1 		= 0.0	
	
	if principle_tds_ded1!='':
		db.principle_tds_ded1 	= principle_tds_ded1
	else:
		db.principle_tds_ded1 	= 0.0	
	
	if remitt1!='':
		db.remitt1 				= remitt1
	else:
		db.remitt1  			= ""
		
	
	if invoice_no2!='':
		db.invoice_no2 			= invoice_no2
	else:
		db.invoice_no2 			= ""

	if client_name2!="":
		db.client_name2 		= client_name2
	else:
		db.client_name2 	 	= ""


	if remitt_amount2!='':
		db.remitt_amount2 = remitt_amount2
	else:
		db.remitt_amount2 = 0.0
	
	if remitt_amount3!='':
		db.remitt_amount3 = remitt_amount3
	else:
		db.remitt_amount3 = 0.0
	
	if remitt_amount4!='':
		db.remitt_amount4 = remitt_amount4
	else:
		db.remitt_amount4 = 0.0
	
	if remitt_amount5!='':
		db.remitt_amount5 = remitt_amount5
	else:
		db.remitt_amount5 = 0.0
	
	if remitt_amount6!='':
		db.remitt_amount6 = remitt_amount6
	else:
		db.remitt_amount6 = 0.0
	
	if remitt_amount7!='':
		db.remitt_amount7 = remitt_amount7
	else:
		db.remitt_amount7 = 0.0
	
	if remitt_amount8!='':
		db.remitt_amount8 = remitt_amount8
	else:
		db.remitt_amount8 = 0.0
	
	if remitt_amount9!='':
		db.remitt_amount9 = remitt_amount9
	else:
		db.remitt_amount9 = 0.0
	
	if remitt_amount10!='':
		db.remitt_amount10 = remitt_amount10
	else:
		db.remitt_amount10 = 0.0

	if principle2!='':
		db.principle2 			= principle2
	else:
		db.principle2 			= 0.0

	if igst2!='':
		db.igst2 				= igst2
	else:
		db.igst2  				= ""

	if igst_value2!='': 	
		db.igst_value2 			= igst_value2
	else:
		db.igst_value2 			= 0
	
	if tds2!='':
		db.tds2 				= tds2
	else:
		db.tds2 				= 0.0
	
	if tds_amount2!='':  
		db.tds_amount2 			= tds_amount2
	else:
		db.tds_amount2  		= 0.0

	if tds_igst2!='':  
		db.tds_igst2 			= tds_igst2
	else:
		db.tds_igst2 			= 0.0

	if tds_igst_amt2!='':  
		db.tds_igst_amt2 		= tds_igst_amt2
	else:
		db.tds_igst_amt2 		= 0.0

	if principle_tds_ded2!='': 
		db.principle_tds_ded2 	= principle_tds_ded2
	else:
		db.principle_tds_ded2 	= 0.0

	if remitt2!='':  
		db.remitt2 				= remitt2
	else:
		db.remitt2 				= ""	
	
	if invoice_no3!='':  
		db.invoice_no3 			= invoice_no3
	else:
		db.invoice_no3 			= ""

	if client_name3!='':  
		db.client_name3 		= client_name3
	else:
		db.client_name3 		= ""

	if principle3!='':  
		db.principle3 			= principle3
	else:
		db.principle3 			= 0.0

	if igst3!='':  
		db.igst3 				= igst3
	else:
		db.igst3 				= ""

	if igst_value3!='':  
		db.igst_value3 			= igst_value3
	else:
		db.igst_value3 			= 0.0

	if tds3!='':
		db.tds3 				= tds3
	else:
		db.tds3 				= 0.0
		
	if tds_amount3!='': 
		db.tds_amount3 			= tds_amount3
	else:
		db.tds_amount3 			= 0.0

	if tds_igst3!='': 
		db.tds_igst3 			= tds_igst3
	else:
		db.tds_igst3 			= 0.0

	if tds_igst_amt3!='': 
		db.tds_igst_amt3 		= tds_igst_amt3
	else:
		db.tds_igst_amt3 		= 0.0

	if principle_tds_ded3!='': 
		db.principle_tds_ded3 	= principle_tds_ded3
	else:
		db.principle_tds_ded3 	= 0.0

	if remitt3!='':  
		db.remitt3 				= remitt3
	else:
		db.remitt3 		      	= ""

	if invoice_no4!='':  
		db.invoice_no4 			= invoice_no4
	else:
		db.invoice_no4 			= ""

	if  client_name4!='':
		db.client_name4 		= client_name4
	else:
		db.client_name4 		= ""

	if principle4!='': 
		db.principle4 			= principle4
	else:
		db.principle4 			= 0.0

	if igst4!='':  
		db.igst4 				= igst4
	else:
		db.igst4 				= ""

	if igst_value4!='': 
		db.igst_value4 			= igst_value4
	else:
		db.igst_value4 			= 0.0

	if tds4!='':
		db.tds4 				= tds4
	else:
		db.tds4 				= 0.0

	if tds_amount4!='':  
		db.tds_amount4 			= tds_amount4
	else:
		db.tds_amount4 			= 0.0

	if tds_igst4!='':  
		db.tds_igst4 			= tds_igst4
	else:
		db.tds_igst4 			= 0.0
	
	if tds_igst_amt4!='':   
		db.tds_igst_amt4 		= tds_igst_amt4
	else:
		db.tds_igst_amt4 		= 0.0

	if principle_tds_ded4!='': 
		db.principle_tds_ded4 	= principle_tds_ded4
	else:
		db.principle_tds_ded4 	= 0.0
	
	if remitt4!='': 
		db.remitt4 				= remitt4
	else:
		db.remitt4 				= ""

	if invoice_no5!='':  
		db.invoice_no5 			= invoice_no5
	else:
		db.invoice_no5 			= ""
	
	if client_name5!="":  
		db.client_name5 		= client_name5
	else:
		db.client_name5 		= ""

	if principle5!='': 
		db.principle5 			= principle5
	else:
		db.principle5 			= 0.0

	if igst5!='':  
		db.igst5 				= igst5
	else:
		db.igst5 				= ""

	if igst_value5!='': 
		db.igst_value5 			= igst_value5
	else:
		db.igst_value5 			= 0.0

	if tds5!='': 
		db.tds5 				= tds5
	else:
		db.tds5 				= 0.0

	if tds_amount5!='': 
		db.tds_amount5 			= tds_amount5
	else:
		db.tds_amount5 			= 0.0

	if tds_igst5!='': 
		db.tds_igst5 			= tds_igst5
	else:
		db.tds_igst5 			= 0.0
	
	if tds_igst_amt5!='': 
		db.tds_igst_amt5 		= tds_igst_amt5
	else:
		db.tds_igst_amt5 		= 0.0

	if principle_tds_ded5!='': 
		db.principle_tds_ded5 	= principle_tds_ded5
	else:
		db.principle_tds_ded5 	= 0.0

	if remitt5!='':  
		db.remitt5 				= remitt5
	else:
		db.remitt5 				= ""

	if invoice_no6!='':  
		db.invoice_no6 			= invoice_no6
	else:
		db.invoice_no6 			= ""

	if client_name6!="": 
		db.client_name6 		= client_name6
	else:
		db.client_name6 		= ""

	if principle6!='':  
		db.principle6 			= principle6
	else:
		db.principle6 	 		= 0.0

	if igst6!='': 
		db.igst6 				= igst6
	else:
		db.igst6 				= ""

	if igst_value6!='': 
		db.igst_value6 			= igst_value6
	else:
		db.igst_value6 			= 0.0

	if tds6!='': 
		db.tds6 				= tds6
	else:
		db.tds6 				= 0.0

	if tds_amount6!='': 
		db.tds_amount6 			= tds_amount6
	else:
		db.tds_amount6 			= 0.0

	if tds_igst6!='': 
		db.tds_igst6 			= tds_igst6
	else:
		db.tds_igst6 			= 0.0

	if tds_igst_amt6!='': 
		db.tds_igst_amt6 		= tds_igst_amt6
	else:
		db.tds_igst_amt6 		= 0.0

	if principle_tds_ded6!='': 
		db.principle_tds_ded6 	= principle_tds_ded6
	else:
		db.principle_tds_ded6 	= 0.0

	if remitt6!='': 
		db.remitt6 				= remitt6
	else:
		db.remitt6 				= ""

	if invoice_no7!='': 
		db.invoice_no7 			= invoice_no7
	else:
		db.invoice_no7 		 	= ""

	if client_name7!='': 
		db.client_name7 		= client_name7
	else:
		db.client_name7 		= ""

	if principle7!='': 
		db.principle7 			= principle7
	else:
		db.principle7 		 	= 0.0

	if igst7!='': 
		db.igst7 				= igst7
	else:
		db.igst7 		 		= ""

	if igst_value7!='': 
		db.igst_value7 			= igst_value7
	else:
		db.igst_value7 			= 0.0

	if tds7!='': 
		db.tds7 				= tds7
	else:
		db.tds7 				= 0.0

	if tds_amount7!='': 
		db.tds_amount7 			= tds_amount7
	else:
		db.tds_amount7 		 	= 0.0

	if tds_igst7!='': 
		db.tds_igst7 			= tds_igst7
	else:
		db.tds_igst7 		 	= 0.0

	if tds_igst_amt7!='': 
		db.tds_igst_amt7 		= tds_igst_amt7
	else:
		db.tds_igst_amt7 		= 0.0

	if principle_tds_ded7!='': 
		db.principle_tds_ded7 	= principle_tds_ded7
	else:
		db.principle_tds_ded7 	= 0.0

	if remitt7!='': 
		db.remitt7 				= remitt7
	else:
		db.remitt7 		 		= 0.0

	if invoice_no8!='': 
		db.invoice_no8 			= invoice_no8
	else:
		db.invoice_no8 		 	= ""

	if client_name8!='': 
		db.client_name8 		= client_name8
	else:
		db.client_name8 		 = ""

	if principle8!='': 
		db.principle8 			= principle8
	else:
		db.principle8 		 	= 0.0

	if igst8!='': 
		db.igst8 				= igst8
	else:
		db.igst8 		 		= ""

	if igst_value8!='': 
		db.igst_value8 			= igst_value8
	else:
		db.igst_value8 		 	= 0.0

	if tds8!='': 
		db.tds8 				= tds8
	else:
		db.tds8 		 		= 0.0

	if tds_amount8!='': 
		db.tds_amount8 			= tds_amount8
	else:
		db.tds_amount8 			= 0.0
		
	if tds_igst8!='': 
		db.tds_igst8 			= tds_igst8
	else:
		db.tds_igst8 			= 0.0
		
	if tds_igst_amt8!='': 
		db.tds_igst_amt8 		= tds_igst_amt8
	else:
		db.tds_igst_amt8 		= 0.0
		
	if principle_tds_ded8!='': 
		db.principle_tds_ded8 	= principle_tds_ded8
	else:
		db.principle_tds_ded8 	= 0.0
		
	if remitt8!='': 
		db.remitt8 				= remitt8
	else:
		db.remitt8 				= 0.0
		
	if invoice_no9!='': 
		db.invoice_no9 			= invoice_no9
	else:
		db.invoice_no9 			= ""
		
	if client_name9!='': 
		db.client_name9 		= client_name9
	else:
		db.client_name9 		= ""
		
	if principle9!='': 
		db.principle9 			= principle9
	else:
		db.principle9 			= 0.0
		
	if igst9!='': 
		db.igst9 				= igst9
	else:
		db.igst9 				= ""
		
	if igst_value9!='': 
		db.igst_value9 			= igst_value9
	else:
		db.igst_value9 			= 0.0
		
	if tds9!='': 
		db.tds9 				= tds9
	else:
		db.tds9 				= 0.0
		
	if tds_amount9!='': 
		db.tds_amount9 			= tds_amount9
	else:
		db.tds_amount9 			= 0.0
		
	if tds_igst9!='': 
		db.tds_igst9 			= tds_igst9
	else:
		db.tds_igst9 			= 0.0
		
	if tds_igst_amt9!='': 
		db.tds_igst_amt9 		= tds_igst_amt9
	else:
		db.tds_igst_amt9 		= 0.0
		
	if principle_tds_ded9!='': 
		db.principle_tds_ded9 	= principle_tds_ded9
	else:
		db.principle_tds_ded9 	= 0.0
		
	if remitt9!='': 
		db.remitt9 				= remitt9
	else:
		db.remitt9 				= 0.0
		
	if  invoice_no10!="":
		db.invoice_no10 		= invoice_no10
	else:
		db.invoice_no10 		= ""
	
	if client_name10!="":  
		db.client_name10 		= client_name10
	else:
		db.client_name10 		= ""
	
	if principle10!='':  
		db.principle10 			= principle10
	else:
		db.principle10 			= 0.0
	
	if igst10!='': 
		db.igst10 				= igst10
	else:
		db.igst10 				= ""
	
	if igst_value10!='': 
		db.igst_value10 		= igst_value10
	else:
		db.igst_value10 		= 0.0
	
	if  tds10!='':
		db.tds10 				= tds10
	else:
		db.tds10 				= 0.0
	
	if tds_amount10!='': 
		db.tds_amount10 		= tds_amount10
	else:
		db.tds_amount10 		= 0.0
	
	if tds_igst10!='':   
		db.tds_igst10 			= tds_igst10
	else:
		db.tds_igst10 			= 0.0

	if tds_igst_amt10!='': 
		db.tds_igst_amt10 		= tds_igst_amt10
	else:
		db.tds_igst_amt10 		= 0.0
	
	if principle_tds_ded10!='':  
		db.principle_tds_ded10 	= principle_tds_ded10
	else:
		db.principle_tds_ded10  = 0.0

	if remitt10!='':
		db.remitt10 			= remitt10
	else:
		db.remitt10 			= ""
	
	db.entry_date 				= datetime.now().date()
	db.save()

	return HttpResponse(json.dumps('done'))