from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

#####################################################
import socket
import platform
import os

def load_dsr(request):
	if request.user.is_authenticated() and models.Users.objects.filter(user = request.user)[0].user_type == 'admin':
		login_user 			= request.user
		
		context={
			'login_user' : login_user,			
		}

		return render_to_response("invoice_display/online_dsr.html",context)
	else:
		return HttpResponseRedirect('/it/user_login')	


@csrf_exempt
def dsr_list(request):	
	single_client = []
	single_array  = []
	cl_array =[]
	s_no 		  = 0
	end_date      = json.loads(request.POST['end_date'])
	start_date    = json.loads(request.POST['start_date'])

	get_client = models.Client.objects.filter(proj_name='CHM',status=1)
	for w in get_client:
		if w.client_name not in single_client:
			single_client.append(w.client_name)



	ship_type       = 0
	stat_dt_format  = start_date
	end_dt_format   = end_date
	for c in sorted(single_client):	
		api_url 	    = "https://chm.bwesglobal.com/hb/get_chm_data/"
		#api_url 	    = "http://0.0.0.0:8004/hb/get_chm_data/"
		api_method      = "GET"
		parameters      = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': c,'ship_type':ship_type,'start_date':str(stat_dt_format),'end_date': str(end_dt_format)}
		response        = requests.get(api_url, params=parameters,verify=False)
		ch_array        = json.loads(response.content)
		ch_array        = ch_array['finance_array']
		
		for t in ch_array:
			#print '------------<>>>>',t['client_name']
			try:
				if t['can_staus']==1:
					g_cancel = 'Cancelled'				
				else:
					g_cancel = ''
			except:
				g_cancel = ''


			check_done = models.invoice.objects.filter(ship_name=t['ship_name'],voyage_no=t['voyage_no'],proj_name='CHM').count()
		
			if check_done==0:
				chk_generated = ''
			else:
				chk_generated = 'Made'
		
			

			
			try:
				start_dht = datetime.strptime(t['start_dht_date'], "%Y-%m-%d  00:00:00+00:00").strftime('%d-%b-%Y')
			except:
				start_dht = ''

			try:
				last_dht = datetime.strptime(t['last_dht_date'], "%Y-%m-%d 00:00:00+00:00").strftime('%d-%b-%Y')
			except:
				last_dht = ''


			try:
				disch_datess = datetime.strptime(t['disch_date'], "%Y-%m-%d").strftime('%d-%b-%Y')
			except:
				disch_datess = ''

			try:
				disch_port = t['disch_port']
			except:
				disch_port = ''

			if t['edit_reasons']!='None' and t['edit_reasons']!='NULL' and t['edit_reasons']!=None and  t['edit_reasons']!='':
				f_reasons = t['edit_reasons']
			else:
				f_reasons = ''

			single_array.append({
				'pic'				: t['vm_name'],
				'vessel' 			: t['ship_name'],
				'voyage_no' 		: t['voyage_no'],
				'nomination_date' 	: t['finance_date'],
				'client' 			: t['client_name'],
				'last_dht' 			: str(last_dht),
				'start_dht' 		: str(start_dht),
				'status'			: g_cancel,
				's_no'				: s_no+1,
				'generated' 		: chk_generated,
				'disch_datex'       : str(disch_datess),
				'disch_port'        : disch_port,
				'f_reasons' 		: f_reasons,


			})
			s_no+=1
			

	context={
	'single_array' : single_array
	}

		
		
	return HttpResponse(json.dumps(context))	
		