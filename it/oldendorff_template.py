from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words

import smbclient 
import os 
from shutil import copyfile    
import shutil
from smb import smb_structs
from smb.SMBConnection import SMBConnection
from nmb.NetBIOS import NetBIOS
smb_structs.SUPPORT_SMB2 = True
import urllib
import tempfile
import urllib2
from smb.SMBHandler import SMBHandler

#####################################################


def get_calling_for_excel_otp(client,calling_pdf,invoice,flag_no,total_row_table):	
	dir  = '/var/www/html/invoice/it/static/pdf/BIM/'+str(client)+'/Online/'
	if not os.path.exists(dir):
		os.makedirs(dir)

	vvds = calling_pdf[0].invoice_no
	split_invoice_no  = vvds.replace('/','_')
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/OLDENDORFF_.xlsx'	
	dstroot       = '/var/www/html/invoice/it/static/pdf/BIM/'+str(client)+'/Online/OLDENDORFF_'+str(split_invoice_no)+'.xlsx'	 
	copyfile(srcfile, dstroot)
	my_dir 	      = os.path.dirname(__file__)	
	location      = '/var/www/html/invoice/it/static/pdf/BIM/'+str(client)+'/Online/OLDENDORFF_'+str(split_invoice_no)+'.xlsx'	
	wb 	   	      = load_workbook('/var/www/html/invoice/it/static/pdf/BIM/'+str(client)+'/Online/OLDENDORFF_'+str(split_invoice_no)+'.xlsx')
	ws     	      = wb.get_sheet_by_name("Invoice")

	s_no = 1
	j  	 = 1	
	i    = 24
	t    = 31
	f    = 56
		
	_styles   		= Style(font=Font(name='Calibri', size=11, bold=True),fill=PatternFill(patternType='solid', fgColor=Color(rgb='DBE5F1')),border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')))
	_styles2  		= Style(font=Font(name='Arial', size=10, italic=True,bold=True,color=colors.RED),alignment=Alignment(horizontal='right'))		
	month_name  	= datetime.strptime(calling_pdf[0].month, "%Y-%m-%d").strftime('%b')
	year 			= datetime.strptime(calling_pdf[0].month, "%Y-%m-%d").strftime('%Y')
	date_name  		= datetime.strptime(calling_pdf[0].month, "%Y-%m-%d").strftime('%d')
	invoice_period  = str(month_name)+' '+str(year)
	try:
		today_monthe  	= calling_pdf[0].today.date().strftime('%b')
	except:
		today_monthe  	= datetime.now().date().strftime('%b')

	month_number        = datetime.strptime(str(calling_pdf[0].month), "%Y-%m-%d").strftime('%m')	
	month_no 			= 12 #int(month_number)-1
	if month_no<10:
		month_str = "0"+str(month_no)
	else:
		month_str = month_no

	get_month   		 	= monthrange(int(year),int(month_str))[1]	
	ws.cell('H5').value  	= calling_pdf[0].invoice_no
	ws.cell('H11').value 	= 'Invoice for period of '+str(invoice_period)
	generate_date 			= str(get_month)+'-'+str(month_name)+'-'+str(year)	
	ws.cell('H6').value  	= generate_date	
	ws.print_options.horizontalCentered = True
	net_total  = 0
	sub_total  = 0
	sum_of_amt = 0
	ws.cell('D10').value  = "Mr Rayjo Paul"

	img 		= Image('/var/www/html/invoice/it/static/pdf/1.png')
	ws.add_image(img, 'A1')
	img2 		= Image('/var/www/html/invoice/it/static/pdf/boss.png',size=[95,50])	
	ws.add_image(img2,'K1')
	img3 		= Image('/var/www/html/invoice/it/static/pdf/stamp.png',size=[190,194])	
	ws.add_image(img3,'I140')


	for x in calling_pdf:
		#if s_no<=30:
		thin_border  = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		my_style     = Style(font=Font(name='Calibri', size=9, bold=False),border=thin_border)
		_row_style   = Style(font=Font(name='Calibri', size=9, bold=False),border=thin_border)
		ws.cell('B'+str(i)).value = s_no			
		ws.cell('C'+str(i)).value = x.ship_name
		ws.cell('H'+str(i)).value = x.qty

		split_day = (x.client.price)/(get_month)
		if x.qty>29:
			billing = x.client.price
		else:
			billing = split_day*(x.qty)
		
		try:	
			ws.cell('I'+str(i)).value = x.client.price
		except:
			ws.cell('I'+str(i)).value = x.client.price
			
		try:
			ws.cell('J'+str(i)).value = billing
		except:
			ws.cell('J'+str(i)).value = (x.client.price * billing)
			
		try:
			net_total+=(billing)
		except:
			net_total+=(billing)
		#last_row = i
		
		#else:
		# if x.qty>29:
		# 	billing = 450
		# else:
		# 	billing = 15*(x.qty)	

		# ws.cell('B'+str(i)).value = s_no		
		# ws.cell('C'+str(i)).value = x.ship_name
		# ws.cell('H'+str(i)).value = x.qty
		# try:	
		# 	ws.cell('I'+str(i)).value = x.client.price
		# except:
		# 	ws.cell('I'+str(i)).value = x.client.price
			
		# try:
		# 	ws.cell('J'+str(i)).value = billing
		# except:
		# 	ws.cell('J'+str(i)).value = (x.client.price * billing)
			
		# try:
		# 	sub_total+=(billing)
		# except:
		# 	sub_total+=(billing)
		s_no+=1
		i+=1
		j+=1
		f+=1
		t+=1

		try:
			sum_of_amt+=(billing)
		except:
			sum_of_amt+=(billing)

	s_no+=1
	i=i+1
	j+=1


	# _border 		= Border(top=Side(style='thin'))
	# style_border  	= Style(font=Font(name='Arial', size=12, italic=False,bold=True,color=colors.BLACK),alignment=Alignment(horizontal='right'))
	# merge 			= ('C'+str(last_row+1)+':I'+str(last_row+1))
	# ws.merge_cells(merge)
	# ws.cell('C'+str(last_row+1)).value  = 'Sub Total (C/f)'	
	# subtotal = net_total 
	# merges = ('C'+str(last_row+2)+':I'+str(last_row+2))
	# ws.merge_cells(merges)
	# ws.cell('C'+str(last_row+2)).value  = 'Sub Total (B/f)'	
	# ws.cell('J'+str(last_row+1)).value  = subtotal
	# ws.cell('J'+str(last_row+2)).value  = subtotal
	# ws.cell('B'+str(last_row+2)).value = ''
	# ws.cell('B'+str(last_row+1)).value = ''	
	ws.cell('J135').value  = "=SUM(J24"":"'J'+str(i)+")"
	rounded_amt = round(sum_of_amt,0)
	#############################################################
	word_amount 		 = convert_money_to_text(rounded_amt)
	ws.cell('B139').value = word_amount
	try:
		due_date  		 = calling_pdf[0].today + timedelta(days=60)
	except:
		due_date  		 = datetime.now().date() + timedelta(days=60)

	due_month 			 = (due_date).strftime('%b')
	due_day   			 = (due_date).strftime('%d')
	due_year   			 = (due_date).strftime('%Y')
	merge_due_date 		 = str(due_day)+'-'+str(due_month)+'-'+str(due_year)
	ws.cell('J140').value = merge_due_date

	target_line 	 = 135	
	extraction_point = int(i)+2
	for xxx in range(extraction_point, target_line):
		ws.row_dimensions[xxx].hidden = True	



	wb.save('/var/www/html/invoice/it/static/pdf/BIM/'+str(client)+'/Online/OLDENDORFF_'+str(split_invoice_no)+'.xlsx')	
	
	username 	= "BWTW059"
	password 	= "sandeep@123"
	clientname  = "OHMSERVER"
	servername  = "OHMSERVER"
	domain 		= 'WORKGROUP'
	ipaddress 	= "172.16.5.100"
	conn 		= SMBConnection(username,password,clientname,servername,domain,use_ntlm_v2=True, sign_options=2, is_direct_tcp=True)	
	conn.connect(ipaddress,445)
	Shares 		= conn.listShares()	
	share_name  = Shares[5].name	
	path_name   = 'Finance/Current/Finance/BIM/Online/'+str(client)+'/'		
	sharedfiles = conn.listPath(share_name,path_name)	
	
	with open('/var/www/html/invoice/it/static/pdf/BIM/'+str(client)+'/Online/OLDENDORFF_'+str(split_invoice_no)+'.xlsx', 'rb') as file:		
		conn.storeFile('Finance',"Finance/Current/Finance/BIM/Online/"+str(client)+"/Online/OLDENDORFF_"+str(split_invoice_no)+'.xlsx', file)	
	conn.close()	
	

	return HttpResponse(json.dumps('done'))


#def oldendorff_template_download(request):
	# import os, tempfile, zipfile
	# from django.core.servers.basehttp import FileWrapper
	# from django.conf import settings
	# import mimetypes	
	# my_dir 	 	   					= os.path.dirname(__file__)	
	# location 	   					= (my_dir+'/Shell/shell_template.xlsx')
	# filename 	   					= my_dir+'/Shell/shell_template.xlsx'
	# download_name  					= "shell_template.xlsx"	
	# wrapper        					= FileWrapper(open(filename))	
	# content_type   					= mimetypes.guess_type(filename)[0]
	# response       					= HttpResponse(wrapper,content_type=content_type)
	# response['Content-Length']      = os.path.getsize(filename)
	# response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	# return response
	#return HttpResponse(json.dumps('done'))


def convert_money_to_text(money):
    try:
        money_number = int(money)
    except ValueError:
        return "You must enter Number"
    if money_number == 0:
        return None
    positions 		= [None for i in range(4)]
    key_range 		= 0
    one_place 		= ["", "One ", "Two ", "Three ", "Four ","Five ", "Six ", "Seven ", "Eight ", "Nine "]
    one_ten_place   = ["Ten ", "Eleven ", "Twelve ", "Thirteen ", "Fourteen ","Fifteen ", "Sixteen ", "Seventeen ", "Eighteen ","Nineteen "]
    ten_place 		= ["Twenty ", "Thirty ", "Forty ", "Fifty ","Sixty ", "Seventy ", "Eighty ", "Ninety "]
    name_of_number  = ["Thousand ", "Lakh ", "Crore "]
    money_number_money_text = ''
    positions[0] = money_number % 1000  # units
    positions[1] = money_number / 1000
    positions[2] = money_number / 100000
    positions[1] = positions[1] - 100 * positions[2]  # thousands
    positions[3] = money_number / 10000000  # crores
    positions[2] = positions[2] - 100 * positions[3]  # lakhs
    for counter in range(3, 0, -1):
        if positions[counter] != 0:
            key_range = counter
            break
    for i in range(key_range, -1, -1):
        if positions[i] == 0:
            continue
        ones = positions[i] % 10  # ones
        tens = positions[i] / 10
        hundreds = positions[i] / 100  # hundreds
        tens = tens - 10 * hundreds  # tens
        if (hundreds > 0):
            money_number_money_text += one_place[hundreds] + "Hundred "
        if (ones > 0 or tens > 0):
            if (hundreds > 0 and i == 0):
                money_number_money_text += "and "
            if (tens == 0):
                money_number_money_text += one_place[ones]
            elif (tens == 1):
                money_number_money_text += one_ten_place[ones]
            else:
                money_number_money_text += ten_place[tens - 2] + \
                    one_place[ones]
        if (i != 0):
            money_number_money_text += name_of_number[i - 1]
    return money_number_money_text