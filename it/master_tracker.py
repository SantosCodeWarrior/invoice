from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
###################################################################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from openpyxl.styles import Border
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt


from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
import pytz

from openpyxl.styles.borders import Border, Side, BORDER_THIN
#####################################################


def master_tracker(request):
	if request.user.is_authenticated():
		login_user 		= request.user		
		app_url 		= request.path		
		split_url 		= app_url.split('/')
		url_nme 		= split_url[2].capitalize()
		IST 			= pytz.timezone('Asia/Kolkata') 
		datetime_ist 	= datetime.now(IST)		 
		now_date 		= datetime_ist.strftime('%Y-%m-%d %H:%M:%S')
		InID 			= models.log_sessions()
		InID.date  		= now_date 
		InID.user_name  = login_user
		InID.url_name   = url_nme
		InID.save()	

		context={
			'login_user' :  login_user
		}			
		return render_to_response("invoice_display/master_tracker.html",context)
		if user.is_anonymous():
			return HttpResponseRedirect('/it/user_login')
	else:
		return HttpResponseRedirect('/it/user_login')


def filter_master_tracker(request):
	start_dates 	= request.GET['start_dates']
	end_dates 		= request.GET['end_dates']
	api_url 	    = "https://chm.bwesglobal.com/hb/get_chm_data/"	
	api_method      = "GET"
	parameters      = {'key' : "938e67ec-e20f-408b-9bd5-5913bfdc1d7b",'name': '','ship_type':'','start_date':str(start_dates),'end_date': str(end_dates)}
	response        = requests.get(api_url, params=parameters,verify=False)
	chmData         = json.loads(response.content)
	vg_list         = chmData['vg_list']
	chm_list 		= []	
	for x in vg_list:
		ship_name = x['ship_name']
		voyage_no = x['voyage_no']
		try:
			qry    			= "SELECT invoice_no FROM invoice.it_invoice where ship_name='"+str(ship_name)+"' and voyage_no='"+str(voyage_no)+"'"
			cursor 			= connection.cursor()
			cursor.execute(qry)
			invoice_details = cursor.fetchall()					
			invoice_nos		= "Done (" +str(invoice_details[0][0])+")"			
		except:
			invoice_nos     = ""

				
		chm_list.append({
			'status'			: x['status'],
			'end_dht'  			: str(x['end_dht_date']),
			'vm_name'			: x['vm_name'],
			'tag_done' 			: invoice_nos, #x['tag_done'],
			'voyage_no'     	: x['voyage_no'],
			'start_dht' 		: str(x['start_dht_date']),
			'pool_name' 		: x['pool_name'],
			'ship_name'			: x['ship_name'],
			'voyage_id' 		: x['voyage_id'],	
			'disch_port'    	: x['discharge_port'],
			'get_resons'		: x['get_resons'],
			'client_name' 		: x['client_name'],
			'nominate_date' 	: str(x['nomination_date']),
			'max_disch_date'    : str(x['max_discharge_date']),
			'pv_complete_date'	: str(x['pv_complete_date']),
		})

	context={
		'chm_list' :  chm_list,	
	}

	return HttpResponse(json.dumps(context))

@csrf_exempt
def export_master_trackerss(request):
	data 				= json.loads(request.POST['datas'])	
	today 				= datetime.now().date()
	format_year 		= (today).strftime('%Y')
	format_month 		= (today).strftime('%b')
	int_mon				= (today).strftime('%m')
	format_date 		= (today).strftime('%d')
	int_mont 			= int(int_mon)
	mrge_date 			= str(format_date)+'-'+str(format_month)+'-'+str(format_year)
	select_dir    		= os.path.dirname(__file__)
	srcfile       		= select_dir+'/static/Master_Tracker.xlsx'
	dstroot       		= select_dir+'/Master_Tracker.xlsx'	 
	copyfile(srcfile,dstroot)
	my_dir 	      		= os.path.dirname(__file__)		
	location      		= (my_dir+'/Master_Tracker.xlsx')
	wb 	   	      		= load_workbook(my_dir+'/Master_Tracker.xlsx')
	ws_usd        		= wb.get_sheet_by_name("Sheet1")
	thin_border   		= Border(left=Side(style='thin',border_style=BORDER_THIN, color='808080'),right=Side(style='thin',border_style=BORDER_THIN, color='808080'),top=Side(style='thin',border_style=BORDER_THIN, color='808080'),bottom=Side(style='thin',border_style=BORDER_THIN, color='808080'))
	#thin_border 		= Border(left=Side(border_style=BORDER_THIN, color='808080'),right=Side(border_style=BORDER_THIN, color='808080'),top=Side(border_style=BORDER_THIN, color='808080'),bottom=Side(border_style=BORDER_THIN, color='808080'))
	my_style      		= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	thin_border   		= Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      		= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	_row_style    		= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	row_alignment		= Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border,alignment=Alignment(horizontal='center'))	
	ii   		  		= 2
	jj 	 		  		= 1
	
	for f in data:
		ship_name 			= f['ship_name']
		voyage_no 			= f['voyage_no']
		vm_name   			= f['vm_name']
		pool_name 			= f['pool_name']
		cl_name   			= f['client_name']
		nomination_date 	= f['nominate_date']
		discharging_port 	= f['disch_port']
		etd_disch_date	 	= f['max_disch_date']
		pv_complete_date    = f['pv_complete_date']
		status 				= f['status']
		start_dht_date 		= f['start_dht']
		end_dht_date 		= f['end_dht']
		done_status 		= f['get_resons']		
		voyageID 			= f['voyage_id']
		f_status 			= f['tag_done']

		ws_usd.cell('A'+str(ii)).value = jj
		ws_usd.cell('B'+str(ii)).value = ship_name
		ws_usd.cell('C'+str(ii)).value = voyage_no
		ws_usd.cell('D'+str(ii)).value = vm_name
		ws_usd.cell('E'+str(ii)).value = pool_name
		ws_usd.cell('F'+str(ii)).value = cl_name
		ws_usd.cell('G'+str(ii)).value = nomination_date
		ws_usd.cell('H'+str(ii)).value = discharging_port
		ws_usd.cell('I'+str(ii)).value = etd_disch_date
		ws_usd.cell('J'+str(ii)).value = pv_complete_date
		ws_usd.cell('K'+str(ii)).value = status
		ws_usd.cell('L'+str(ii)).value = start_dht_date
		ws_usd.cell('M'+str(ii)).value = end_dht_date
		ws_usd.cell('N'+str(ii)).value = done_status
		ws_usd.cell('O'+str(ii)).value = voyageID
		ws_usd.cell('P'+str(ii)).value = f_status

		ws_usd.cell('A'+str(ii)).style = row_alignment	
		ws_usd.cell('B'+str(ii)).style = _row_style	
		ws_usd.cell('C'+str(ii)).style = _row_style	
		ws_usd.cell('D'+str(ii)).style = _row_style	
		ws_usd.cell('E'+str(ii)).style = _row_style	
		ws_usd.cell('F'+str(ii)).style = _row_style	
		ws_usd.cell('G'+str(ii)).style = _row_style	
		ws_usd.cell('H'+str(ii)).style = _row_style	
		ws_usd.cell('I'+str(ii)).style = _row_style	
		ws_usd.cell('J'+str(ii)).style = _row_style	
		ws_usd.cell('K'+str(ii)).style = _row_style	
		ws_usd.cell('L'+str(ii)).style = _row_style	
		ws_usd.cell('M'+str(ii)).style = _row_style
		ws_usd.cell('N'+str(ii)).style = _row_style
		ws_usd.cell('O'+str(ii)).style = _row_style
		ws_usd.cell('P'+str(ii)).style = _row_style
		wb.save(my_dir+'/Master_Tracker.xlsx')
		ii+=1
		jj+=1
				
	return HttpResponse(json.dumps('done'))


def xexport_tracker_download(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes	
	today 							= datetime.now().date()
	format_year 					= (today).strftime('%Y')
	format_month 					= (today).strftime('%b')
	int_mon							= (today).strftime('%m')
	format_date 					= (today).strftime('%d')
	int_mont 						= int(int_mon)
	mrge_date 						= str(format_date)+'-'+str(format_month)+'-'+str(format_year)	
	my_dir 	 	   					= os.path.dirname(__file__)	
	location 	   					= (my_dir+'/Master_Tracker.xlsx')
	filename 	   					= my_dir+'/Master_Tracker.xlsx'
	download_name  					= 'Master Tracker.xlsx'	
	wrapper        					= FileWrapper(open(filename))	
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name	
	return response