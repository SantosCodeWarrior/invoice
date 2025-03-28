from django.shortcuts import render
from django.shortcuts import render,render_to_response
from it import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
#####################################################


def view_rec_inr(request):
	if request.user.is_authenticated():
		login_user = request.user	
		

		context={		
			'login_user'   : login_user,		
		}
		return render_to_response('invoice_display/view_rec_inr.html',context)
	else:
		return HttpResponseRedirect('/it/user_login')

def filter_rec_inr(request):
	calling_inr = models.remittance_data_inr.objects.all().order_by('remittance_date')
	inr_list 	= []
	for x in calling_inr:
		
		if x.remarks!=None:
			rem = x.remarks
		else:
			rem = ''

		try:	
			remittance_date = x.remittance_date.strftime('%d-%b-%Y')
		except:
			remittance_date = ''

		try:
			sum_of_remitt_amount = x.remitt_amount + x.remitt_amount2 + x.remitt_amount3 + x.remitt_amount4 + x.remitt_amount5 + x.remitt_amount6 + x.remitt_amount7 + x.remitt_amount8 + x.remitt_amount9 + x.remitt_amount10
			net_remitt_amt 		 = round(x.total_remitt,1)
		except:
			net_remitt_amt 		 = 0.0

		diff = x.amount_received - x.total_remitt

		if (abs(diff))<=10:
			status = 'RECONCILED'
			diff_color = ''
		else:
			status = 'NOT RECONCILED'
			diff_color = '#fa8c8c'

		
		# if x.amount_received == round(x.total_remitt,0):
		# 	status = 'RECONCILED'
		# 	diff_color = ''
		# else:
		# 	status = 'NOT RECONCILED'
		# 	diff_color = '#fa8c8c'

		



		inr_list.append({
			'remittance_date'			: str(remittance_date),
			'fremittance_date'			: str(x.remittance_date),
			'reference_no'				: x.reference_no,
			'amount_received'			: x.amount_received,
			'status'					: status,
			'total_remitt'				: net_remitt_amt, #x.total_remitt,
			'charges_any'				: x.charges_any,
			'remitt_amount'				: x.remitt_amount,
			'invoice_no1'				: x.invoice_no1,
			'client_name1'				: x.client_name1,
			'principle1'				: x.principle1,
			'igst1'						: x.igst1,
			'igst_value1'				: x.igst_value1,
			'tds1'						: x.tds1,
			'tds_amount1'				: x.tds_amount1,
			'tds_igst1'					: x.tds_igst1,
			'tds_igst_amt1'				: x.tds_igst_amt1,
			'principle_tds_ded1'		: x.principle_tds_ded1,
			'remitt1'					: x.remitt1,
			'invoice_no2'				: x.invoice_no2,
			'client_name2'				: x.client_name2,
			'principle2'				: x.principle2,
			'igst2'						: x.igst2,
			'igst_value2'				: x.igst_value2,
			'tds2'						: x.tds2,
			'tds_amount2'				: x.tds_amount2,
			'tds_igst2'					: x.tds_igst2,
			'tds_igst_amt2'				: x.tds_igst_amt2,
			'principle_tds_ded2'		: x.principle_tds_ded2,
			'remitt2'					: x.remitt2,
			'invoice_no3'				: x.invoice_no3,
			'client_name3'				: x.client_name3,
			'principle3'				: x.principle3,
			'igst3'						: x.igst3,
			'igst_value3'				: x.igst_value3,
			'tds3'						: x.tds3,
			'tds_amount3'				: x.tds_amount3,
			'tds_igst3'					: x.tds_igst3,
			'tds_igst_amt3'				: x.tds_igst_amt3,
			'principle_tds_ded3'		: x.principle_tds_ded3,
			'remitt3'					: x.remitt3,
			'invoice_no4'				: x.invoice_no4,
			'client_name4'				: x.client_name4,
			'principle4'				: x.principle4,
			'igst4'						: x.igst4,
			'igst_value4'				: x.igst_value4,
			'tds4'						: x.tds4,
			'tds_amount4'				: x.tds_amount4,
			'tds_igst4'					: x.tds_igst4,
			'tds_igst_amt4'				: x.tds_igst_amt4,
			'principle_tds_ded4'		: x.principle_tds_ded4,
			'remitt4'					: x.remitt4,
			'invoice_no5'				: x.invoice_no5,
			'client_name5'				: x.client_name5,
			'principle5'				: x.principle5,
			'igst5'						: x.igst5,
			'igst_value5'				: x.igst_value5,
			'tds5'						: x.tds5,
			'tds_amount5'				: x.tds_amount5,
			'tds_igst5'					: x.tds_igst5,
			'tds_igst_amt5'				: x.tds_igst_amt5,
			'principle_tds_ded5'		: x.principle_tds_ded5,
			'remitt5'					: x.remitt5,
			'invoice_no6'				: x.invoice_no6,
			'client_name6'				: x.client_name6,
			'principle6'				: x.principle6,
			'igst6'						: x.igst6,
			'igst_value6'				: x.igst_value6,
			'tds6'						: x.tds6,
			'tds_amount6'				: x.tds_amount6,
			'tds_igst6'					: x.tds_igst6,
			'tds_igst_amt6'				: x.tds_igst_amt6,
			'principle_tds_ded6'		: x.principle_tds_ded6,
			'remitt6'					: x.remitt6,
			'invoice_no7'				: x.invoice_no7,
			'client_name7'				: x.client_name7,
			'principle7'				: x.principle7,
			'igst7'						: x.igst7,
			'igst_value7'				: x.igst_value7,
			'tds7'						: x.tds7,
			'tds_amount7'				: x.tds_amount7,
			'tds_igst7'					: x.tds_igst7,
			'tds_igst_amt7'				: x.tds_igst_amt7,
			'principle_tds_ded7'		: x.principle_tds_ded7,
			'remitt7'					: x.remitt7,
			'invoice_no8'				: x.invoice_no8,
			'client_name8'				: x.client_name8,
			'principle8'				: x.principle8,
			'igst8'						: x.igst8,
			'igst_value8'				: x.igst_value8,
			'tds8'						: x.tds8,
			'tds_amount8'				: x.tds_amount8,
			'tds_igst8'					: x.tds_igst8,
			'tds_igst_amt8'				: x.tds_igst_amt8,
			'principle_tds_ded8'		: x.principle_tds_ded8,
			'remitt8'					: x.remitt8,
			'invoice_no9'				: x.invoice_no9,
			'client_name9'				: x.client_name9,
			'principle9'				: x.principle9,
			'igst9'						: x.igst9,
			'igst_value9'				: x.igst_value9,
			'tds9'						: x.tds9,
			'tds_amount9'				: x.tds_amount9,
			'tds_igst9'					: x.tds_igst9,
			'tds_igst_amt9'				: x.tds_igst_amt9,
			'principle_tds_ded9'		: x.principle_tds_ded9,
			'remitt9'					: x.remitt9,
			'invoice_no10'				: x.invoice_no10,
			'client_name10'				: x.client_name10,
			'principle10'				: x.principle10,
			'igst10'					: x.igst10,
			'igst_value10'				: x.igst_value10,
			'tds10'						: x.tds10,
			'tds_amount10'				: x.tds_amount10,
			'tds_igst10'				: x.tds_igst10,
			'tds_igst_amt10'			: x.tds_igst_amt10,
			'principle_tds_ded10'		: x.principle_tds_ded10,
			'remitt10'					: x.remitt10,
			'remarks'					: rem,
			'entry_date'				: str(x.entry_date),
			'ids'						: x.id,
			'remitt_amount'				: x.remitt_amount,
			'remitt_amount2'			: x.remitt_amount2,
			'remitt_amount3'			: x.remitt_amount3,
			'remitt_amount4'			: x.remitt_amount4,
			'remitt_amount5'			: x.remitt_amount5,
			'remitt_amount6'			: x.remitt_amount6,
			'remitt_amount7'			: x.remitt_amount7,
			'remitt_amount8'			: x.remitt_amount8,
			'remitt_amount9'			: x.remitt_amount9,
			'remitt_amount10'			: x.remitt_amount10,

			'diff_color'				: diff_color,
			})

	context={
	'inr_list' : inr_list
	}

	return HttpResponse(json.dumps(context))



def update_remiitance_inrs(request):
	e_get_id 					= json.loads(request.GET['e_get_id'])
	remittance_date 			= json.loads(request.GET['remittance_date'])
	reference_no 				= json.loads(request.GET['reference_no'])
	amount_received 			= json.loads(request.GET['amount_received'])
	status 						= json.loads(request.GET['status'])
	total_remitt 				= json.loads(request.GET['total_remitt'])
	charges_any 				= json.loads(request.GET['charges_any'])
	remitt_amount1 				= json.loads(request.GET['remitt_amount1'])
	invoice_no1 				= json.loads(request.GET['invoice_no1'])
	client_name1 				= json.loads(request.GET['client_name1'])
	e_principle1 				= json.loads(request.GET['e_principle1'])
	e_igst1 					= json.loads(request.GET['e_igst1'])
	e_igst_value1 				= json.loads(request.GET['e_igst_value1'])
	e_tds1 						= json.loads(request.GET['e_tds1'])
	e_tds_amount1 				= json.loads(request.GET['e_tds_amount1'])
	e_tds_igst1 				= json.loads(request.GET['e_tds_igst1'])
	e_tds_igst_amt1 			= json.loads(request.GET['e_tds_igst_amt1'])
	e_principle_tds_ded1 		= json.loads(request.GET['e_principle_tds_ded1'])
	e_remitt1 					= json.loads(request.GET['e_remitt1'])
	e_remitt_amount2 			= json.loads(request.GET['e_remitt_amount2'])
	e_invoice_no2 				= json.loads(request.GET['e_invoice_no2'])
	e_client_name2 				= json.loads(request.GET['e_client_name2'])
	e_principle2 				= json.loads(request.GET['e_principle2'])
	e_igst2 					= json.loads(request.GET['e_igst2'])
	e_igst_value2 				= json.loads(request.GET['e_igst_value2'])
	e_tds2 						= json.loads(request.GET['e_tds2'])
	e_tds_amount2 				= json.loads(request.GET['e_tds_amount2'])
	e_tds_igst2 				= json.loads(request.GET['e_tds_igst2'])
	e_tds_igst_amt2 			= json.loads(request.GET['e_tds_igst_amt2'])
	e_principle_tds_ded2 		= json.loads(request.GET['e_principle_tds_ded2'])
	e_remitt2 					= json.loads(request.GET['e_remitt2'])
	e_remitt_amount3 			= json.loads(request.GET['e_remitt_amount3'])
	e_invoice_no3 				= json.loads(request.GET['e_invoice_no3'])
	e_client_name3 				= json.loads(request.GET['e_client_name3'])
	e_principle3 				= json.loads(request.GET['e_principle3'])
	e_igst3 					= json.loads(request.GET['e_igst3'])
	e_igst_value3 				= json.loads(request.GET['e_igst_value3'])
	e_tds3 						= json.loads(request.GET['e_tds3'])
	e_tds_amount3 				= json.loads(request.GET['e_tds_amount3'])
	e_tds_igst3 				= json.loads(request.GET['e_tds_igst3'])
	e_tds_igst_amt3 			= json.loads(request.GET['e_tds_igst_amt3'])
	e_principle_tds_ded3 		= json.loads(request.GET['e_principle_tds_ded3'])
	e_remitt3 					= json.loads(request.GET['e_remitt3'])
	e_remitt_amount4 			= json.loads(request.GET['e_remitt_amount4'])
	e_invoice_no4 				= json.loads(request.GET['e_invoice_no4'])
	e_client_name4 				= json.loads(request.GET['e_client_name4'])
	e_principle4 				= json.loads(request.GET['e_principle4'])
	e_igst4 					= json.loads(request.GET['e_igst4'])
	e_igst_value4 				= json.loads(request.GET['e_igst_value4'])
	e_tds4 						= json.loads(request.GET['e_tds4'])
	e_tds_amount4 				= json.loads(request.GET['e_tds_amount4'])
	e_tds_igst4 				= json.loads(request.GET['e_tds_igst4'])
	e_tds_igst_amt4 			= json.loads(request.GET['e_tds_igst_amt4'])
	e_principle_tds_ded4 		= json.loads(request.GET['e_principle_tds_ded4'])
	e_remitt4 					= json.loads(request.GET['e_remitt4'])
	e_remitt_amount5 			= json.loads(request.GET['e_remitt_amount5'])
	e_invoice_no5 				= json.loads(request.GET['e_invoice_no5'])
	e_client_name5 				= json.loads(request.GET['e_client_name5'])
	e_principle5 				= json.loads(request.GET['e_principle5'])
	e_igst5 					= json.loads(request.GET['e_igst5'])
	e_igst_value5 				= json.loads(request.GET['e_igst_value5'])
	e_tds5 						= json.loads(request.GET['e_tds5'])
	e_tds_amount5 				= json.loads(request.GET['e_tds_amount5'])
	e_tds_igst5 				= json.loads(request.GET['e_tds_igst5'])
	e_tds_igst_amt5 			= json.loads(request.GET['e_tds_igst_amt5'])
	e_principle_tds_ded5 		= json.loads(request.GET['e_principle_tds_ded5'])
	e_remitt5 					= json.loads(request.GET['e_remitt5'])
	e_remitt_amount6 			= json.loads(request.GET['e_remitt_amount6'])
	e_invoice_no6 				= json.loads(request.GET['e_invoice_no6'])
	e_client_name6 				= json.loads(request.GET['e_client_name6'])
	e_principle6 				= json.loads(request.GET['e_principle6'])
	e_igst6 					= json.loads(request.GET['e_igst6'])
	e_igst_value6 				= json.loads(request.GET['e_igst_value6'])
	e_tds6 						= json.loads(request.GET['e_tds6'])
	e_tds_amount6 				= json.loads(request.GET['e_tds_amount6'])
	e_tds_igst6 				= json.loads(request.GET['e_tds_igst6'])
	e_tds_igst_amt6 			= json.loads(request.GET['e_tds_igst_amt6'])
	e_principle_tds_ded6 		= json.loads(request.GET['e_principle_tds_ded6'])
	e_remitt6 					= json.loads(request.GET['e_remitt6'])
	e_remitt_amount7 			= json.loads(request.GET['e_remitt_amount7'])
	e_invoice_no7 				= json.loads(request.GET['e_invoice_no7'])
	e_client_name7 				= json.loads(request.GET['e_client_name7'])
	e_principle7 				= json.loads(request.GET['e_principle7'])
	e_igst7 					= json.loads(request.GET['e_igst7'])
	e_igst_value7 				= json.loads(request.GET['e_igst_value7'])
	e_tds7 						= json.loads(request.GET['e_tds7'])
	e_tds_amount7 				= json.loads(request.GET['e_tds_amount7'])
	e_tds_igst7 				= json.loads(request.GET['e_tds_igst7'])
	e_tds_igst_amt7 			= json.loads(request.GET['e_tds_igst_amt7'])
	e_principle_tds_ded7 		= json.loads(request.GET['e_principle_tds_ded7'])
	e_remitt7 					= json.loads(request.GET['e_remitt7'])
	e_remitt_amount8 			= json.loads(request.GET['e_remitt_amount8'])
	e_invoice_no8 				= json.loads(request.GET['e_invoice_no8'])
	e_client_name8 				= json.loads(request.GET['e_client_name8'])
	e_principle8 				= json.loads(request.GET['e_principle8'])
	e_igst8 					= json.loads(request.GET['e_igst8'])
	e_igst_value8 				= json.loads(request.GET['e_igst_value8'])
	e_tds8 						= json.loads(request.GET['e_tds8'])
	e_tds_amount8 				= json.loads(request.GET['e_tds_amount8'])
	e_tds_igst8 				= json.loads(request.GET['e_tds_igst8'])
	e_tds_igst_amt8 			= json.loads(request.GET['e_tds_igst_amt8'])
	e_principle_tds_ded8 		= json.loads(request.GET['e_principle_tds_ded8'])
	e_remitt8 					= json.loads(request.GET['e_remitt8'])
	e_remitt_amount9 			= json.loads(request.GET['e_remitt_amount9'])
	e_invoice_no9 				= json.loads(request.GET['e_invoice_no9'])
	e_client_name9 				= json.loads(request.GET['e_client_name9'])
	e_principle9 				= json.loads(request.GET['e_principle9'])
	e_igst9 					= json.loads(request.GET['e_igst9'])
	e_igst_value9 				= json.loads(request.GET['e_igst_value9'])
	e_tds9 						= json.loads(request.GET['e_tds9'])
	e_tds_amount9 				= json.loads(request.GET['e_tds_amount9'])
	e_tds_igst9 				= json.loads(request.GET['e_tds_igst9'])
	e_tds_igst_amt9 			= json.loads(request.GET['e_tds_igst_amt9'])
	e_principle_tds_ded9 		= json.loads(request.GET['e_principle_tds_ded9'])
	e_remitt9 					= json.loads(request.GET['e_remitt9'])
	e_remitt_amount10 			= json.loads(request.GET['e_remitt_amount10'])
	e_invoice_no10 				= json.loads(request.GET['e_invoice_no10'])
	e_client_name10 			= json.loads(request.GET['e_client_name10'])
	e_principle10 				= json.loads(request.GET['e_principle10'])
	e_igst10 					= json.loads(request.GET['e_igst10'])
	e_igst_value10 				= json.loads(request.GET['e_igst_value10'])
	e_tds10 					= json.loads(request.GET['e_tds10'])
	e_tds_amount10 				= json.loads(request.GET['e_tds_amount10'])
	e_tds_igst10 				= json.loads(request.GET['e_tds_igst10'])
	e_tds_igst_amt10 			= json.loads(request.GET['e_tds_igst_amt10'])
	e_principle_tds_ded10 		= json.loads(request.GET['e_principle_tds_ded10'])
	e_remitt10 					= json.loads(request.GET['e_remitt10'])
	

	CHK = models.remittance_data_inr.objects.filter(id=e_get_id).count()
	if CHK!=0:
		UpD = models.remittance_data_inr.objects.filter(id=e_get_id).first()
	else:
		UpD = models.remittance_data_inr()

	UpD.id 						= e_get_id
	UpD.remittance_date 		= remittance_date 		
	UpD.reference_no 			= reference_no 		
	UpD.amount_received 		= amount_received 		
	UpD.status 					= status 		
	UpD.total_remitt 			= total_remitt 		
	UpD.charges_any 			= charges_any 		
	UpD.remitt_amount1 			= remitt_amount1 		
	UpD.invoice_no1 			= invoice_no1 		
	UpD.client_name1 			= client_name1 		
	UpD.principle1 				= e_principle1 		
	UpD.igst1 					= e_igst1 		
	UpD.igst_value1 			= e_igst_value1 		
	UpD.tds1 					= e_tds1 		
	UpD.tds_amount1 			= e_tds_amount1 		
	UpD.tds_igst1 				= e_tds_igst1 		
	UpD.tds_igst_amt1 			= e_tds_igst_amt1 		
	UpD.principle_tds_ded1 		= e_principle_tds_ded1 		
	UpD.remitt1 				= e_remitt1 		
	UpD.remitt_amount2 			= e_remitt_amount2 		
	UpD.invoice_no2 			= e_invoice_no2 		
	UpD.client_name2 			= e_client_name2 		
	UpD.principle2 				= e_principle2 		
	UpD.igst2 					= e_igst2 		
	UpD.igst_value2 			= e_igst_value2 		
	UpD.tds2 					= e_tds2 		
	UpD.tds_amount2 			= e_tds_amount2 		
	UpD.tds_igst2 				= e_tds_igst2 		
	UpD.tds_igst_amt2 			= e_tds_igst_amt2 		
	UpD.principle_tds_ded2 		= e_principle_tds_ded2 		
	UpD.remitt2 				= e_remitt2 		
	UpD.remitt_amount3 			= e_remitt_amount3 		
	UpD.invoice_no3 			= e_invoice_no3 		
	UpD.client_name3 			= e_client_name3 		
	UpD.principle3 				= e_principle3 		
	UpD.igst3 					= e_igst3 		
	UpD.igst_value3 			= e_igst_value3 		
	UpD.tds3 					= e_tds3 		
	UpD.tds_amount3 			= e_tds_amount3 		
	UpD.tds_igst3 				= e_tds_igst3 		
	UpD.tds_igst_amt3 			= e_tds_igst_amt3 		
	UpD.principle_tds_ded3 		= e_principle_tds_ded3 		
	UpD.remitt3 				= e_remitt3 		
	UpD.remitt_amount4 			= e_remitt_amount4 		
	UpD.invoice_no4 			= e_invoice_no4 		
	UpD.client_name4 			= e_client_name4 		
	UpD.principle4 				= e_principle4 		
	UpD.igst4 					= e_igst4 		
	UpD.igst_value4 			= e_igst_value4 		
	UpD.tds4 					= e_tds4 		
	UpD.tds_amount4 			= e_tds_amount4 		
	UpD.tds_igst4 				= e_tds_igst4 		
	UpD.tds_igst_amt4 			= e_tds_igst_amt4 		
	UpD.principle_tds_ded4 		= e_principle_tds_ded4 		
	UpD.remitt4 				= e_remitt4 		
	UpD.remitt_amount5 			= e_remitt_amount5 		
	UpD.invoice_no5 			= e_invoice_no5 		
	UpD.client_name5 			= e_client_name5 		
	UpD.principle5 				= e_principle5 		
	UpD.igst5 					= e_igst5 		
	UpD.igst_value5 			= e_igst_value5 		
	UpD.tds5 					= e_tds5 		
	UpD.tds_amount5 			= e_tds_amount5 		
	UpD.tds_igst5 				= e_tds_igst5 		
	UpD.tds_igst_amt5 			= e_tds_igst_amt5 		
	UpD.principle_tds_ded5 		= e_principle_tds_ded5 		
	UpD.remitt5 				= e_remitt5 		
	UpD.remitt_amount6 			= e_remitt_amount6 		
	UpD.invoice_no6 			= e_invoice_no6 		
	UpD.client_name6 			= e_client_name6 		
	UpD.principle6 				= e_principle6 		
	UpD.igst6 					= e_igst6 		
	UpD.igst_value6 			= e_igst_value6 		
	UpD.tds6 					= e_tds6 		
	UpD.tds_amount6 			= e_tds_amount6 		
	UpD.tds_igst6 				= e_tds_igst6 		
	UpD.tds_igst_amt6 			= e_tds_igst_amt6 		
	UpD.principle_tds_ded6 		= e_principle_tds_ded6 		
	UpD.remitt6 				= e_remitt6 		
	UpD.remitt_amount7 			= e_remitt_amount7 		
	UpD.invoice_no7 			= e_invoice_no7 		
	UpD.client_name7 			= e_client_name7 		
	UpD.principle7 				= e_principle7 		
	UpD.igst7 					= e_igst7 		
	UpD.igst_value7 			= e_igst_value7 		
	UpD.tds7 					= e_tds7 		
	UpD.tds_amount7 			= e_tds_amount7 		
	UpD.tds_igst7 				= e_tds_igst7 		
	UpD.tds_igst_amt7 			= e_tds_igst_amt7 		
	UpD.principle_tds_ded7 		= e_principle_tds_ded7 		
	UpD.remitt7 				= e_remitt7 		
	UpD.remitt_amount8 			= e_remitt_amount8 		
	UpD.invoice_no8 			= e_invoice_no8 		
	UpD.client_name8 			= e_client_name8 		
	UpD.principle8 				= e_principle8 		
	UpD.igst8 					= e_igst8 		
	UpD.igst_value8 			= e_igst_value8 		
	UpD.tds8 					= e_tds8 		
	UpD.tds_amount8 			= e_tds_amount8 		
	UpD.tds_igst8 				= e_tds_igst8 		
	UpD.tds_igst_amt8 			= e_tds_igst_amt8 		
	UpD.principle_tds_ded8 		= e_principle_tds_ded8 		
	UpD.remitt8 				= e_remitt8 		
	UpD.remitt_amount9 			= e_remitt_amount9 		
	UpD.invoice_no9 			= e_invoice_no9 		
	UpD.client_name9 			= e_client_name9 		
	UpD.principle9 				= e_principle9 		
	UpD.igst9 					= e_igst9 		
	UpD.igst_value9 			= e_igst_value9 		
	UpD.tds9 					= e_tds9 		
	UpD.tds_amount9 			= e_tds_amount9 		
	UpD.tds_igst9 				= e_tds_igst9 		
	UpD.tds_igst_amt9 			= e_tds_igst_amt9 		
	UpD.principle_tds_ded9 		= e_principle_tds_ded9 		
	UpD.remitt9 				= e_remitt9 		
	UpD.remitt_amount10 		= e_remitt_amount10 		
	UpD.invoice_no10 			= e_invoice_no10 		
	UpD.client_name10 			= e_client_name10 		
	UpD.principle10 			= e_principle10 		
	UpD.igst10 					= e_igst10 		
	UpD.igst_value10 			= e_igst_value10 		
	UpD.tds10 					= e_tds10 		
	UpD.tds_amount10 			= e_tds_amount10 		
	UpD.tds_igst10 				= e_tds_igst10 		
	UpD.tds_igst_amt10 			= e_tds_igst_amt10 		
	UpD.principle_tds_ded10 	= e_principle_tds_ded10 		
	UpD.remitt10 				= e_remitt10

	UpD.save()		

	return HttpResponse(json.dumps('done'))


def calculation_remittance_inr(request):
	e_n 		= json.loads(request.GET['n'])	
	e_invoice 	= json.loads(request.GET['e_invoice'])
	
	amountss    = 0
	query1 		= "SELECT sum(ins.price* ins.qty*ins.rate),cl.client_name FROM invoice.it_invoice ins,invoice.it_client cl where (cl.id=ins.client_id) and ins.invoice_no='"+str(e_invoice)+"';"	
	cursor 		= connection.cursor()
	cursor.execute(query1)
	totamts 	= cursor.fetchall()
	calc_qry    = []
	try:
		amt = round(totamts[0][0],1)
	except:
		amt = 0.0
	calc_qry.append({
		'amountss' 		: amt,
		'client_name' 	: totamts[0][1],
		'sno' 			: int(e_n),
	})	
	
	return HttpResponse(json.dumps(calc_qry))

	

def get_pdf_viewerx_inr(request):	
	invc 	= request.GET['invc']
	urls 	= models.invoice.objects.filter(invoice_no=invc).first()
	url 	= urls.url
	print '-------',invc

	return HttpResponse(json.dumps(url))